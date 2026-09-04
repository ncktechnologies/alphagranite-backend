import asyncio
import calendar
import csv
import io
import logging
import os
from datetime import date, datetime, timedelta
from decimal import Decimal
from functools import partial
from typing import Any
from typing import Optional

from openpyxl import Workbook
from sqlalchemy import Numeric, case, null
from sqlalchemy import cast
from sqlalchemy import func
from sqlalchemy import select

from src.app.database.account import Account
from src.app.database.business_job import BusinessJob
from src.app.database.fab import Fab
from src.app.database.shop_cut_plan import ShopCutPlan
from src.app.database.stone_color import StoneColor
from src.app.database.stone_thickness import StoneThickness
from src.app.database.stone_type import StoneType
from src.app.interface.generated_schemas import CutList
from src.app.interface.generated_schemas import InstallCompletion
from src.app.service.background import send_email_with_attachments
from src.app.utils.config import ADMIN_EMAIL
from src.app.utils.config import SessionLocal

logger = logging.getLogger("monthly_status_report")

_scheduler_task: Optional[asyncio.Task] = None
_last_trigger_key: Optional[str] = None


DETAIL_COLUMNS = [
    "FabType",
    "FabID",
    "JobNumber",
    "Account",
    "JobName",
    "Areas",
    "StoneType",
    "StoneColor",
    "StoneThickness",
    "TotalSqFt",
    "CutSqFt",
    "CutPercent",
    "WJLinFt",
    "WJPercent",
    "EdgingLinFt",
    "EdgingPercent",
    "CNCLinFt",
    "CNCPercent",
    "MiterLinFt",
    "MiterPercent",
    "TouchupSqFt",
    "TouchupPercent",
    "EstCompletionDate",
    "PercentageComplete",
]

SUMMARY_COLUMNS = [
    "Account",
    "JobNumber",
    "JobName",
    "TotalSqFt_sum",
    "CompletedSqFt_sum",
    "Rows",
    "PercentComplete",
]


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _month_bounds(year: int, month: int) -> tuple[datetime, datetime]:
    start_dt = datetime(year, month, 1)
    if month == 12:
        end_dt = datetime(year + 1, 1, 1) - timedelta(microseconds=1)
    else:
        end_dt = datetime(year, month + 1, 1) - timedelta(microseconds=1)
    return start_dt, end_dt


def _to_percent(numerator: float, denominator: float) -> float:
    if denominator <= 0:
        return 0.0
    return round((numerator / denominator) * 100, 2)


def _normalize_work_percent(value: Any) -> float:
    """Normalize work percentage values to 0..100.

    Some sources store work percentage as 0..1 and others as 0..100.
    """
    pct = _to_float(value)
    if pct <= 0:
        return 0.0
    if pct <= 1:
        pct = pct * 100
    return round(min(pct, 100.0), 2)


def _rows_to_csv_bytes(columns: list[str], rows: list[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns)
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in columns})
    return buf.getvalue().encode("utf-8")


def _rows_to_xlsx_bytes(detail_rows: list[dict], summary_rows: list[dict]) -> bytes:
    wb = Workbook()

    ws_detail = wb.active
    ws_detail.title = "End of Month Status"
    for col_idx, col_name in enumerate(DETAIL_COLUMNS, start=1):
        ws_detail.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(detail_rows, start=2):
        for col_idx, col_name in enumerate(DETAIL_COLUMNS, start=1):
            ws_detail.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    ws_summary = wb.create_sheet(title="Account Job Summary")
    for col_idx, col_name in enumerate(SUMMARY_COLUMNS, start=1):
        ws_summary.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(summary_rows, start=2):
        for col_idx, col_name in enumerate(SUMMARY_COLUMNS, start=1):
            ws_summary.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def safe_cast_numeric(column):
    """Safely cast a column to numeric, returning null if not a valid number."""
    return case(
        (column.op("~")(r"^(?:[0-9]+(?:\.[0-9]*)?|\.[0-9]+)$"), cast(column, Numeric)),
        else_=null(),
    )


async def _build_report_rows(year: int, month: int) -> tuple[list[dict], list[dict]]:
    async with SessionLocal() as db:
        cut_sqft_subq = (
            select(
                CutList.fab_id.label("fab_id"),
                func.max(safe_cast_numeric(CutList.total_sqft)).label("cut_sqft"),
            )
            .group_by(CutList.fab_id)
            .subquery()
        )

        completed_sqft_subq = (
            select(
                InstallCompletion.fab_id.label("fab_id"),
                func.sum(safe_cast_numeric(InstallCompletion.total_sqft_installed)).label("completed_sqft"),
            )
            .group_by(InstallCompletion.fab_id)
            .subquery()
        )

        avg_work_pct_subq = (
            select(
                ShopCutPlan.fab_id.label("fab_id"),
                func.avg(cast(ShopCutPlan.work_percentage, Numeric)).label("avg_work_pct"),
            )
            .group_by(ShopCutPlan.fab_id)
            .subquery()
        )

        detail_query = (
            select(
                Fab.fab_type,
                Fab.id,
                BusinessJob.job_number,
                Account.name,
                BusinessJob.name,
                Fab.input_area,
                StoneType.name,
                StoneColor.name,
                StoneThickness.thickness,
                Fab.total_sqft,
                cut_sqft_subq.c.cut_sqft,
                Fab.wj_linft,
                Fab.edging_linft,
                Fab.cnc_linft,
                Fab.miter_linft,
                completed_sqft_subq.c.completed_sqft,
                Fab.shop_est_completion_date,
                avg_work_pct_subq.c.avg_work_pct,
            )
            .select_from(Fab)
            .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
            .join(Account, Account.id == BusinessJob.account_id, isouter=True)
            .join(StoneType, StoneType.id == Fab.stone_type_id, isouter=True)
            .join(StoneColor, StoneColor.id == Fab.stone_color_id, isouter=True)
            .join(StoneThickness, StoneThickness.id == Fab.stone_thickness_id, isouter=True)
            .join(cut_sqft_subq, cut_sqft_subq.c.fab_id == Fab.id, isouter=True)
            .join(completed_sqft_subq, completed_sqft_subq.c.fab_id == Fab.id, isouter=True)
            .join(avg_work_pct_subq, avg_work_pct_subq.c.fab_id == Fab.id, isouter=True)
            .where(
                Fab.cutlist_complete.is_(True),
                func.lower(func.coalesce(Fab.current_stage, "")) != "install_completion",
            )
            .order_by(Fab.id.asc())
        )

        detail_result = (await db.execute(detail_query)).all()

        detail_rows: list[dict] = []
        for (
            fab_type,
            fab_id,
            job_number,
            account_name,
            job_name,
            areas,
            stone_type,
            stone_color,
            stone_thickness,
            total_sqft_raw,
            cut_sqft_raw,
            wj_linft_raw,
            edging_linft_raw,
            cnc_linft_raw,
            miter_linft_raw,
            completed_sqft_raw,
            est_completion_date,
            avg_work_pct_raw,
        ) in detail_result:
            total_sqft = round(_to_float(total_sqft_raw), 2)
            cut_sqft = round(_to_float(cut_sqft_raw), 2)
            wj_linft = round(_to_float(wj_linft_raw), 2)
            edging_linft = round(_to_float(edging_linft_raw), 2)
            cnc_linft = round(_to_float(cnc_linft_raw), 2)
            miter_linft = round(_to_float(miter_linft_raw), 2)
            completed_sqft = round(_to_float(completed_sqft_raw), 2)
            avg_work_pct = _normalize_work_percent(avg_work_pct_raw)

            linear_total = wj_linft + edging_linft + cnc_linft + miter_linft

            detail_rows.append(
                {
                    "FabType": fab_type,
                    "FabID": fab_id,
                    "JobNumber": job_number,
                    "Account": account_name,
                    "JobName": job_name,
                    "Areas": areas,
                    "StoneType": stone_type,
                    "StoneColor": stone_color,
                    "StoneThickness": stone_thickness,
                    "TotalSqFt": total_sqft,
                    "CutSqFt": cut_sqft,
                    "CutPercent": _to_percent(cut_sqft, total_sqft),
                    "WJLinFt": wj_linft,
                    "WJPercent": _to_percent(wj_linft, linear_total),
                    "EdgingLinFt": edging_linft,
                    "EdgingPercent": _to_percent(edging_linft, linear_total),
                    "CNCLinFt": cnc_linft,
                    "CNCPercent": _to_percent(cnc_linft, linear_total),
                    "MiterLinFt": miter_linft,
                    "MiterPercent": _to_percent(miter_linft, linear_total),
                    "TouchupSqFt": completed_sqft,
                    "TouchupPercent": _to_percent(completed_sqft, total_sqft),
                    "EstCompletionDate": est_completion_date.isoformat() if est_completion_date else None,
                    "PercentageComplete": avg_work_pct,
                }
            )

        summary_map: dict[tuple[str | None, str | None, str | None], dict[str, float | int | str | None]] = {}
        for row in detail_rows:
            account_name = row.get("Account")
            job_number = row.get("JobNumber")
            job_name = row.get("JobName")
            key = (account_name, job_number, job_name)

            total_sqft = _to_float(row.get("TotalSqFt"))
            work_pct = _to_float(row.get("PercentageComplete"))
            completed_equivalent = total_sqft * (work_pct / 100.0)

            if key not in summary_map:
                summary_map[key] = {
                    "Account": account_name,
                    "JobNumber": job_number,
                    "JobName": job_name,
                    "TotalSqFt_sum": 0.0,
                    "CompletedSqFt_sum": 0.0,
                    "Rows": 0,
                }

            summary_map[key]["TotalSqFt_sum"] = _to_float(summary_map[key]["TotalSqFt_sum"]) + total_sqft
            summary_map[key]["CompletedSqFt_sum"] = _to_float(summary_map[key]["CompletedSqFt_sum"]) + completed_equivalent
            summary_map[key]["Rows"] = int(summary_map[key]["Rows"]) + 1

        summary_rows: list[dict] = []
        for _, summary in sorted(
            summary_map.items(),
            key=lambda item: (
                str(item[0][0] or "").lower(),
                str(item[0][1] or "").lower(),
                str(item[0][2] or "").lower(),
            ),
        ):
            total_sqft_sum = round(_to_float(summary["TotalSqFt_sum"]), 2)
            completed_sqft_sum = round(_to_float(summary["CompletedSqFt_sum"]), 2)
            summary_rows.append(
                {
                    "Account": summary["Account"],
                    "JobNumber": summary["JobNumber"],
                    "JobName": summary["JobName"],
                    "TotalSqFt_sum": total_sqft_sum,
                    "CompletedSqFt_sum": completed_sqft_sum,
                    "Rows": int(summary["Rows"]),
                    "PercentComplete": _to_percent(completed_sqft_sum, total_sqft_sum),
                }
            )

        return detail_rows, summary_rows


async def send_monthly_end_of_month_status_report(
    year: int,
    month: int,
    recipients_override: Optional[str] = None,
) -> dict:
    recipients = recipients_override or os.getenv("EOM_STATUS_REPORT_RECIPIENTS", ADMIN_EMAIL)
    recipients = ",".join([email.strip() for email in recipients.split(",") if email.strip()])
    if not recipients:
        logger.warning("Skipping end-of-month status report: no recipients configured")
        return {"sent": False, "reason": "no_recipients", "recipients": "", "detail_rows": 0, "summary_rows": 0}

    detail_rows, summary_rows = await _build_report_rows(year, month)

    month_label = f"{calendar.month_name[month]} {year}"
    csv_detail = _rows_to_csv_bytes(DETAIL_COLUMNS, detail_rows)
    csv_summary = _rows_to_csv_bytes(SUMMARY_COLUMNS, summary_rows)
    xlsx_combined = _rows_to_xlsx_bytes(detail_rows, summary_rows)

    attachments = [
        (f"end_of_month_status_{year}_{month:02d}.csv", csv_detail, "text/csv"),
        (f"end_of_month_summary_{year}_{month:02d}.csv", csv_summary, "text/csv"),
        (
            f"end_of_month_status_{year}_{month:02d}.xlsx",
            xlsx_combined,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    ]

    body = (
        f"<p>Attached are the End of Month Status reports for <strong>{month_label}</strong>.</p>"
        f"<ul>"
        f"<li>Detailed status report (CSV)</li>"
        f"<li>Account/job completion summary (CSV)</li>"
        f"<li>Combined workbook with both reports (XLSX)</li>"
        f"</ul>"
    )

    loop = asyncio.get_running_loop()
    send_fn = partial(
        send_email_with_attachments,
        to_email=recipients,
        subject=f"End of Month Status Report - {month_label}",
        body=body,
        attachments=attachments,
        is_html=True,
    )
    await loop.run_in_executor(None, send_fn)

    logger.info(
        "End-of-month status report email sent for %s to %s (detail_rows=%s, summary_rows=%s)",
        month_label,
        recipients,
        len(detail_rows),
        len(summary_rows),
    )

    return {
        "sent": True,
        "reason": None,
        "recipients": recipients,
        "detail_rows": len(detail_rows),
        "summary_rows": len(summary_rows),
        "year": year,
        "month": month,
    }


async def _scheduler_loop() -> None:
    global _last_trigger_key

    while True:
        now = datetime.now()
        trigger_key = now.strftime("%Y-%m-%d %H:%M")

        if now.day == 1 and now.hour == 0 and now.minute == 0 and trigger_key != _last_trigger_key:
            previous_day = date(now.year, now.month, 1) - timedelta(days=1)
            try:
                await send_monthly_end_of_month_status_report(previous_day.year, previous_day.month)
            except Exception:
                logger.exception("Failed sending scheduled end-of-month status report")
            _last_trigger_key = trigger_key

        await asyncio.sleep(20)


def start_monthly_status_report_scheduler() -> None:
    global _scheduler_task

    is_enabled = os.getenv("EOM_STATUS_REPORT_ENABLED", "true").strip().lower() == "true"
    if not is_enabled:
        logger.info("End-of-month status report scheduler is disabled")
        return

    if _scheduler_task and not _scheduler_task.done():
        return

    _scheduler_task = asyncio.create_task(_scheduler_loop())
    logger.info("End-of-month status report scheduler started")


async def stop_monthly_status_report_scheduler() -> None:
    global _scheduler_task

    if _scheduler_task is None:
        return

    _scheduler_task.cancel()
    try:
        await _scheduler_task
    except asyncio.CancelledError:
        pass
    finally:
        _scheduler_task = None
        logger.info("End-of-month status report scheduler stopped")
