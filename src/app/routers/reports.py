from __future__ import annotations
# pyright: reportGeneralTypeIssues=false, reportMissingImports=false

import csv
import calendar
import io
import json
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import and_, case, cast, func, literal_column, or_, select
from sqlalchemy import Numeric, String
from sqlalchemy.ext.asyncio import AsyncSession

from src.app.database import get_db
from src.app.database.account import Account
from src.app.database.business_job import BusinessJob
from src.app.database.department import Department
from src.app.database.edge import Edge
from src.app.database.fab import Fab
from src.app.database.installer_rate_history import InstallerRateHistory
from src.app.database.service_level_setting import ServiceLevelSetting
from src.app.database.installer_job_timer_session import InstallerJobTimerSession
from src.app.database.templater_job_timer_session import TemplaterJobTimerSession
from src.app.database.shop_cut_plan import ShopCutPlan
from src.app.database.stone_color import StoneColor
from src.app.database.stone_thickness import StoneThickness
from src.app.database.stone_type import StoneType
from src.app.database.user import User
from src.app.interface.generated_schemas import CNCDrafting, CostOfStone, CutList, DraftingSession, InstallCompletion, InstallScheduling, PlanningSection, ResurfaceScheduling, Revision, ShopRevision, Templating
from src.app.interface.response_wrappers import SuccessResponse, success_response
from src.app.middleware.jwt_auth import get_current_user
from src.app.routers.fabs import FAB_STAGES, PUNCHOUT_REDIRECT_FAB_TYPES, _active_shop_cut_plan_visibility_filter, _get_shop_current_stage, _pending_cnc_widget_filter, _stage_filter_condition
from src.app.service.monthly_end_of_month_status_report import send_monthly_end_of_month_status_report
from src.app.utils.helpers import error_response

router = APIRouter()


class InstallerRateUpsert(BaseModel):
    installer_id: int = Field(..., gt=0)
    hourly_rate: float = Field(..., gt=0)
    effective_from: datetime
    effective_to: Optional[datetime] = None
    is_active: bool = True


class RedoPatchRequest(BaseModel):
    no_of_pieces: Optional[int] = Field(default=None, ge=0)
    sqft: Optional[float] = Field(default=None, ge=0)
    cost_per_sqft: Optional[float] = Field(default=None, ge=0)
    total_cost: Optional[float] = Field(default=None, ge=0)
    department: Optional[str] = None
    person_name: Optional[str] = None
    reason: Optional[str] = None


class MonthlyCutCompletionPatchRequest(BaseModel):
    revenue: Optional[float] = Field(default=None, ge=0)
    cost_of_stone: Optional[float] = Field(default=None, ge=0)
    revenue_per_sq_ft: Optional[float] = Field(default=None, ge=0)


class MonthlyInstallCompletionPatchRequest(BaseModel):
    revenue: Optional[float] = Field(default=None, ge=0)
    sq_ft: Optional[float] = Field(default=None, ge=0)
    revenue_per_sq_ft: Optional[float] = Field(default=None, ge=0)
    installer_id: Optional[int] = Field(default=None, gt=0)
    installer_name: Optional[str] = None


class DailyInstallCompletionPatchRequest(BaseModel):
    revenue: Optional[float] = Field(default=None, ge=0)
    sq_ft: Optional[float] = Field(default=None, ge=0)
    installer_id: Optional[int] = Field(default=None, gt=0)
    installer_name: Optional[str] = None


class InstallationTemplateDashboardPatchRequest(BaseModel):
    type: str = Field(..., description="'templater' or 'installer'")
    fab_id: int
    job_id: Optional[int] = None
    installer_id: Optional[int] = None
    activity_complete: Optional[bool] = Field(None, alias="Activity Complete")
    sqft_templated: Optional[float] = Field(None, alias="sqft templated")
    sqft_not_templated: Optional[float] = Field(None, alias="sqft not templated")
    reason: Optional[str] = None
    duration: Optional[int] = None


class ServiceLevelSettingUpdate(BaseModel):
    target_days: Optional[float] = Field(default=None, ge=0, description="Days at or below = green")
    at_risk_days: Optional[float] = Field(default=None, ge=0, description="Yellow window: days beyond target before red")
    is_applicable: Optional[bool] = None


class ServiceLevelSettingCreate(BaseModel):
    fab_type: str = Field(..., min_length=1, max_length=100)
    stage_name: str = Field(..., min_length=1, max_length=100)
    target_days: float = Field(default=1.0, ge=0)
    at_risk_days: float = Field(default=0.0, ge=0)
    is_applicable: bool = True


@router.get("/redos", response_model=SuccessResponse[dict])
@router.get("/reports/redos", response_model=SuccessResponse[dict])
async def get_ag_redo_report(
    from_date: Optional[date] = Query(None, description="Inclusive from date filter"),
    to_date: Optional[date] = Query(None, description="Inclusive to date filter"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return AG REDO FAB rows with complete report fields."""
    _ = current_user
    start_dt, end_dt = _range_bounds(from_date, to_date)

    filters = [func.lower(Fab.fab_type) == "ag redo"]
    _apply_datetime_filters(filters, Fab.created_at, start_dt, end_dt)

    rows = (
        await db.execute(
            select(
                Fab.created_at,
                Fab.fab_type,
                Fab.id,
                BusinessJob.job_number,
                Account.name,
                BusinessJob.name,
                StoneType.name,
                StoneColor.name,
                Edge.name,
                StoneThickness.thickness,
                Fab.input_area,
                Fab.no_of_pieces,
                Fab.total_sqft,
                Fab.cost_per_sqft,
                Fab.redo_total_sqft,
                Fab.notes,
                Fab.redo_department,
                Fab.redo_requested_by,
            )
            .select_from(Fab)
            .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
            .join(Account, Account.id == BusinessJob.account_id, isouter=True)
            .join(StoneType, StoneType.id == Fab.stone_type_id, isouter=True)
            .join(StoneColor, StoneColor.id == Fab.stone_color_id, isouter=True)
            .join(Edge, Edge.id == Fab.edge_id, isouter=True)
            .join(StoneThickness, StoneThickness.id == Fab.stone_thickness_id, isouter=True)
            .add_columns(Fab.cost_of_stone)
            .where(and_(*filters))
            .order_by(Fab.created_at.desc(), Fab.id.desc())
        )
    ).all()

    department_ids = sorted({int(row[16]) for row in rows if row[16] is not None})
    requested_by_ids = sorted({int(row[17]) for row in rows if row[17] is not None})

    department_name_map: dict[int, str] = {}
    if department_ids:
        department_rows = (
            await db.execute(
                select(Department.id, Department.name).where(Department.id.in_(department_ids))
            )
        ).all()
        department_name_map = {int(row[0]): row[1] for row in department_rows}

    requested_by_name_map: dict[int, str] = {}
    if requested_by_ids:
        requested_by_rows = (
            await db.execute(
                select(User.id, User.first_name, User.last_name).where(User.id.in_(requested_by_ids))
            )
        ).all()
        requested_by_name_map = {
            int(row[0]): (f"{(row[1] or '').strip()} {(row[2] or '').strip()}".strip() or f"User {row[0]}")
            for row in requested_by_rows
        }

    departments = (await db.execute(select(Department.name).order_by(Department.name.asc()))).all()
    department_options = ", ".join(row[0] for row in departments if row[0])

    data = []
    for (
        created_at,
        fab_type,
        fab_id,
        job_number,
        account_name,
        job_name,
        stone_type,
        stone_color,
        edge_name,
        stone_thickness,
        input_area,
        no_of_pieces,
        sqft,
        cost_per_sqft_raw,
        redo_total_sqft_raw,
        notes,
        redo_department,
        redo_requested_by,
        cost_of_stone_raw,
    ) in rows:
        cost_per_sqft = round(_to_float(cost_per_sqft_raw), 2) if cost_per_sqft_raw is not None else None
        sqft_value = round(_to_float(sqft), 2)
        redo_sqft_raw_value = _to_float(redo_total_sqft_raw)
        if redo_sqft_raw_value <= 0:
            # Backward-compatibility for older rows where redo_total_sqft wasn't populated.
            redo_sqft_raw_value = _to_float(sqft)
        redo_sqft_value = round(redo_sqft_raw_value, 2)
        total_cost = round(_to_float(cost_per_sqft_raw) * redo_sqft_raw_value * 2.1, 2)
        cost_of_stone_value = round(_to_float(cost_of_stone_raw), 2)

        info_parts = [
            account_name,
            job_name,
            stone_type,
            stone_color,
            stone_thickness,
        ]
        fab_info = " - ".join(str(part).strip() for part in info_parts if part and str(part).strip()) or None

        data.append(
            {
                "fab_created_date": created_at.date().isoformat() if created_at else None,
                "fab_type": fab_type,
                "fab_id": fab_id,
                "job_number": job_number,
                "job_name": job_name,
                "account_name": account_name,
                "stone_type_name": stone_type,
                "stone_color_name": stone_color,
                "edge_name": edge_name,
                "stone_thickness_value": stone_thickness,
                "input_area": input_area,
                "fab_info": fab_info,
                "no_of_pieces": int(_to_float(no_of_pieces)),
                "sqft": sqft_value,
                "cost_per_sqft": cost_per_sqft,
                "redo_total_sqft": redo_sqft_value,
                "total_cost": total_cost,
                "cost_of_stone": cost_of_stone_value,
                "redo_department": redo_department,
                "redo_requested_by": redo_requested_by,
                "department": department_name_map.get(int(redo_department), None) if redo_department is not None else None,
                "person_name": requested_by_name_map.get(int(redo_requested_by), None) if redo_requested_by is not None else None,
                "note": _notes_to_text(notes),
                "reason": _notes_to_text(notes),
                "department_options": department_options,
            }
        )

    summary = {
        "total_sqft": round(sum(_to_float(row.get("redo_total_sqft")) for row in data), 2),
        "total_cost": round(sum(_to_float(row.get("total_cost")) for row in data), 2),
        "total_cost_of_stone": round(sum(_to_float(row.get("cost_of_stone")) for row in data), 2),
    }

    return success_response(
        {
            "summary": summary,
            "rows": data,
        },
        "AG REDO report data retrieved successfully",
    )


@router.patch("/redos/{redo_id}", response_model=SuccessResponse[dict])
@router.patch("/reports/redos/{redo_id}", response_model=SuccessResponse[dict])
async def patch_redo_record(
    redo_id: int,
    patch: RedoPatchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update editable AG Redo row fields by FAB/redo id."""
    if all(
        value is None
        for value in [
            patch.no_of_pieces,
            patch.sqft,
            patch.cost_per_sqft,
            patch.total_cost,
            patch.department,
            patch.person_name,
        ]
    ):
        raise error_response("At least one field is required", 400)

    fab = (await db.execute(select(Fab).where(Fab.id == redo_id))).scalar_one_or_none()
    if not fab:
        raise error_response("Redo record not found", 404)
    if (fab.fab_type or "").strip().lower() != "ag redo":
        raise error_response("Record is not an AG redo FAB", 400)

    now = datetime.now()

    if patch.no_of_pieces is not None:
        fab.no_of_pieces = patch.no_of_pieces
    if patch.sqft is not None:
        # Persist to redo_total_sqft for AG redo cost calculations and keep total_sqft in sync.
        fab.redo_total_sqft = patch.sqft
        fab.total_sqft = patch.sqft
    if patch.department is not None:
        normalized_department = patch.department.strip()
        if not normalized_department:
            fab.redo_department = None
        else:
            department_row = (
                await db.execute(
                    select(Department.id).where(func.lower(Department.name) == normalized_department.lower()).limit(1)
                )
            ).first()
            if department_row is None:
                raise error_response("Department not found", 400)
            fab.redo_department = int(department_row[0])
    if patch.person_name is not None:
        normalized_person_name = patch.person_name.strip()
        if not normalized_person_name:
            fab.redo_requested_by = None
        else:
            user = await _resolve_user_by_name(db, normalized_person_name)
            if user is None:
                raise error_response("Person not found", 400)
            fab.redo_requested_by = int(user.id)

    cost_record: Optional[CostOfStone] = None
    if patch.cost_per_sqft is not None or patch.total_cost is not None:
        if fab.cost_of_stone_id:
            cost_record = await db.get(CostOfStone, fab.cost_of_stone_id)
        if cost_record is None:
            cost_record = (
                await db.execute(select(CostOfStone).where(CostOfStone.fab_id == fab.id))
            ).scalar_one_or_none()

        if cost_record is None:
            cost_record = CostOfStone(
                fab_id=fab.id,
                stone_color_id=fab.stone_color_id,
                stone_type_id=fab.stone_type_id,
                total_sqft=str(fab.total_sqft) if fab.total_sqft is not None else None,
                status_id=1,
                created_at=now,
                updated_at=now,
                updated_by=current_user.id,
            )
            db.add(cost_record)
            await db.flush()

        if patch.cost_per_sqft is not None:
            cost_record.cost_per_sqft = f"{patch.cost_per_sqft:.2f}"

        total_cost_value: Optional[float] = None
        if patch.total_cost is not None:
            total_cost_value = patch.total_cost
        elif patch.cost_per_sqft is not None:
            total_cost_value = patch.cost_per_sqft * _to_float(fab.total_sqft) * 2.1

        if total_cost_value is not None:
            total_cost_value = round(total_cost_value, 2)
            cost_record.total_cost = f"{total_cost_value:.2f}"
            fab.cost_of_stone = total_cost_value

        cost_record.updated_at = now
        cost_record.updated_by = current_user.id
        if fab.cost_of_stone_id is None:
            fab.cost_of_stone_id = cost_record.id

    fab.updated_at = now
    fab.updated_by = current_user.id

    await db.commit()

    current_cost_per_sqft = None
    current_total_cost = None
    if cost_record is not None:
        current_cost_per_sqft = round(_to_float(cost_record.cost_per_sqft), 2) if cost_record.cost_per_sqft is not None else None
        current_total_cost = round(_to_float(cost_record.total_cost), 2) if cost_record.total_cost is not None else round(_to_float(fab.cost_of_stone), 2)
    elif fab.cost_of_stone_id:
        existing_cost = await db.get(CostOfStone, fab.cost_of_stone_id)
        if existing_cost:
            current_cost_per_sqft = round(_to_float(existing_cost.cost_per_sqft), 2) if existing_cost.cost_per_sqft is not None else None
            current_total_cost = round(_to_float(existing_cost.total_cost), 2) if existing_cost.total_cost is not None else round(_to_float(fab.cost_of_stone), 2)

    if current_total_cost is None and fab.cost_of_stone is not None:
        current_total_cost = round(_to_float(fab.cost_of_stone), 2)

    department_name = None
    if fab.redo_department is not None:
        department_name = (
            await db.execute(select(Department.name).where(Department.id == fab.redo_department).limit(1))
        ).scalar_one_or_none()

    person_name = None
    if fab.redo_requested_by is not None:
        user_row = (
            await db.execute(
                select(User.first_name, User.last_name).where(User.id == fab.redo_requested_by).limit(1)
            )
        ).first()
        if user_row is not None:
            person_name = (f"{(user_row[0] or '').strip()} {(user_row[1] or '').strip()}".strip() or f"User {fab.redo_requested_by}")

    effective_sqft = _to_float(fab.redo_total_sqft)
    if effective_sqft <= 0:
        effective_sqft = _to_float(fab.total_sqft)

    return success_response(
        {
            "redo_id": fab.id,
            "no_of_pieces": fab.no_of_pieces,
            "sqft": round(effective_sqft, 2),
            "cost_per_sqft": current_cost_per_sqft,
            "total_cost": current_total_cost,
            "department": department_name,
            "person_name": person_name,
            "reason": None,
        },
        "Redo record updated successfully",
    )


def _range_bounds(start_date: Optional[date], end_date: Optional[date]) -> tuple[Optional[datetime], Optional[datetime]]:
    start_dt = datetime.combine(start_date, time.min) if start_date else None
    end_dt = datetime.combine(end_date, time.max) if end_date else None
    return start_dt, end_dt


def _month_bounds(year: int, month: int) -> tuple[datetime, datetime]:
    start_dt = datetime(year, month, 1)
    if month == 12:
        end_dt = datetime(year + 1, 1, 1) - timedelta(microseconds=1)
    else:
        end_dt = datetime(year, month + 1, 1) - timedelta(microseconds=1)
    return start_dt, end_dt


def _parse_month_input(month_value: str) -> Optional[int]:
    text = (month_value or "").strip()
    if not text:
        return None

    if text.isdigit():
        month_num = int(text)
        return month_num if 1 <= month_num <= 12 else None

    lower = text.lower()
    for idx, name in enumerate(calendar.month_name):
        if idx == 0:
            continue
        if lower == name.lower():
            return idx
    for idx, name in enumerate(calendar.month_abbr):
        if idx == 0:
            continue
        if lower == name.lower():
            return idx

    return None


def _days_between(start_value: Optional[datetime], end_value: Optional[datetime]) -> Optional[int]:
    if start_value is None or end_value is None:
        return None
    return max((end_value.date() - start_value.date()).days, 0)


def _days_between_count_partial_day_as_full(
    start_value: Optional[datetime],
    end_value: Optional[datetime],
) -> Optional[int]:
    """Return elapsed days where any positive partial day counts as one full day."""
    if start_value is None or end_value is None:
        return None

    delta = end_value - start_value
    if delta <= timedelta(0):
        return 0

    full_days = delta.days
    return full_days if delta == timedelta(days=full_days) else full_days + 1


def _apply_datetime_filters(filters: list, field, start_dt: Optional[datetime], end_dt: Optional[datetime]) -> None:
    if start_dt is not None:
        filters.append(field >= start_dt)
    if end_dt is not None:
        filters.append(field <= end_dt)


def _to_float(value) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


async def _resolve_user_by_name(db: AsyncSession, installer_name: str) -> Optional[User]:
    normalized = (installer_name or "").strip().lower()
    if not normalized:
        return None
    like_value = f"%{normalized}%"
    return (
        await db.execute(
            select(User)
            .where(
                or_(
                    func.lower(User.first_name).like(like_value),
                    func.lower(User.last_name).like(like_value),
                    func.lower(User.username).like(like_value),
                    func.lower(User.email).like(like_value),
                    func.lower(
                        func.concat(
                            func.coalesce(User.first_name, ""),
                            " ",
                            func.coalesce(User.last_name, ""),
                        )
                    ).like(like_value),
                )
            )
            .order_by(User.id.asc())
            .limit(1)
        )
    ).scalar_one_or_none()


async def _get_latest_install_completion_for_fab(
    db: AsyncSession,
    fab_id: int,
) -> Optional[InstallCompletion]:
    return (
        await db.execute(
            select(InstallCompletion)
            .where(InstallCompletion.fab_id == fab_id)
            .order_by(InstallCompletion.completion_date.desc(), InstallCompletion.id.desc())
            .limit(1)
        )
    ).scalar_one_or_none()


def _safe_numeric_col(col):
    """Cast a text column to Numeric, silently returning NULL for non-numeric values.

    Uses PostgreSQL regexp_replace to strip non-numeric characters and NULLIF to
    convert the empty-string case to NULL before casting, so rows that contain
    placeholder text (e.g. 'string') never raise InvalidTextRepresentationError.
    """
    # Cast to TEXT first so regexp_replace works regardless of whether the column
    # is stored as text or a numeric type (double precision, numeric, etc.).
    # Then strip non-numeric chars, convert empty string to NULL, and cast to NUMERIC.
    text_col = cast(col, String)
    return cast(func.nullif(func.regexp_replace(text_col, '[^0-9.]', '', 'g'), ''), Numeric)


def _rows_from_mapping(value: dict) -> list[dict]:
    return [{"metric": k, "value": v} for k, v in value.items()]


def _format_duration_hhmm(total_seconds: int) -> str:
    safe_seconds = max(int(total_seconds or 0), 0)
    hours = safe_seconds // 3600
    minutes = (safe_seconds % 3600) // 60
    return f"{hours:02d}:{minutes:02d}"


def _notes_to_text(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, list):
        joined = " | ".join(str(item).strip() for item in value if str(item).strip())
        return joined or None
    text = str(value).strip()
    return text or None


def _safe_div(numerator: float, denominator: float) -> float:
    if denominator == 0:
        return 0.0
    return numerator / denominator


def _week_windows_for_month(year: int, month: int, week_ending_weekday: int = 4) -> list[dict]:
    month_start = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    month_end = date(year, month, last_day)

    days_until_week_end = (week_ending_weekday - month_start.weekday()) % 7
    current_week_end = month_start + timedelta(days=days_until_week_end)

    windows: list[dict] = []
    while current_week_end <= month_end:
        week_start = current_week_end - timedelta(days=6)
        overlap_start = max(week_start, month_start)
        overlap_end = min(current_week_end, month_end)
        windows.append(
            {
                "week_start": week_start,
                "week_end": current_week_end,
                "overlap_start": overlap_start,
                "overlap_end": overlap_end,
                "number_of_days": (overlap_end - overlap_start).days + 1,
            }
        )
        current_week_end = current_week_end + timedelta(days=7)

    if not windows or windows[-1]["overlap_end"] < month_end:
        trailing_week_end = month_end
        trailing_week_start = trailing_week_end - timedelta(days=6)
        overlap_start = max(trailing_week_start, month_start)
        windows.append(
            {
                "week_start": trailing_week_start,
                "week_end": trailing_week_end,
                "overlap_start": overlap_start,
                "overlap_end": month_end,
                "number_of_days": (month_end - overlap_start).days + 1,
            }
        )

    return windows


def _parse_payroll_overrides(raw: Optional[str]) -> tuple[dict, Optional[str]]:
    if not raw:
        return {}, None
    try:
        decoded = json.loads(raw)
    except json.JSONDecodeError:
        return {}, "payroll_overrides_json must be valid JSON"
    if not isinstance(decoded, dict):
        return {}, "payroll_overrides_json must decode to an object"
    return decoded, None


def _payroll_value(overrides: dict, week_key: str, field: str, default: float = 0.0) -> float:
    weekly_block = overrides.get(week_key, {})
    default_block = overrides.get("_default", {})

    if isinstance(weekly_block, dict) and field in weekly_block:
        return _to_float(weekly_block.get(field))
    if isinstance(default_block, dict) and field in default_block:
        return _to_float(default_block.get(field))
    return default


def _client_layout_sections(report_key: str, data: dict) -> Optional[list[tuple[str, list[dict]]]]:
    if report_key == "redo-analysis":
        period = data.get("period", {})
        summary = data.get("summary", {})
        redo_by_stage = data.get("redo_by_stage", [])
        accounts = data.get("top_accounts_with_redo", [])
        jobs = data.get("top_jobs_with_redo", [])

        total_redo = sum(int(_to_float(row.get("redo_count", 0))) for row in redo_by_stage) or 0
        stage_rows = []
        for row in redo_by_stage:
            stage_count = int(_to_float(row.get("redo_count", 0)))
            stage_rows.append(
                {
                    "stage": row.get("stage"),
                    "redo_count": stage_count,
                    "redo_share_percent": round((stage_count / total_redo) * 100, 2) if total_redo else 0.0,
                }
            )

        hotspot_rows = []
        for idx, row in enumerate(accounts, start=1):
            hotspot_rows.append(
                {
                    "rank": idx,
                    "category": "Account",
                    "reference": row.get("account_id"),
                    "name": row.get("account_name"),
                    "redo_count": int(_to_float(row.get("redo_count", 0))),
                    "redo_revenue": round(_to_float(row.get("redo_revenue", 0)), 2),
                }
            )
        for idx, row in enumerate(jobs, start=1):
            hotspot_rows.append(
                {
                    "rank": idx,
                    "category": "Job",
                    "reference": row.get("job_number") or row.get("job_id"),
                    "name": row.get("job_name"),
                    "redo_count": int(_to_float(row.get("redo_count", 0))),
                    "redo_revenue": "",
                }
            )

        summary_rows = [
            {
                "period_start": period.get("start_date"),
                "period_end": period.get("end_date"),
                "total_fabs": int(_to_float(summary.get("total_fabs", 0))),
                "revised_fabs": int(_to_float(summary.get("revised_fabs", 0))),
                "redo_rate_percent": round(_to_float(summary.get("redo_rate_percent", 0)), 2),
                "revision_events": int(_to_float(summary.get("revision_events", 0))),
            }
        ]

        return [
            ("redo_summary", summary_rows),
            ("redo_by_stage", stage_rows),
            ("redo_hotspots", hotspot_rows),
        ]

    if report_key == "shop-status":
        period = data.get("period", {})
        stage_rows = []
        for row in data.get("stage_status", []):
            fab_count = int(_to_float(row.get("fab_count", 0)))
            stalled_count = int(_to_float(row.get("stalled_over_14_days", 0)))
            stage_rows.append(
                {
                    "stage": row.get("stage"),
                    "fab_count": fab_count,
                    "avg_age_days": round(_to_float(row.get("avg_age_days", 0)), 2),
                    "max_age_days": round(_to_float(row.get("max_age_days", 0)), 2),
                    "stalled_over_14_days": stalled_count,
                    "stalled_rate_percent": round((stalled_count / fab_count) * 100, 2) if fab_count else 0.0,
                }
            )

        alerts_rows = []
        for row in stage_rows:
            if row["stalled_over_14_days"] <= 0:
                continue
            stalled_rate = _to_float(row["stalled_rate_percent"])
            priority = "high" if stalled_rate >= 30 else "medium" if stalled_rate >= 15 else "low"
            alerts_rows.append(
                {
                    "stage": row["stage"],
                    "priority": priority,
                    "stalled_over_14_days": row["stalled_over_14_days"],
                    "stalled_rate_percent": row["stalled_rate_percent"],
                }
            )

        cover_rows = [
            {
                "report": "End of Month Shop Status",
                "period_start": period.get("start_date"),
                "period_end": period.get("end_date"),
                "generated_at": datetime.now().isoformat(),
            }
        ]

        return [
            ("report_cover", cover_rows),
            ("shop_status", stage_rows),
            ("shop_alerts", alerts_rows),
        ]

    return None


def _report_sections(report_key: str, data: dict, layout: str = "default") -> list[tuple[str, list[dict]]]:
    if layout == "client":
        sections = _client_layout_sections(report_key, data)
        if sections is not None:
            return sections

    if report_key == "overview":
        kpi_rows = _rows_from_mapping(data.get("kpis", {}))
        return [
            ("kpis", kpi_rows),
            ("stage_breakdown", data.get("stage_breakdown", [])),
        ]

    if report_key == "redo-analysis":
        summary_rows = _rows_from_mapping(data.get("summary", {}))
        return [
            ("summary", summary_rows),
            ("redo_by_stage", data.get("redo_by_stage", [])),
            ("top_accounts_with_redo", data.get("top_accounts_with_redo", [])),
            ("top_jobs_with_redo", data.get("top_jobs_with_redo", [])),
        ]

    if report_key == "shop-status":
        return [("stage_status", data.get("stage_status", []))]

    if report_key == "install-performance":
        summary_rows = _rows_from_mapping(data.get("summary", {}))
        return [
            ("summary", summary_rows),
            ("installer_breakdown", data.get("installer_breakdown", [])),
        ]

    if report_key == "installation-template":
        summary_rows = _rows_from_mapping(data.get("summary", {}))
        return [
            ("summary", summary_rows),
            ("installation_template_rows", data.get("rows", [])),
        ]

    if report_key == "installation-template-dashboard":
        summary_rows = _rows_from_mapping(data.get("summary", {}))
        return [
            ("summary", summary_rows),
            ("installation_template_groups", data.get("groups", [])),
            ("installation_template_rows", data.get("rows", [])),
        ]

    if report_key == "monthly-install-completion":
        summary_rows = _rows_from_mapping(data.get("summary", {}))
        return [
            ("summary", summary_rows),
            ("daily_totals", data.get("daily_totals", [])),
            ("rows", data.get("rows", [])),
        ]

    if report_key == "daily-install-completion":
        summary_rows = _rows_from_mapping(data.get("summary", {}))
        return [
            ("summary", summary_rows),
            ("daily_totals", data.get("daily_totals", [])),
            ("rows", data.get("rows", [])),
        ]

    if report_key == "monthly-cut-completion":
        summary_rows = _rows_from_mapping(data.get("summary", {}))
        return [
            ("summary", summary_rows),
            ("daily_totals", data.get("daily_totals", [])),
            ("rows", data.get("rows", [])),
        ]

    if report_key == "turnaround-times":
        stats_rows = _rows_from_mapping(data.get("summary", {}))
        return [
            ("summary", stats_rows),
            ("rows", data.get("rows", [])),
        ]

    if report_key == "service-level":
        summary_rows = _rows_from_mapping(data.get("summary", {}))
        widget_rows = []
        widgets = data.get("widgets", {})
        if widgets:
            for metric_key, metric_value in widgets.items():
                if isinstance(metric_value, dict):
                    widget_rows.append({"metric": metric_key, "value": json.dumps(metric_value)})
                else:
                    widget_rows.append({"metric": metric_key, "value": metric_value})
        return [
            ("widgets", widget_rows),
            ("summary", summary_rows),
            ("stage_bottleneck_heat_map", data.get("stage_bottleneck_heat_map", [])),
            ("fab_status_rows", data.get("fab_status_rows", [])),
            ("cycle_time_stats", data.get("cycle_time_stats", [])),
            ("aging_backlog", data.get("aging_backlog", [])),
            ("breach_details", data.get("breach_details", [])),
        ]

    if report_key == "weekly-trends":
        return [("weekly_trends", data.get("weekly_trends", []))]

    if report_key == "weekly-fabrication-labor-cost":
        monthly_report = data.get("monthly_report", {})
        totals_rows = _rows_from_mapping(monthly_report.get("totals", {}))
        return [
            ("monthly_totals", totals_rows),
            ("weekly_breakdown", monthly_report.get("weekly_breakdown", [])),
            ("annual_monthly_summary", data.get("annual_monthly_summary", [])),
        ]

    if report_key == "weekly-installer-labor-cost":
        monthly_report = data.get("monthly_report", {})
        totals_rows = _rows_from_mapping(monthly_report.get("totals", {}))
        return [
            ("monthly_totals", totals_rows),
            ("weekly_breakdown", monthly_report.get("weekly_breakdown", [])),
            ("annual_monthly_summary", data.get("annual_monthly_summary", [])),
        ]

    if report_key == "management-packet":
        rows = []
        for block_name, block_value in data.items():
            if isinstance(block_value, dict):
                rows.append({"section": block_name, "payload": json.dumps(block_value)})
            else:
                rows.append({"section": block_name, "payload": str(block_value)})
        return [("packet", rows)]

    return [("data", [{"payload": json.dumps(data)}])]


@router.get("/reports/owner/weekly-fabrication-labor-cost", response_model=SuccessResponse[dict])
async def get_owner_weekly_fabrication_labor_cost_report(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    total_employees: int = Query(40, ge=0, description="Display header value for total employees"),
    overhead_per_week: float = Query(38512.69, ge=0, description="Default overhead amount per week"),
    week_ending_weekday: int = Query(4, ge=0, le=6, description="Week ending day: Monday=0 ... Sunday=6"),
    payroll_overrides_json: Optional[str] = Query(
        None,
        description=(
            "Optional JSON object keyed by week-ending date (YYYY-MM-DD) for external payroll values. "
            "Supported fields: head_count, shop_management, wages_basic_shop_yard, overtime_shop_yard, "
            "cost_of_overtime_pct, total_labor_cost, regular_hours, overtime_hours, overhead_per_week. "
            "Use _default object for defaults."
        ),
    ),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Weekly fabrication labor cost analysis with monthly and annual summaries."""
    payroll_overrides, payroll_error = _parse_payroll_overrides(payroll_overrides_json)
    if payroll_error:
        return success_response(None, payroll_error, status_code=400)

    async def _compute_month(month_num: int) -> dict:
        windows = _week_windows_for_month(year, month_num, week_ending_weekday)
        weekly_rows: list[dict] = []

        for window in windows:
            week_start = window["week_start"]
            week_end = window["week_end"]
            week_key = week_end.isoformat()
            week_start_dt = datetime.combine(week_start, time.min)
            week_end_dt = datetime.combine(week_end, time.max)

            cut_metrics = (
                await db.execute(
                    select(func.sum(_safe_numeric_col(Fab.saw_cut_lnft)), func.sum(_safe_numeric_col(Fab.total_sqft))).where(
                        Fab.shop_date_schedule >= week_start_dt,
                        Fab.shop_date_schedule <= week_end_dt,
                    )
                )
            ).first()

            completion_metrics = (
                await db.execute(
                    select(
                        func.sum(_safe_numeric_col(InstallCompletion.total_sqft_installed)),
                        func.sum(_safe_numeric_col(Fab.revenue)),
                        func.sum(_safe_numeric_col(Fab.gp)),
                    )
                    .join(Fab, Fab.id == InstallCompletion.fab_id, isouter=True)
                    .where(
                        InstallCompletion.is_completed.is_(True),
                        InstallCompletion.completion_date >= week_start_dt,
                        InstallCompletion.completion_date <= week_end_dt,
                    )
                )
            ).first()

            cut_sqft_saw = _to_float(cut_metrics[0] if cut_metrics and cut_metrics[0] is not None else (cut_metrics[1] if cut_metrics else 0.0))
            completed_sqft = _to_float(completion_metrics[0] if completion_metrics else 0.0)
            gross_revenue = _to_float(completion_metrics[1] if completion_metrics else 0.0)
            gross_profit = _to_float(completion_metrics[2] if completion_metrics else 0.0)

            number_of_days = int(window["number_of_days"])
            avg_sqft_per_day = _safe_div(completed_sqft, number_of_days)
            avg_revenue_per_day = _safe_div(gross_revenue, number_of_days)

            head_count = _payroll_value(payroll_overrides, week_key, "head_count", float(total_employees))
            shop_management = _payroll_value(payroll_overrides, week_key, "shop_management", 0.0)
            wages_basic_shop_yard = _payroll_value(payroll_overrides, week_key, "wages_basic_shop_yard", 0.0)
            overtime_shop_yard = _payroll_value(payroll_overrides, week_key, "overtime_shop_yard", 0.0)
            regular_hours = _payroll_value(payroll_overrides, week_key, "regular_hours", 0.0)
            overtime_hours = _payroll_value(payroll_overrides, week_key, "overtime_hours", 0.0)
            week_overhead = _payroll_value(payroll_overrides, week_key, "overhead_per_week", overhead_per_week)

            total_labor_cost_override = _payroll_value(payroll_overrides, week_key, "total_labor_cost", -1.0)
            total_labor_cost = (
                total_labor_cost_override
                if total_labor_cost_override >= 0
                else (wages_basic_shop_yard + overtime_shop_yard)
            )

            overtime_pct_override = _payroll_value(payroll_overrides, week_key, "cost_of_overtime_pct", -1.0)
            cost_of_overtime_pct = (
                overtime_pct_override
                if overtime_pct_override >= 0
                else (_safe_div(overtime_shop_yard, wages_basic_shop_yard) * 100)
            )

            total_hours = regular_hours + overtime_hours
            overtime_hours_pct = _safe_div(overtime_hours, total_hours) * 100
            shop_labor_per_hour = _safe_div(total_labor_cost, total_hours)
            shop_overhead_per_hour = _safe_div(week_overhead, total_hours)
            shop_labor_overhead_per_hour = shop_labor_per_hour + shop_overhead_per_hour
            manpower_cost_per_hour = shop_labor_per_hour

            sqft_per_labor_hour = _safe_div(completed_sqft, total_hours)
            shop_productivity_sqft_per_hour = _safe_div(cut_sqft_saw, total_hours)
            labor_cost_per_sq_ft = _safe_div(total_labor_cost, completed_sqft)
            labor_cost_pct_per_dollar_sold = _safe_div(total_labor_cost, gross_revenue) * 100
            shop_overhead_cost_per_sqft = _safe_div(week_overhead, completed_sqft)
            shop_total_cost_per_sqft = labor_cost_per_sq_ft + shop_overhead_cost_per_sqft
            gross_profit_per_sf_completed = _safe_div(gross_profit, completed_sqft)
            gross_profit_less_shop_total_cost_psf = gross_profit_per_sf_completed - shop_total_cost_per_sqft
            gross_revenue_per_sqft_fabricated = _safe_div(gross_revenue, completed_sqft)

            weekly_rows.append(
                {
                    "week_ending": week_key,
                    "number_of_days": number_of_days,
                    "cut_sqft_saw": round(cut_sqft_saw, 2),
                    "wj_sqft": 0.0,
                    "completed_sqft": round(completed_sqft, 2),
                    "average_sqft_per_day": round(avg_sqft_per_day, 2),
                    "gross_revenue": round(gross_revenue, 2),
                    "gross_profit": round(gross_profit, 2),
                    "average_revenue_per_day": round(avg_revenue_per_day, 2),
                    "total_head_count_inc_yard": round(head_count, 2),
                    "shop_management": round(shop_management, 2),
                    "wages_basic_shop_yard": round(wages_basic_shop_yard, 2),
                    "overtime_shop_yard": round(overtime_shop_yard, 2),
                    "cost_of_overtime_pct": round(cost_of_overtime_pct, 2),
                    "total_labor_cost": round(total_labor_cost, 2),
                    "regular_hours": round(regular_hours, 2),
                    "overtime_hours": round(overtime_hours, 2),
                    "overtime_hours_pct": round(overtime_hours_pct, 2),
                    "total_hours": round(total_hours, 2),
                    "shop_labor_per_hour": round(shop_labor_per_hour, 2),
                    "shop_overhead_per_hour": round(shop_overhead_per_hour, 2),
                    "shop_labor_overhead_per_hour": round(shop_labor_overhead_per_hour, 2),
                    "manpower_cost_per_hour": round(manpower_cost_per_hour, 2),
                    "sqft_per_labor_hour": round(sqft_per_labor_hour, 2),
                    "shop_productivity_sqft_per_hour": round(shop_productivity_sqft_per_hour, 2),
                    "labor_cost_per_sq_ft": round(labor_cost_per_sq_ft, 2),
                    "labor_cost_pct_per_dollar_sold": round(labor_cost_pct_per_dollar_sold, 2),
                    "shop_overhead_cost_per_sqft": round(shop_overhead_cost_per_sqft, 2),
                    "shop_total_cost_per_sqft": round(shop_total_cost_per_sqft, 2),
                    "gross_profit_per_sf_completed": round(gross_profit_per_sf_completed, 2),
                    "gross_profit_less_shop_total_cost_psf": round(gross_profit_less_shop_total_cost_psf, 2),
                    "gross_revenue_per_sqft_fabricated": round(gross_revenue_per_sqft_fabricated, 2),
                    "overhead_per_week": round(week_overhead, 2),
                }
            )

        week_count = len(weekly_rows)

        totals = {
            "number_of_weeks": week_count,
            "number_of_days": int(sum(_to_float(row["number_of_days"]) for row in weekly_rows)),
            "cut_sqft_saw": round(sum(_to_float(row["cut_sqft_saw"]) for row in weekly_rows), 2),
            "wj_sqft": round(sum(_to_float(row["wj_sqft"]) for row in weekly_rows), 2),
            "completed_sqft": round(sum(_to_float(row["completed_sqft"]) for row in weekly_rows), 2),
            "gross_revenue": round(sum(_to_float(row["gross_revenue"]) for row in weekly_rows), 2),
            "gross_profit": round(sum(_to_float(row["gross_profit"]) for row in weekly_rows), 2),
            "shop_management": round(sum(_to_float(row["shop_management"]) for row in weekly_rows), 2),
            "wages_basic_shop_yard": round(sum(_to_float(row["wages_basic_shop_yard"]) for row in weekly_rows), 2),
            "overtime_shop_yard": round(sum(_to_float(row["overtime_shop_yard"]) for row in weekly_rows), 2),
            "total_labor_cost": round(sum(_to_float(row["total_labor_cost"]) for row in weekly_rows), 2),
            "regular_hours": round(sum(_to_float(row["regular_hours"]) for row in weekly_rows), 2),
            "overtime_hours": round(sum(_to_float(row["overtime_hours"]) for row in weekly_rows), 2),
            "total_hours": round(sum(_to_float(row["total_hours"]) for row in weekly_rows), 2),
            "total_head_count_inc_yard": round(_safe_div(sum(_to_float(row["total_head_count_inc_yard"]) for row in weekly_rows), week_count), 2) if week_count else 0.0,
            "overhead_per_week": round(sum(_to_float(row["overhead_per_week"]) for row in weekly_rows), 2),
        }

        totals["average_sqft_per_day"] = round(_safe_div(totals["completed_sqft"], totals["number_of_days"]), 2)
        totals["average_revenue_per_day"] = round(_safe_div(totals["gross_revenue"], totals["number_of_days"]), 2)
        totals["cost_of_overtime_pct"] = round(_safe_div(totals["overtime_shop_yard"], totals["wages_basic_shop_yard"]) * 100, 2)
        totals["overtime_hours_pct"] = round(_safe_div(totals["overtime_hours"], totals["total_hours"]) * 100, 2)
        totals["shop_labor_per_hour"] = round(_safe_div(totals["total_labor_cost"], totals["total_hours"]), 2)
        totals["shop_overhead_per_hour"] = round(_safe_div(totals["overhead_per_week"], totals["total_hours"]), 2)
        totals["shop_labor_overhead_per_hour"] = round(totals["shop_labor_per_hour"] + totals["shop_overhead_per_hour"], 2)
        totals["manpower_cost_per_hour"] = totals["shop_labor_per_hour"]
        totals["sqft_per_labor_hour"] = round(_safe_div(totals["completed_sqft"], totals["total_hours"]), 2)
        totals["shop_productivity_sqft_per_hour"] = round(_safe_div(totals["cut_sqft_saw"], totals["total_hours"]), 2)
        totals["labor_cost_per_sq_ft"] = round(_safe_div(totals["total_labor_cost"], totals["completed_sqft"]), 2)
        totals["labor_cost_pct_per_dollar_sold"] = round(_safe_div(totals["total_labor_cost"], totals["gross_revenue"]) * 100, 2)
        totals["shop_overhead_cost_per_sqft"] = round(_safe_div(totals["overhead_per_week"], totals["completed_sqft"]), 2)
        totals["shop_total_cost_per_sqft"] = round(totals["labor_cost_per_sq_ft"] + totals["shop_overhead_cost_per_sqft"], 2)
        totals["gross_profit_per_sf_completed"] = round(_safe_div(totals["gross_profit"], totals["completed_sqft"]), 2)
        totals["gross_profit_less_shop_total_cost_psf"] = round(totals["gross_profit_per_sf_completed"] - totals["shop_total_cost_per_sqft"], 2)
        totals["gross_revenue_per_sqft_fabricated"] = round(_safe_div(totals["gross_revenue"], totals["completed_sqft"]), 2)

        return {
            "weekly_breakdown": weekly_rows,
            "totals": totals,
            "month": calendar.month_name[month_num],
            "month_number": month_num,
        }

    monthly_report = await _compute_month(month)

    annual_monthly_summary = []
    for month_num in range(1, 13):
        month_data = await _compute_month(month_num)
        totals = month_data["totals"]
        annual_monthly_summary.append(
            {
                "month": calendar.month_name[month_num],
                "month_number": month_num,
                "number_of_weeks": int(_to_float(totals.get("number_of_weeks", 0))),
                "completed_sqft": round(_to_float(totals.get("completed_sqft", 0)), 2),
                "gross_revenue": round(_to_float(totals.get("gross_revenue", 0)), 2),
                "gross_profit": round(_to_float(totals.get("gross_profit", 0)), 2),
                "total_labor_cost": round(_to_float(totals.get("total_labor_cost", 0)), 2),
                "total_hours": round(_to_float(totals.get("total_hours", 0)), 2),
                "labor_cost_pct_per_dollar_sold": round(_to_float(totals.get("labor_cost_pct_per_dollar_sold", 0)), 2),
                "gross_profit_less_shop_total_cost_psf": round(_to_float(totals.get("gross_profit_less_shop_total_cost_psf", 0)), 2),
            }
        )

    month_start = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    month_end = date(year, month, last_day)

    return success_response(
        {
            "title": "Weekly Fabrication Labor Cost Analysis",
            "period": {
                "start_date": month_start.isoformat(),
                "end_date": month_end.isoformat(),
            },
            "display": {
                "total_employee": total_employees,
                "default_overhead_per_week": round(overhead_per_week, 2),
                "week_ending_weekday": week_ending_weekday,
            },
            "payroll_source": {
                "mode": "external_overrides",
                "override_fields": [
                    "head_count",
                    "shop_management",
                    "wages_basic_shop_yard",
                    "overtime_shop_yard",
                    "cost_of_overtime_pct",
                    "total_labor_cost",
                    "regular_hours",
                    "overtime_hours",
                    "overhead_per_week",
                ],
                "provided_weekly_keys": sorted([k for k in payroll_overrides.keys() if k != "_default"]),
                "has_default_override": isinstance(payroll_overrides.get("_default"), dict),
            },
            "monthly_report": monthly_report,
            "annual_monthly_summary": annual_monthly_summary,
        },
        "Owner weekly fabrication labor cost analysis report generated",
    )


@router.get("/reports/owner/weekly-installer-labor-cost", response_model=SuccessResponse[dict])
async def get_owner_weekly_installer_labor_cost_report(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    total_employees: int = Query(40, ge=0, description="Display header value for total employees"),
    overhead_per_week: float = Query(38512.69, ge=0, description="Default overhead amount per week"),
    week_ending_weekday: int = Query(4, ge=0, le=6, description="Week ending day: Monday=0 ... Sunday=6"),
    payroll_overrides_json: Optional[str] = Query(
        None,
        description=(
            "Optional JSON object keyed by week-ending date (YYYY-MM-DD) for external payroll values. "
            "Supported fields: sub_contractor_head_count, wages_sub_contractor, head_count, "
            "wages_basic_installer, overtime_installer, overtime_pct, total_labor_cost, "
            "regular_hours, overtime_hours, overhead_per_week. Use _default object for defaults."
        ),
    ),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Weekly installer labor cost analysis with monthly and annual summaries."""
    payroll_overrides, payroll_error = _parse_payroll_overrides(payroll_overrides_json)
    if payroll_error:
        return success_response(None, payroll_error, status_code=400)

    async def _compute_month(month_num: int) -> dict:
        windows = _week_windows_for_month(year, month_num, week_ending_weekday)
        weekly_rows: list[dict] = []

        for window in windows:
            week_start = window["week_start"]
            week_end = window["week_end"]
            week_key = week_end.isoformat()
            week_start_dt = datetime.combine(week_start, time.min)
            week_end_dt = datetime.combine(week_end, time.max)

            install_sqft_row = (
                await db.execute(
                    select(func.sum(_safe_numeric_col(InstallScheduling.total_sqft))).where(
                        InstallScheduling.scheduled_install_date >= week_start_dt,
                        InstallScheduling.scheduled_install_date <= week_end_dt,
                    )
                )
            ).first()

            completion_metrics = (
                await db.execute(
                    select(
                        func.sum(_safe_numeric_col(InstallCompletion.total_sqft_installed)),
                        func.sum(_safe_numeric_col(Fab.revenue)),
                        func.sum(_safe_numeric_col(Fab.gp)),
                    )
                    .join(Fab, Fab.id == InstallCompletion.fab_id, isouter=True)
                    .where(
                        InstallCompletion.is_completed.is_(True),
                        InstallCompletion.completion_date >= week_start_dt,
                        InstallCompletion.completion_date <= week_end_dt,
                    )
                )
            ).first()

            install_sqft = _to_float(install_sqft_row[0] if install_sqft_row else 0.0)
            completed_sqft = _to_float(completion_metrics[0] if completion_metrics else 0.0)
            gross_revenue = _to_float(completion_metrics[1] if completion_metrics else 0.0)
            gross_profit = _to_float(completion_metrics[2] if completion_metrics else 0.0)

            number_of_days = int(window["number_of_days"])
            avg_sqft_per_day = _safe_div(completed_sqft, number_of_days)
            avg_revenue_per_day = _safe_div(gross_revenue, number_of_days)

            sub_contractor_head_count = _payroll_value(payroll_overrides, week_key, "sub_contractor_head_count", 0.0)
            wages_sub_contractor = _payroll_value(payroll_overrides, week_key, "wages_sub_contractor", 0.0)
            total_head_count = _payroll_value(payroll_overrides, week_key, "head_count", float(total_employees))
            wages_basic_installer = _payroll_value(payroll_overrides, week_key, "wages_basic_installer", 0.0)
            overtime_installer = _payroll_value(payroll_overrides, week_key, "overtime_installer", 0.0)
            regular_hours = _payroll_value(payroll_overrides, week_key, "regular_hours", 0.0)
            overtime_hours = _payroll_value(payroll_overrides, week_key, "overtime_hours", 0.0)
            week_overhead = _payroll_value(payroll_overrides, week_key, "overhead_per_week", overhead_per_week)

            total_labor_cost_override = _payroll_value(payroll_overrides, week_key, "total_labor_cost", -1.0)
            total_labor_cost = (
                total_labor_cost_override
                if total_labor_cost_override >= 0
                else (wages_basic_installer + overtime_installer + wages_sub_contractor)
            )

            overtime_pct_override = _payroll_value(payroll_overrides, week_key, "overtime_pct", -1.0)
            overtime_pct = (
                overtime_pct_override
                if overtime_pct_override >= 0
                else (_safe_div(overtime_installer, wages_basic_installer) * 100)
            )

            total_hours = regular_hours + overtime_hours
            overtime_total_hours_pct = _safe_div(overtime_hours, total_hours) * 100
            hourly_labor_cost_all_installers = _safe_div(total_labor_cost, total_hours)
            hourly_overhead_cost_all_installers = _safe_div(week_overhead, total_hours)
            hourly_cost_all_installers_inc_overhead = hourly_labor_cost_all_installers + hourly_overhead_cost_all_installers
            hourly_cost_per_installer_inc_overhead = _safe_div(hourly_cost_all_installers_inc_overhead, total_head_count)

            sqft_per_labor_hour = _safe_div(completed_sqft, total_hours)
            installer_productivity_sqft_per_hour = _safe_div(install_sqft, total_hours)
            labor_cost_per_sq_ft = _safe_div(total_labor_cost, completed_sqft)
            labor_cost_pct_per_dollar_sold = _safe_div(total_labor_cost, gross_revenue) * 100
            overhead_cost_per_sqft_installed = _safe_div(week_overhead, completed_sqft)
            cost_to_install_per_sqft = labor_cost_per_sq_ft + overhead_cost_per_sqft_installed
            gross_profit_per_sf_installed = _safe_div(gross_profit, completed_sqft)
            gross_profit_less_installer_total_cost_psf = gross_profit_per_sf_installed - cost_to_install_per_sqft
            gross_revenue_per_sq_ft = _safe_div(gross_revenue, completed_sqft)

            weekly_rows.append(
                {
                    "week_ending": week_key,
                    "number_of_days_per_week": number_of_days,
                    "install_sqft_per_week": round(install_sqft, 2),
                    "completed_sqft_per_week": round(completed_sqft, 2),
                    "average_sqft_per_day": round(avg_sqft_per_day, 2),
                    "gross_revenue": round(gross_revenue, 2),
                    "gross_profit": round(gross_profit, 2),
                    "average_revenue_per_day": round(avg_revenue_per_day, 2),
                    "sub_contractor_head_count": round(sub_contractor_head_count, 2),
                    "wages_sub_contractor": round(wages_sub_contractor, 2),
                    "total_head_count": round(total_head_count, 2),
                    "wages_basic_installer": round(wages_basic_installer, 2),
                    "overtime_installer": round(overtime_installer, 2),
                    "overtime_pct": round(overtime_pct, 2),
                    "total_labor_cost": round(total_labor_cost, 2),
                    "regular_hours": round(regular_hours, 2),
                    "overtime_hours": round(overtime_hours, 2),
                    "overtime_total_hours_pct": round(overtime_total_hours_pct, 2),
                    "total_hours": round(total_hours, 2),
                    "hourly_labor_cost_all_installers": round(hourly_labor_cost_all_installers, 2),
                    "hourly_overhead_cost_all_installers": round(hourly_overhead_cost_all_installers, 2),
                    "hourly_cost_all_installers_inc_overhead": round(hourly_cost_all_installers_inc_overhead, 2),
                    "hourly_cost_per_installer_inc_overhead": round(hourly_cost_per_installer_inc_overhead, 2),
                    "sqft_per_labor_hour": round(sqft_per_labor_hour, 2),
                    "installer_productivity_sqft_per_hour": round(installer_productivity_sqft_per_hour, 2),
                    "labor_cost_per_sq_ft": round(labor_cost_per_sq_ft, 2),
                    "labor_cost_pct_per_dollar_sold": round(labor_cost_pct_per_dollar_sold, 2),
                    "overhead_cost_per_sqft_installed": round(overhead_cost_per_sqft_installed, 2),
                    "cost_to_install_per_sqft": round(cost_to_install_per_sqft, 2),
                    "gross_profit_per_sf_installed": round(gross_profit_per_sf_installed, 2),
                    "gross_profit_less_installer_total_cost_psf": round(gross_profit_less_installer_total_cost_psf, 2),
                    "gross_revenue_per_sq_ft": round(gross_revenue_per_sq_ft, 2),
                    "overhead_per_week": round(week_overhead, 2),
                }
            )

        week_count = len(weekly_rows)

        totals = {
            "number_of_weeks": week_count,
            "number_of_days_per_week": int(sum(_to_float(row["number_of_days_per_week"]) for row in weekly_rows)),
            "install_sqft_per_week": round(sum(_to_float(row["install_sqft_per_week"]) for row in weekly_rows), 2),
            "completed_sqft_per_week": round(sum(_to_float(row["completed_sqft_per_week"]) for row in weekly_rows), 2),
            "gross_revenue": round(sum(_to_float(row["gross_revenue"]) for row in weekly_rows), 2),
            "gross_profit": round(sum(_to_float(row["gross_profit"]) for row in weekly_rows), 2),
            "sub_contractor_head_count": round(_safe_div(sum(_to_float(row["sub_contractor_head_count"]) for row in weekly_rows), week_count), 2) if week_count else 0.0,
            "wages_sub_contractor": round(sum(_to_float(row["wages_sub_contractor"]) for row in weekly_rows), 2),
            "total_head_count": round(_safe_div(sum(_to_float(row["total_head_count"]) for row in weekly_rows), week_count), 2) if week_count else 0.0,
            "wages_basic_installer": round(sum(_to_float(row["wages_basic_installer"]) for row in weekly_rows), 2),
            "overtime_installer": round(sum(_to_float(row["overtime_installer"]) for row in weekly_rows), 2),
            "total_labor_cost": round(sum(_to_float(row["total_labor_cost"]) for row in weekly_rows), 2),
            "regular_hours": round(sum(_to_float(row["regular_hours"]) for row in weekly_rows), 2),
            "overtime_hours": round(sum(_to_float(row["overtime_hours"]) for row in weekly_rows), 2),
            "total_hours": round(sum(_to_float(row["total_hours"]) for row in weekly_rows), 2),
            "overhead_per_week": round(sum(_to_float(row["overhead_per_week"]) for row in weekly_rows), 2),
        }

        totals["average_sqft_per_day"] = round(_safe_div(totals["completed_sqft_per_week"], totals["number_of_days_per_week"]), 2)
        totals["average_revenue_per_day"] = round(_safe_div(totals["gross_revenue"], totals["number_of_days_per_week"]), 2)
        totals["overtime_pct"] = round(_safe_div(totals["overtime_installer"], totals["wages_basic_installer"]) * 100, 2)
        totals["overtime_total_hours_pct"] = round(_safe_div(totals["overtime_hours"], totals["total_hours"]) * 100, 2)
        totals["hourly_labor_cost_all_installers"] = round(_safe_div(totals["total_labor_cost"], totals["total_hours"]), 2)
        totals["hourly_overhead_cost_all_installers"] = round(_safe_div(totals["overhead_per_week"], totals["total_hours"]), 2)
        totals["hourly_cost_all_installers_inc_overhead"] = round(totals["hourly_labor_cost_all_installers"] + totals["hourly_overhead_cost_all_installers"], 2)
        totals["hourly_cost_per_installer_inc_overhead"] = round(_safe_div(totals["hourly_cost_all_installers_inc_overhead"], totals["total_head_count"]), 2)
        totals["sqft_per_labor_hour"] = round(_safe_div(totals["completed_sqft_per_week"], totals["total_hours"]), 2)
        totals["installer_productivity_sqft_per_hour"] = round(_safe_div(totals["install_sqft_per_week"], totals["total_hours"]), 2)
        totals["labor_cost_per_sq_ft"] = round(_safe_div(totals["total_labor_cost"], totals["completed_sqft_per_week"]), 2)
        totals["labor_cost_pct_per_dollar_sold"] = round(_safe_div(totals["total_labor_cost"], totals["gross_revenue"]) * 100, 2)
        totals["overhead_cost_per_sqft_installed"] = round(_safe_div(totals["overhead_per_week"], totals["completed_sqft_per_week"]), 2)
        totals["cost_to_install_per_sqft"] = round(totals["labor_cost_per_sq_ft"] + totals["overhead_cost_per_sqft_installed"], 2)
        totals["gross_profit_per_sf_installed"] = round(_safe_div(totals["gross_profit"], totals["completed_sqft_per_week"]), 2)
        totals["gross_profit_less_installer_total_cost_psf"] = round(totals["gross_profit_per_sf_installed"] - totals["cost_to_install_per_sqft"], 2)
        totals["gross_revenue_per_sq_ft"] = round(_safe_div(totals["gross_revenue"], totals["completed_sqft_per_week"]), 2)

        return {
            "weekly_breakdown": weekly_rows,
            "totals": totals,
            "month": calendar.month_name[month_num],
            "month_number": month_num,
        }

    monthly_report = await _compute_month(month)

    annual_monthly_summary = []
    for month_num in range(1, 13):
        month_data = await _compute_month(month_num)
        totals = month_data["totals"]
        annual_monthly_summary.append(
            {
                "month": calendar.month_name[month_num],
                "month_number": month_num,
                "number_of_weeks": int(_to_float(totals.get("number_of_weeks", 0))),
                "completed_sqft": round(_to_float(totals.get("completed_sqft_per_week", 0)), 2),
                "gross_revenue": round(_to_float(totals.get("gross_revenue", 0)), 2),
                "gross_profit": round(_to_float(totals.get("gross_profit", 0)), 2),
                "total_labor_cost": round(_to_float(totals.get("total_labor_cost", 0)), 2),
                "total_hours": round(_to_float(totals.get("total_hours", 0)), 2),
                "labor_cost_pct_per_dollar_sold": round(_to_float(totals.get("labor_cost_pct_per_dollar_sold", 0)), 2),
                "gross_profit_less_installer_total_cost_psf": round(_to_float(totals.get("gross_profit_less_installer_total_cost_psf", 0)), 2),
            }
        )

    month_start = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    month_end = date(year, month, last_day)

    return success_response(
        {
            "title": "Weekly Installer Labor Cost Analysis",
            "period": {
                "start_date": month_start.isoformat(),
                "end_date": month_end.isoformat(),
            },
            "display": {
                "total_employee": total_employees,
                "default_overhead_per_week": round(overhead_per_week, 2),
                "week_ending_weekday": week_ending_weekday,
            },
            "payroll_source": {
                "mode": "external_overrides",
                "override_fields": [
                    "sub_contractor_head_count",
                    "wages_sub_contractor",
                    "head_count",
                    "wages_basic_installer",
                    "overtime_installer",
                    "overtime_pct",
                    "total_labor_cost",
                    "regular_hours",
                    "overtime_hours",
                    "overhead_per_week",
                ],
                "provided_weekly_keys": sorted([k for k in payroll_overrides.keys() if k != "_default"]),
                "has_default_override": isinstance(payroll_overrides.get("_default"), dict),
            },
            "monthly_report": monthly_report,
            "annual_monthly_summary": annual_monthly_summary,
        },
        "Owner weekly installer labor cost analysis report generated",
    )


def _csv_bytes(report_key: str, data: dict, layout: str = "default") -> bytes:
    sections = _report_sections(report_key, data, layout=layout)
    buf = io.StringIO()
    writer = csv.writer(buf)

    for section_name, rows in sections:
        writer.writerow([section_name])
        if not rows:
            writer.writerow(["no_data"])
            writer.writerow([])
            continue

        headers = sorted({k for row in rows for k in row.keys()})
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row.get(h, "") for h in headers])
        writer.writerow([])

    return buf.getvalue().encode("utf-8")


def _xlsx_bytes(report_key: str, data: dict, layout: str = "default") -> bytes:
    from openpyxl import Workbook  # type: ignore[reportMissingImports]

    sections = _report_sections(report_key, data, layout=layout)
    wb = Workbook()
    # Remove the default sheet and create one per section.
    active_sheet = wb.active
    if active_sheet is not None:
        wb.remove(active_sheet)

    for idx, (section_name, rows) in enumerate(sections):
        ws = wb.create_sheet(title=section_name[:31] or f"sheet_{idx + 1}")
        if not rows:
            ws.cell(row=1, column=1, value="no_data")
            continue

        headers = sorted({k for row in rows for k in row.keys()})
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=str(row.get(header, "")))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def _unwrap_success_data(response) -> dict:
    payload = json.loads(response.body.decode("utf-8"))
    return payload.get("data", {})


@router.get("/reports/owner/overview", response_model=SuccessResponse[dict])
async def get_owner_overview_report(
    start_date: Optional[date] = Query(None, description="Inclusive start date filter"),
    end_date: Optional[date] = Query(None, description="Inclusive end date filter"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Owner summary KPIs across jobs, fabs, finance, and install pipeline."""
    start_dt, end_dt = _range_bounds(start_date, end_date)

    fab_filters = []
    _apply_datetime_filters(fab_filters, Fab.created_at, start_dt, end_dt)

    job_filters = []
    _apply_datetime_filters(job_filters, BusinessJob.created_at, start_dt, end_dt)

    completed_install_exists = (
        select(InstallCompletion.id)
        .where(
            InstallCompletion.fab_id == Fab.id,
            InstallCompletion.is_completed.is_(True),
        )
        .exists()
    )
    incomplete_install_exists = (
        select(InstallCompletion.id)
        .where(
            InstallCompletion.fab_id == Fab.id,
            InstallCompletion.is_completed.is_(False),
        )
        .exists()
    )

    completed_install_filters = [completed_install_exists]
    _apply_datetime_filters(completed_install_filters, Fab.updated_at, start_dt, end_dt)

    pending_install_filters = [
        or_(
            Fab.current_stage == "install_scheduling",
            and_(
                Fab.current_stage == "install_completion",
                incomplete_install_exists,
            ),
        )
    ]
    _apply_datetime_filters(pending_install_filters, Fab.created_at, start_dt, end_dt)

    templating_exists = (
        select(Templating.id)
        .where(Templating.fab_id == Fab.id)
        .limit(1)
        .exists()
    )
    active_fab_filters = [
        or_(
            Fab.template_needed.is_(False),
            and_(Fab.template_needed.is_(True), templating_exists),
        ),
    ]
    _apply_datetime_filters(active_fab_filters, Fab.created_at, start_dt, end_dt)

    total_jobs = (
        await db.execute(select(func.count(BusinessJob.id)).where(and_(*job_filters)) if job_filters else select(func.count(BusinessJob.id)))
    ).scalar() or 0
    total_fabs = (
        await db.execute(select(func.count(Fab.id)).where(and_(*fab_filters)) if fab_filters else select(func.count(Fab.id)))
    ).scalar() or 0

    completed_installs = (
        await db.execute(select(func.count(Fab.id)).where(and_(*completed_install_filters)))
    ).scalar() or 0
    pending_installs = (
        await db.execute(select(func.count(Fab.id)).where(and_(*pending_install_filters)))
    ).scalar() or 0
    active_fabs = (
        await db.execute(select(func.count(Fab.id)).where(and_(*active_fab_filters)))
    ).scalar() or 0

    revenue_filters = [Fab.revenue.isnot(None)]
    _apply_datetime_filters(revenue_filters, Fab.created_at, start_dt, end_dt)
    total_revenue = (
        await db.execute(select(func.sum(Fab.revenue)).where(and_(*revenue_filters)))
    ).scalar() or 0

    gp_filters = [Fab.gp.isnot(None)]
    _apply_datetime_filters(gp_filters, Fab.created_at, start_dt, end_dt)
    total_gp = (
        await db.execute(select(func.sum(Fab.gp)).where(and_(*gp_filters)))
    ).scalar() or 0

    stage_query = (
        select(Fab.current_stage, func.count(Fab.id).label("count"))
        .group_by(Fab.current_stage)
        .order_by(func.count(Fab.id).desc())
    )
    if fab_filters:
        stage_query = stage_query.where(and_(*fab_filters))

    stage_rows = (await db.execute(stage_query)).all()
    stage_counts_dict = {str(stage_name): int(stage_count or 0) for stage_name, stage_count in stage_rows if stage_name}

    final_programming_filters = [_stage_filter_condition("final_programming")]
    _apply_datetime_filters(final_programming_filters, Fab.created_at, start_dt, end_dt)
    stage_counts_dict["final_programming"] = (
        await db.execute(select(func.count(Fab.id)).where(and_(*final_programming_filters)))
    ).scalar() or 0

    cut_list_filters = [_stage_filter_condition("cut_list")]
    _apply_datetime_filters(cut_list_filters, Fab.created_at, start_dt, end_dt)
    stage_counts_dict["cut_list"] = (
        await db.execute(select(func.count(Fab.id)).where(and_(*cut_list_filters)))
    ).scalar() or 0

    install_completion_filters = [_stage_filter_condition("install_completion")]
    _apply_datetime_filters(install_completion_filters, Fab.created_at, start_dt, end_dt)
    stage_counts_dict["install_completion"] = (
        await db.execute(select(func.count(Fab.id)).where(and_(*install_completion_filters)))
    ).scalar() or 0

    slabsmith_pending_filters = [
        or_(Fab.current_stage == "sales_ct", Fab.current_stage == "revision"),
        or_(Fab.slab_smith_ag_needed.is_(True), Fab.slab_smith_cust_needed.is_(True)),
        Fab.slabsmith_completed_date.is_(None),
    ]
    _apply_datetime_filters(slabsmith_pending_filters, Fab.created_at, start_dt, end_dt)
    stage_counts_dict["slab_smith_request"] = (
        await db.execute(select(func.count(Fab.id)).where(and_(*slabsmith_pending_filters)))
    ).scalar() or 0

    cnc_filters = [_pending_cnc_widget_filter()]
    _apply_datetime_filters(cnc_filters, Fab.created_at, start_dt, end_dt)
    stage_counts_dict["cnc"] = (
        await db.execute(select(func.count(Fab.id)).where(and_(*cnc_filters)))
    ).scalar() or 0

    cost_of_stone_queue_filters = [
        Fab.sct_completed.is_(True),
        or_(
            Fab.cost_of_stone.is_(None),
            func.trim(cast(Fab.cost_of_stone, String)) == "",
        ),
    ]
    _apply_datetime_filters(cost_of_stone_queue_filters, Fab.created_at, start_dt, end_dt)
    stage_counts_dict["cost_of_stone"] = (
        await db.execute(select(func.count(Fab.id)).where(and_(*cost_of_stone_queue_filters)))
    ).scalar() or 0

    shop_est_or_install_filter = or_(
        Fab.shop_est_completion_date.isnot(None),
        Fab.fab_type.in_(PUNCHOUT_REDIRECT_FAB_TYPES),
    )
    already_scheduled_for_install_exists = (
        select(InstallScheduling.id)
        .where(
            InstallScheduling.fab_id == Fab.id,
            InstallScheduling.installer_id.isnot(None),
            InstallScheduling.scheduled_install_date.isnot(None),
        )
        .exists()
    )
    install_scheduling_filters = [
        Fab.status_id == 1,
        shop_est_or_install_filter,
        ~already_scheduled_for_install_exists,
    ]
    _apply_datetime_filters(install_scheduling_filters, Fab.created_at, start_dt, end_dt)
    stage_counts_dict["install_scheduling"] = (
        await db.execute(select(func.count(Fab.id)).where(and_(*install_scheduling_filters)))
    ).scalar() or 0

    # Keep Shop count aligned with /fabs?current_stage=shop semantics.
    shop_filters = [
        _stage_filter_condition("shop"),
        _active_shop_cut_plan_visibility_filter(),
    ]
    _apply_datetime_filters(shop_filters, Fab.created_at, start_dt, end_dt)
    stage_counts_dict["shop"] = (
        await db.execute(select(func.count(Fab.id)).where(and_(*shop_filters)))
    ).scalar() or 0

    stage_display_labels = {
        "slab_smith_request": "Slabsmith",
        "final_programming": "Final_programming",
        "cnc": "CNC",
    }

    stage_breakdown_rows: list[dict] = []
    for stage_name in FAB_STAGES:
        count_value = int(stage_counts_dict.get(stage_name, 0))
        if count_value <= 0 and stage_name not in {"slab_smith_request", "final_programming", "cnc"}:
            continue
        stage_breakdown_rows.append(
            {
                "stage": stage_display_labels.get(stage_name, stage_name),
                "count": count_value,
            }
        )

    fab_stage_set = set(FAB_STAGES)
    for stage_name, count_value in stage_counts_dict.items():
        if not stage_name or stage_name in fab_stage_set:
            continue
        if int(count_value or 0) <= 0:
            continue
        stage_breakdown_rows.append(
            {
                "stage": stage_name,
                "count": int(count_value or 0),
            }
        )

    stage_breakdown = sorted(
        stage_breakdown_rows,
        key=lambda row: (-int(row.get("count", 0)), str(row.get("stage", "")).lower()),
    )

    completion_rate = round((completed_installs / total_fabs) * 100, 2) if total_fabs else 0.0
    gp_margin = round((_to_float(total_gp) / _to_float(total_revenue)) * 100, 2) if _to_float(total_revenue) else 0.0

    return success_response(
        {
            "period": {
                "start_date": start_date.isoformat() if start_date else None,
                "end_date": end_date.isoformat() if end_date else None,
            },
            "kpis": {
                "total_jobs": total_jobs,
                "total_fabs": total_fabs,
                "active_fabs": active_fabs,
                "pending_installs": pending_installs,
                "completed_installs": completed_installs,
                "completion_rate_percent": completion_rate,
                "total_revenue": round(_to_float(total_revenue), 2),
                "gross_profit": round(_to_float(total_gp), 2),
                "gross_margin_percent": gp_margin,
            },
            "stage_breakdown": stage_breakdown,
        },
        "Owner overview report generated",
    )


@router.get("/reports/owner/redo-analysis", response_model=SuccessResponse[dict])
async def get_owner_redo_analysis_report(
    start_date: Optional[date] = Query(None, description="Inclusive start date filter"),
    end_date: Optional[date] = Query(None, description="Inclusive end date filter"),
    month: Optional[int] = Query(None, ge=1, le=12, description="Optional month filter"),
    year: Optional[int] = Query(None, ge=2000, le=2100, description="Optional year for month filter"),
    top_n: int = Query(10, ge=1, le=50, description="Top N groups to return"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Analyze AG redo performance with monthly cut completion baseline and annual summary."""

    if month is not None or year is not None:
        effective_year = year or date.today().year
        effective_month = month or date.today().month
        start_dt, end_dt = _month_bounds(effective_year, effective_month)
        period_payload = {
            "mode": "month",
            "start_date": start_dt.date().isoformat(),
            "end_date": end_dt.date().isoformat(),
            "month": effective_month,
            "year": effective_year,
        }
    else:
        start_dt, end_dt = _range_bounds(start_date, end_date)
        if start_dt is None and end_dt is None:
            default_year = date.today().year
            default_month = date.today().month
            start_dt, end_dt = _month_bounds(default_year, default_month)
            period_payload = {
                "mode": "month",
                "start_date": start_dt.date().isoformat(),
                "end_date": end_dt.date().isoformat(),
                "month": default_month,
                "year": default_year,
            }
        else:
            period_payload = {
                "mode": "date_range",
                "start_date": start_dt.date().isoformat() if start_dt else None,
                "end_date": end_dt.date().isoformat() if end_dt else None,
                "month": None,
                "year": None,
            }

    current_month_start, current_month_end = _month_bounds(date.today().year, date.today().month)
    annual_year = year or (start_dt.year if start_dt else date.today().year)

    normalized_plan = func.lower(func.trim(PlanningSection.plan_name))
    ag_redo_filter = func.lower(func.coalesce(Fab.fab_type, "")) == "ag redo"

    # Widget 1: non-AG redo FABs where CUT + WJ are 100% using shop plan actual_end_date within selected period.
    cut_wj_completed_fabs_subquery = (
        select(ShopCutPlan.fab_id.label("fab_id"))
        .select_from(ShopCutPlan)
        .join(PlanningSection, PlanningSection.id == ShopCutPlan.planning_section_id)
        .where(
            ShopCutPlan.work_percentage >= 100,
            ShopCutPlan.actual_end_date.isnot(None),
            normalized_plan.in_(["cut", "wj"]),
            ShopCutPlan.actual_end_date >= start_dt,
            ShopCutPlan.actual_end_date <= end_dt,
        )
        .group_by(ShopCutPlan.fab_id)
        .having(func.count(func.distinct(normalized_plan)) == 2)
        .subquery("cut_wj_completed_fabs_subquery")
    )

    total_fabs_for_period = (
        await db.execute(
            select(func.count(Fab.id))
            .select_from(Fab)
            .join(cut_wj_completed_fabs_subquery, cut_wj_completed_fabs_subquery.c.fab_id == Fab.id)
            .where(~ag_redo_filter)
        )
    ).scalar() or 0

    # Current month AG redo widgets.
    current_month_ag_redo_count = (
        await db.execute(
            select(func.count(Fab.id)).where(
                ag_redo_filter,
                Fab.created_at >= current_month_start,
                Fab.created_at <= current_month_end,
            )
        )
    ).scalar() or 0

    # Keep AG redo rate behavior compatible with previous endpoint semantics.
    previous_total_fabs = (
        await db.execute(
            select(func.count(Fab.id)).where(Fab.created_at >= start_dt, Fab.created_at <= end_dt)
        )
    ).scalar() or 0
    previous_revised_fabs = (
        await db.execute(
            select(func.count(Fab.id)).where(
                Fab.revised.is_(True),
                Fab.created_at >= start_dt,
                Fab.created_at <= end_dt,
            )
        )
    ).scalar() or 0
    redo_rate = round((previous_revised_fabs / previous_total_fabs) * 100, 2) if previous_total_fabs else 0.0

    selected_period_ag_redo_rows = (
        await db.execute(
            select(
                Fab.id,
                Fab.job_id,
                Fab.redo_department,
                Fab.redo_requested_by,
                Fab.redo_total_sqft,
                Fab.cost_per_sqft,
                Fab.total_sqft,
                Fab.created_at,
                Fab.input_area,
                Fab.stone_type_id,
                Fab.stone_color_id,
                Fab.stone_thickness_id,
                Fab.edge_id,
            )
            .where(
                ag_redo_filter,
                Fab.created_at >= start_dt,
                Fab.created_at <= end_dt,
            )
            .order_by(Fab.created_at.desc(), Fab.id.desc())
        )
    ).all()

    def _redo_sqft(row) -> float:
        redo_sqft_value = _to_float(row.redo_total_sqft)
        if redo_sqft_value > 0:
            return redo_sqft_value
        return _to_float(row.total_sqft)

    def _redo_value(row) -> float:
        return _to_float(row.cost_per_sqft) * _redo_sqft(row) * 2.1

    def _redo_total_cost(row) -> float:
        return _to_float(row.cost_per_sqft) * 2.1 * _redo_sqft(row)

    total_ag_redo_sqft = round(sum(_redo_sqft(row) for row in selected_period_ag_redo_rows), 2)
    total_ag_redo_value = round(sum(_redo_value(row) for row in selected_period_ag_redo_rows), 2)

    dept_ids = sorted({int(row.redo_department) for row in selected_period_ag_redo_rows if row.redo_department is not None})
    requested_by_ids = sorted({int(row.redo_requested_by) for row in selected_period_ag_redo_rows if row.redo_requested_by is not None})
    job_ids = sorted({int(row.job_id) for row in selected_period_ag_redo_rows if row.job_id is not None})
    stone_type_ids = sorted({int(row.stone_type_id) for row in selected_period_ag_redo_rows if row.stone_type_id is not None})
    stone_color_ids = sorted({int(row.stone_color_id) for row in selected_period_ag_redo_rows if row.stone_color_id is not None})
    stone_thickness_ids = sorted({int(row.stone_thickness_id) for row in selected_period_ag_redo_rows if row.stone_thickness_id is not None})
    edge_ids = sorted({int(row.edge_id) for row in selected_period_ag_redo_rows if row.edge_id is not None})

    department_name_map: dict[int, str] = {}
    if dept_ids:
        dept_rows = (await db.execute(select(Department.id, Department.name).where(Department.id.in_(dept_ids)))).all()
        department_name_map = {row[0]: row[1] for row in dept_rows}

    employee_name_map: dict[int, str] = {}
    if requested_by_ids:
        employee_rows = (await db.execute(select(User.id, User.first_name, User.last_name).where(User.id.in_(requested_by_ids)))).all()
        employee_name_map = {
            row[0]: (f"{(row[1] or '').strip()} {(row[2] or '').strip()}".strip() or f"User {row[0]}")
            for row in employee_rows
        }

    job_account_map: dict[int, dict] = {}
    if job_ids:
        job_rows = (
            await db.execute(
                select(
                    BusinessJob.id,
                    BusinessJob.job_number,
                    BusinessJob.name,
                    Account.id,
                    Account.name,
                )
                .select_from(BusinessJob)
                .join(Account, Account.id == BusinessJob.account_id, isouter=True)
                .where(BusinessJob.id.in_(job_ids))
            )
        ).all()
        job_account_map = {
            row[0]: {
                "job_number": row[1],
                "job_name": row[2],
                "account_id": row[3],
                "account_name": row[4],
            }
            for row in job_rows
        }

    stone_type_map: dict[int, str] = {}
    if stone_type_ids:
        stone_type_rows = (await db.execute(select(StoneType.id, StoneType.name).where(StoneType.id.in_(stone_type_ids)))).all()
        stone_type_map = {row[0]: row[1] for row in stone_type_rows}

    stone_color_map: dict[int, str] = {}
    if stone_color_ids:
        stone_color_rows = (await db.execute(select(StoneColor.id, StoneColor.name).where(StoneColor.id.in_(stone_color_ids)))).all()
        stone_color_map = {row[0]: row[1] for row in stone_color_rows}

    stone_thickness_map: dict[int, str] = {}
    if stone_thickness_ids:
        stone_thickness_rows = (await db.execute(select(StoneThickness.id, StoneThickness.thickness).where(StoneThickness.id.in_(stone_thickness_ids)))).all()
        stone_thickness_map = {row[0]: row[1] for row in stone_thickness_rows}

    edge_map: dict[int, str] = {}
    if edge_ids:
        edge_rows = (await db.execute(select(Edge.id, Edge.name).where(Edge.id.in_(edge_ids)))).all()
        edge_map = {row[0]: row[1] for row in edge_rows}

    redo_by_department_map: dict[int, dict] = {}
    redo_by_employee_map: dict[int, dict] = {}
    redo_total_cost_rows: list[dict] = []
    top_accounts_map: dict[int, dict] = {}
    top_jobs_map: dict[int, dict] = {}

    for row in selected_period_ag_redo_rows:
        department_id = int(row.redo_department) if row.redo_department is not None else None
        employee_id = int(row.redo_requested_by) if row.redo_requested_by is not None else None
        redo_sqft_value = _redo_sqft(row)
        redo_value = _redo_value(row)
        redo_total_cost_value = _redo_total_cost(row)
        job_meta = job_account_map.get(int(row.job_id) if row.job_id is not None else -1, {})

        if department_id is not None:
            dept_bucket = redo_by_department_map.setdefault(
                department_id,
                {
                    "department_id": department_id,
                    "department_name": department_name_map.get(department_id, f"Department {department_id}"),
                    "redo_count": 0,
                    "redo_sqft": 0.0,
                    "redo_value": 0.0,
                    "redo_total_cost": 0.0,
                },
            )
            dept_bucket["redo_count"] += 1
            dept_bucket["redo_sqft"] += redo_sqft_value
            dept_bucket["redo_value"] += redo_value
            dept_bucket["redo_total_cost"] += redo_total_cost_value

        if employee_id is not None:
            employee_bucket = redo_by_employee_map.setdefault(
                employee_id,
                {
                    "employee_id": employee_id,
                    "employee_name": employee_name_map.get(employee_id, f"User {employee_id}"),
                    "redo_count": 0,
                    "redo_sqft": 0.0,
                    "redo_value": 0.0,
                    "redo_total_cost": 0.0,
                },
            )
            employee_bucket["redo_count"] += 1
            employee_bucket["redo_sqft"] += redo_sqft_value
            employee_bucket["redo_value"] += redo_value
            employee_bucket["redo_total_cost"] += redo_total_cost_value

        account_id = job_meta.get("account_id")
        if account_id is not None:
            account_bucket = top_accounts_map.setdefault(
                int(account_id),
                {
                    "account_id": int(account_id),
                    "account_name": job_meta.get("account_name") or "Unassigned Account",
                    "redo_count": 0,
                },
            )
            account_bucket["redo_count"] += 1

        if row.job_id is not None:
            job_bucket = top_jobs_map.setdefault(
                int(row.job_id),
                {
                    "job_id": int(row.job_id),
                    "job_number": job_meta.get("job_number"),
                    "job_name": job_meta.get("job_name"),
                    "redo_count": 0,
                },
            )
            job_bucket["redo_count"] += 1

        redo_total_cost_rows.append(
            {
                "fab_id": row.id,
                "job_id": row.job_id,
                "job_number": job_meta.get("job_number"),
                "job_name": job_meta.get("job_name"),
                "account_id": job_meta.get("account_id"),
                "account_name": job_meta.get("account_name"),
                "input_area": row.input_area,
                "stone_type_name": stone_type_map.get(int(row.stone_type_id), None) if row.stone_type_id else None,
                "stone_color_name": stone_color_map.get(int(row.stone_color_id), None) if row.stone_color_id else None,
                "stone_thickness_value": stone_thickness_map.get(int(row.stone_thickness_id), None) if row.stone_thickness_id else None,
                "edge_name": edge_map.get(int(row.edge_id), None) if row.edge_id else None,
                "cost_per_sqft": round(_to_float(row.cost_per_sqft), 2),
                "redo_total_sqft": round(redo_sqft_value, 2),
                "redo_total_cost": round(redo_total_cost_value, 2),
                "created_at": row.created_at.isoformat() if row.created_at else None,
            }
        )

    redo_by_department = sorted(
        [
            {
                **item,
                "redo_sqft": round(item["redo_sqft"], 2),
                "redo_value": round(item["redo_value"], 2),
                "redo_total_cost": round(item["redo_total_cost"], 2),
            }
            for item in redo_by_department_map.values()
        ],
        key=lambda x: (-x["redo_count"], (x["department_name"] or "")),
    )

    redo_by_employee = sorted(
        [
            {
                **item,
                "redo_sqft": round(item["redo_sqft"], 2),
                "redo_value": round(item["redo_value"], 2),
                "redo_total_cost": round(item["redo_total_cost"], 2),
            }
            for item in redo_by_employee_map.values()
        ],
        key=lambda x: (-x["redo_count"], (x["employee_name"] or "")),
    )

    redo_total_cost_rows = sorted(redo_total_cost_rows, key=lambda x: x["redo_total_cost"], reverse=True)
    top_accounts_with_redo = sorted(top_accounts_map.values(), key=lambda x: x["redo_count"], reverse=True)[:top_n]
    top_jobs_with_redo = sorted(top_jobs_map.values(), key=lambda x: x["redo_count"], reverse=True)[:top_n]

    # Keep stage breakdown for compatibility; scoped to AG redo rows.
    stage_redo_query = (
        select(Fab.current_stage, func.count(Fab.id).label("redo_count"))
        .where(
            ag_redo_filter,
            Fab.created_at >= start_dt,
            Fab.created_at <= end_dt,
        )
        .group_by(Fab.current_stage)
        .order_by(func.count(Fab.id).desc())
    )
    stage_redo = [
        {"stage": row[0] or "unknown", "redo_count": int(row[1] or 0)}
        for row in (await db.execute(stage_redo_query)).all()
    ]

    # Annual summary by month.
    month_bucket_expr = func.date_trunc(literal_column("'month'"), ShopCutPlan.actual_end_date)
    annual_cut_wj_rows = (
        await db.execute(
            select(
                month_bucket_expr.label("month_bucket"),
                ShopCutPlan.fab_id,
            )
            .select_from(ShopCutPlan)
            .join(PlanningSection, PlanningSection.id == ShopCutPlan.planning_section_id)
            .where(
                ShopCutPlan.work_percentage >= 100,
                ShopCutPlan.actual_end_date.isnot(None),
                normalized_plan.in_(["cut", "wj"]),
                func.extract("year", ShopCutPlan.actual_end_date) == annual_year,
            )
            .group_by(month_bucket_expr, ShopCutPlan.fab_id)
            .having(func.count(func.distinct(normalized_plan)) == 2)
        )
    ).all()

    annual_fab_count_by_month: dict[int, int] = defaultdict(int)
    for month_bucket, _fab_id in annual_cut_wj_rows:
        month_number = month_bucket.month if month_bucket else None
        if month_number is not None:
            annual_fab_count_by_month[month_number] += 1

    annual_redo_rows = (
        await db.execute(
            select(
                func.extract("month", Fab.created_at).label("month_no"),
                func.count(Fab.id).label("redo_count"),
                func.coalesce(func.sum(Fab.redo_total_sqft), 0).label("redo_sqft"),
                func.coalesce(func.sum(Fab.cost_per_sqft * Fab.redo_total_sqft), 0).label("redo_value"),
            )
            .where(
                ag_redo_filter,
                func.extract("year", Fab.created_at) == annual_year,
            )
            .group_by(func.extract("month", Fab.created_at))
        )
    ).all()

    annual_redo_map = {
        int(row.month_no): {
            "redo_count": int(row.redo_count or 0),
            "redo_sqft": round(_to_float(row.redo_sqft), 2),
            "redo_value": round(_to_float(row.redo_value), 2),
        }
        for row in annual_redo_rows
    }

    annual_summary = []
    previous_redo_count = None
    previous_redo_value = None
    for month_no in range(1, 13):
        month_name = calendar.month_name[month_no]
        total_fabs_month = int(annual_fab_count_by_month.get(month_no, 0))
        redo_stats = annual_redo_map.get(month_no, {"redo_count": 0, "redo_sqft": 0.0, "redo_value": 0.0})
        redo_count = int(redo_stats["redo_count"])
        redo_sqft = round(_to_float(redo_stats["redo_sqft"]), 2)
        redo_value = round(_to_float(redo_stats["redo_value"]), 2)

        change_in_redos_value = 0 if previous_redo_count is None else (redo_count - previous_redo_count)
        increase_decrease_value = 0.0 if previous_redo_value is None else round(redo_value - previous_redo_value, 2)
        redo_percent_value = round((redo_count / total_fabs_month) * 100, 2) if total_fabs_month else 0.0

        annual_summary.append(
            {
                "month": month_name,
                "month_number": month_no,
                "total_number_of_fabs": total_fabs_month,
                "total_number_of_ag_redo_fabs": redo_count,
                "change_in_number_of_redos_value": change_in_redos_value,
                "redo_percent_value": redo_percent_value,
                "total_square_footage": redo_sqft,
                "total_redo_value": redo_value,
                "increase_decrease_value": increase_decrease_value,
            }
        )

        previous_redo_count = redo_count
        previous_redo_value = redo_value

    summary = {
        "total_fabs": total_fabs_for_period,
        "ag_redo_fabs_current_month": current_month_ag_redo_count,
        "ag_redo_fabs_period": len(selected_period_ag_redo_rows),
        "redo_rate_percent": redo_rate,
        "total_ag_redo_sq_ft": total_ag_redo_sqft,
        "total_ag_redo_dollar_value": total_ag_redo_value,
        "redo_total_cost": round(sum(row["redo_total_cost"] for row in redo_total_cost_rows), 2),
    }

    return success_response(
        {
            "period": period_payload,
            "summary": summary,
            "redo_by_stage": stage_redo,
            "redo_by_department": redo_by_department,
            "redo_by_employee": redo_by_employee,
            "redo_total_cost_rows": redo_total_cost_rows,
            "top_accounts_with_redo": top_accounts_with_redo,
            "top_jobs_with_redo": top_jobs_with_redo,
            "redo_annual_summary": annual_summary,
        },
        "Owner redo analysis report generated",
    )


@router.get("/reports/owner/shop-status", response_model=SuccessResponse[dict])
async def get_owner_shop_status_report(
    start_date: Optional[date] = Query(None, description="Inclusive start date filter"),
    end_date: Optional[date] = Query(None, description="Inclusive end date filter"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Current shop load by stage with aging and stalled-work indicators."""
    start_dt, end_dt = _range_bounds(start_date, end_date)

    filters = []
    _apply_datetime_filters(filters, Fab.created_at, start_dt, end_dt)

    def _normalize_stage_name(value: Optional[str]) -> str:
        text = (value or "").strip().lower().replace("_", " ").replace("-", " ")
        return " ".join(text.split())

    excluded_stages = {
        "slabsmith",
        "predraft review",
        "predraftreview",
        "cutlist",
        "cut list",
    }
    allowed_shop_plan_stages = {
        "cnc": "cnc",
        "cut": "cut",
        "wj": "wj",
        "miter": "miter",
        "edging": "edging",
        "resurfacing": "resurfacing",
        "touch up": "touch up",
        "touchup": "touch up",
    }

    now_dt = datetime.now()
    stage_fab_map: dict[str, dict[int, dict]] = defaultdict(dict)

    non_shop_query = (
        select(
            Fab.id,
            Fab.current_stage,
            Fab.total_sqft,
            Fab.saw_cut_lnft,
            Fab.updated_at,
            Fab.created_at,
        )
        .select_from(Fab)
        .where(Fab.current_stage != "shop")
    )
    if filters:
        non_shop_query = non_shop_query.where(and_(*filters))

    non_shop_rows = (await db.execute(non_shop_query)).all()
    for fab_id, current_stage, total_sqft, saw_cut_lnft, updated_at, created_at in non_shop_rows:
        normalized_stage = _normalize_stage_name(current_stage)
        if normalized_stage in excluded_stages:
            continue

        stage_key = normalized_stage or "unknown"
        reference_dt = updated_at or created_at
        age_days = max((now_dt.date() - reference_dt.date()).days, 0) if reference_dt else 0
        stage_fab_map[stage_key][int(fab_id)] = {
            "age_days": age_days,
            "sqft": _to_float(total_sqft),
            "linear_ft": _to_float(saw_cut_lnft),
        }

    shop_query = (
        select(
            Fab.id,
            Fab.total_sqft,
            Fab.saw_cut_lnft,
            Fab.updated_at,
            Fab.created_at,
            ShopCutPlan.updated_at,
            ShopCutPlan.actual_end_date,
            ShopCutPlan.actual_start_date,
            PlanningSection.plan_name,
        )
        .select_from(Fab)
        .join(ShopCutPlan, ShopCutPlan.fab_id == Fab.id)
        .join(PlanningSection, PlanningSection.id == ShopCutPlan.planning_section_id, isouter=True)
        .where(Fab.current_stage == "shop")
    )
    if filters:
        shop_query = shop_query.where(and_(*filters))

    shop_rows = (await db.execute(shop_query)).all()
    for (
        fab_id,
        total_sqft,
        saw_cut_lnft,
        fab_updated_at,
        fab_created_at,
        plan_updated_at,
        plan_actual_end_at,
        plan_actual_start_at,
        plan_name,
    ) in shop_rows:
        normalized_plan_name = _normalize_stage_name(plan_name)
        if normalized_plan_name in excluded_stages:
            continue
        mapped_stage = allowed_shop_plan_stages.get(normalized_plan_name)
        if mapped_stage is None:
            continue

        reference_dt = plan_updated_at or plan_actual_end_at or plan_actual_start_at or fab_updated_at or fab_created_at
        age_days = max((now_dt.date() - reference_dt.date()).days, 0) if reference_dt else 0

        existing = stage_fab_map[mapped_stage].get(int(fab_id))
        if existing is None:
            stage_fab_map[mapped_stage][int(fab_id)] = {
                "age_days": age_days,
                "sqft": _to_float(total_sqft),
                "linear_ft": _to_float(saw_cut_lnft),
            }
        else:
            existing["age_days"] = max(existing["age_days"], age_days)
            existing["sqft"] = max(existing["sqft"], _to_float(total_sqft))
            existing["linear_ft"] = max(existing["linear_ft"], _to_float(saw_cut_lnft))

    stage_status = []
    for stage_name, fab_entries in stage_fab_map.items():
        fab_values = list(fab_entries.values())
        fab_count = len(fab_values)
        if fab_count == 0:
            continue

        age_days_values = [int(item["age_days"]) for item in fab_values]
        stalled_count = sum(1 for age_days in age_days_values if age_days > 14)
        total_sqft_value = round(sum(_to_float(item["sqft"]) for item in fab_values), 2)
        total_linear_ft_value = round(sum(_to_float(item["linear_ft"]) for item in fab_values), 2)

        stage_status.append(
            {
                "stage": stage_name,
                "fab_count": fab_count,
                "avg_age_days": round(sum(age_days_values) / fab_count, 2),
                "max_age_days": float(max(age_days_values)),
                "stalled_over_14_days": int(stalled_count),
                "total_sqft": total_sqft_value,
                "total_linear_ft": total_linear_ft_value,
            }
        )

    stage_status.sort(key=lambda row: (row["fab_count"], row["total_sqft"], row["total_linear_ft"]), reverse=True)

    unique_fab_ids = {fab_id for per_stage in stage_fab_map.values() for fab_id in per_stage.keys()}

    summary = {
        "stage_count": len(stage_status),
        "total_fabs": len(unique_fab_ids),
        "total_stalled_over_14_days": int(sum(int(row["stalled_over_14_days"]) for row in stage_status)),
        "total_sqft": round(sum(_to_float(row["total_sqft"]) for row in stage_status), 2),
        "total_linear_ft": round(sum(_to_float(row["total_linear_ft"]) for row in stage_status), 2),
    }

    return success_response(
        {
            "period": {
                "start_date": start_date.isoformat() if start_date else None,
                "end_date": end_date.isoformat() if end_date else None,
            },
            "summary": summary,
        },
        "Owner shop status report generated",
    )


@router.get("/reports/owner/revision-report", response_model=SuccessResponse[dict])
async def get_owner_revision_report(
    start_date: Optional[date] = Query(None, description="Inclusive start date filter"),
    end_date: Optional[date] = Query(None, description="Inclusive end date filter"),
    month: Optional[int] = Query(None, ge=1, le=12, description="Optional month filter when date range is not provided"),
    year: Optional[int] = Query(None, ge=2000, le=2100, description="Optional year filter when month filter is used"),
    revision_type: Optional[str] = Query(None, description="Optional revision type filter (sales, client, cad, template, shop)"),
    fab_type: Optional[str] = Query(None, description="Optional FAB type filter"),
    job_name: Optional[str] = Query(None, description="Optional job name filter"),
    account_name: Optional[str] = Query(None, description="Optional account name filter"),
    top_n: int = Query(10, ge=1, le=100, description="Top N rows for account/job client-revision ranking"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Combined Sales CT and Shop revision report with active FAB filtering and client revision hotspots."""
    _ = current_user

    if start_date is not None or end_date is not None:
        effective_start = start_date or end_date
        effective_end = end_date or start_date
        start_dt, end_dt = _range_bounds(effective_start, effective_end)
        period_payload = {
            "mode": "date_range",
            "start_date": effective_start.isoformat() if effective_start else None,
            "end_date": effective_end.isoformat() if effective_end else None,
            "month": None,
            "year": None,
        }
    elif month is not None or year is not None:
        effective_year = year or date.today().year
        effective_month = month or date.today().month
        start_dt, end_dt = _month_bounds(effective_year, effective_month)
        period_payload = {
            "mode": "month",
            "start_date": start_dt.date().isoformat(),
            "end_date": end_dt.date().isoformat(),
            "month": effective_month,
            "year": effective_year,
        }
    else:
        start_dt, end_dt = None, None
        period_payload = {
            "mode": "all_time",
            "start_date": None,
            "end_date": None,
            "month": None,
            "year": None,
        }

    templating_exists = (
        select(Templating.id)
        .where(Templating.fab_id == Fab.id)
        .limit(1)
        .exists()
    )
    active_fab_condition = and_(
        Fab.status_id == 1,
        or_(
            Fab.template_needed.is_(False),
            and_(Fab.template_needed.is_(True), templating_exists),
        ),
    )

    normalized_revision_type = (revision_type or "").strip().lower()
    valid_revision_types = {"sales", "client", "cad", "template", "shop"}
    if normalized_revision_type and normalized_revision_type not in valid_revision_types:
        return success_response(
            None,
            "Invalid revision_type. Use one of: sales, client, cad, template, shop",
            status_code=400,
        )

    revision_query = (
        select(
            Revision.id.label("revision_id"),
            Revision.fab_id,
            Revision.revision_type,
            Revision.revision_reason,
            Revision.revision_notes,
            Revision.requested_by,
            Revision.assigned_to,
            Revision.created_at,
            BusinessJob.id.label("job_id"),
            BusinessJob.job_number,
            BusinessJob.name.label("job_name"),
            Account.id.label("account_id"),
            Account.name.label("account_name"),
            Fab.fab_type,
            Fab.input_area,
            StoneType.name.label("stone_type_name"),
            StoneColor.name.label("stone_color_name"),
            StoneThickness.thickness.label("stone_thickness_value"),
            Edge.name.label("edge_name"),
        )
        .select_from(Revision)
        .join(Fab, Fab.id == Revision.fab_id)
        .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
        .join(Account, Account.id == BusinessJob.account_id, isouter=True)
        .join(StoneType, StoneType.id == Fab.stone_type_id, isouter=True)
        .join(StoneColor, StoneColor.id == Fab.stone_color_id, isouter=True)
        .join(StoneThickness, StoneThickness.id == Fab.stone_thickness_id, isouter=True)
        .join(Edge, Edge.id == Fab.edge_id, isouter=True)
        .where(active_fab_condition)
        .order_by(Revision.created_at.desc(), Revision.id.desc())
    )
    if start_dt is not None:
        revision_query = revision_query.where(Revision.created_at >= start_dt)
    if end_dt is not None:
        revision_query = revision_query.where(Revision.created_at <= end_dt)
    if normalized_revision_type and normalized_revision_type != "shop":
        revision_query = revision_query.where(func.lower(func.coalesce(Revision.revision_type, "")) == normalized_revision_type)
    elif normalized_revision_type == "shop":
        revision_query = revision_query.where(literal_column("1=0"))
    if fab_type:
        revision_query = revision_query.where(func.lower(Fab.fab_type) == fab_type.strip().lower())
    if job_name:
        revision_query = revision_query.where(BusinessJob.name.ilike(f"%{job_name.strip()}%"))
    if account_name:
        revision_query = revision_query.where(Account.name.ilike(f"%{account_name.strip()}%"))
    revision_rows = (await db.execute(revision_query)).all()

    shop_revision_query = (
        select(
            ShopRevision.id.label("shop_revision_id"),
            ShopRevision.fab_id,
            ShopRevision.revision_note,
            ShopRevision.revision_feedback,
            ShopRevision.requested_by,
            ShopRevision.assigned_to,
            ShopRevision.created_at,
            ShopRevision.revision_completed,
            ShopRevision.completed_at,
            BusinessJob.id.label("job_id"),
            BusinessJob.job_number,
            BusinessJob.name.label("job_name"),
            Account.id.label("account_id"),
            Account.name.label("account_name"),
            Fab.fab_type,
            Fab.input_area,
            StoneType.name.label("stone_type_name"),
            StoneColor.name.label("stone_color_name"),
            StoneThickness.thickness.label("stone_thickness_value"),
            Edge.name.label("edge_name"),
        )
        .select_from(ShopRevision)
        .join(Fab, Fab.id == ShopRevision.fab_id)
        .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
        .join(Account, Account.id == BusinessJob.account_id, isouter=True)
        .join(StoneType, StoneType.id == Fab.stone_type_id, isouter=True)
        .join(StoneColor, StoneColor.id == Fab.stone_color_id, isouter=True)
        .join(StoneThickness, StoneThickness.id == Fab.stone_thickness_id, isouter=True)
        .join(Edge, Edge.id == Fab.edge_id, isouter=True)
        .where(active_fab_condition)
        .order_by(ShopRevision.created_at.desc(), ShopRevision.id.desc())
    )
    if start_dt is not None:
        shop_revision_query = shop_revision_query.where(ShopRevision.created_at >= start_dt)
    if end_dt is not None:
        shop_revision_query = shop_revision_query.where(ShopRevision.created_at <= end_dt)
    if normalized_revision_type and normalized_revision_type != "shop":
        shop_revision_query = shop_revision_query.where(literal_column("1=0"))
    if fab_type:
        shop_revision_query = shop_revision_query.where(func.lower(Fab.fab_type) == fab_type.strip().lower())
    if job_name:
        shop_revision_query = shop_revision_query.where(BusinessJob.name.ilike(f"%{job_name.strip()}%"))
    if account_name:
        shop_revision_query = shop_revision_query.where(Account.name.ilike(f"%{account_name.strip()}%"))
    shop_revision_rows = (await db.execute(shop_revision_query)).all()

    user_ids: set[int] = set()
    for row in revision_rows:
        if row.requested_by is not None:
            user_ids.add(int(row.requested_by))
        if row.assigned_to is not None:
            user_ids.add(int(row.assigned_to))
    for row in shop_revision_rows:
        if row.requested_by is not None:
            user_ids.add(int(row.requested_by))
        if row.assigned_to is not None:
            user_ids.add(int(row.assigned_to))

    user_name_map: dict[int, str] = {}
    if user_ids:
        user_rows = (
            await db.execute(
                select(User.id, User.first_name, User.last_name).where(User.id.in_(list(user_ids)))
            )
        ).all()
        user_name_map = {
            row[0]: (f"{(row[1] or '').strip()} {(row[2] or '').strip()}".strip() or f"User {row[0]}")
            for row in user_rows
        }

    revision_type_map = {
        "sales": "Sales",
        "client": "Client",
        "cad": "CAD",
        "template": "Template",
    }

    sales_ct_revisions = []
    active_sales_fab_ids: set[int] = set()
    for row in revision_rows:
        normalized_type = (row.revision_type or "").strip().lower()
        mapped_type = revision_type_map.get(normalized_type, (row.revision_type or "Unknown").strip() or "Unknown")
        active_sales_fab_ids.add(int(row.fab_id))
        sales_ct_revisions.append(
            {
                "revision_id": row.revision_id,
                "fab_id": row.fab_id,
                "revision_type": normalized_type or None,
                "revision_type_label": mapped_type,
                "revision_reason": row.revision_reason,
                "revision_notes": row.revision_notes,
                "requested_by": row.requested_by,
                "requested_by_name": user_name_map.get(row.requested_by, f"User {row.requested_by}" if row.requested_by else None),
                "assigned_to": row.assigned_to,
                "assigned_to_name": user_name_map.get(row.assigned_to, f"User {row.assigned_to}" if row.assigned_to else None),
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "job_id": row.job_id,
                "job_number": row.job_number,
                "job_name": row.job_name,
                "account_id": row.account_id,
                "account_name": row.account_name,
                "fab_type": row.fab_type,
                "input_area": row.input_area,
                "stone_type_name": row.stone_type_name,
                "stone_color_name": row.stone_color_name,
                "stone_thickness_value": row.stone_thickness_value,
                "edge_name": row.edge_name,
            }
        )

    shop_revisions = []
    active_shop_fab_ids: set[int] = set()
    for row in shop_revision_rows:
        active_shop_fab_ids.add(int(row.fab_id))
        shop_revisions.append(
            {
                "shop_revision_id": row.shop_revision_id,
                "fab_id": row.fab_id,
                "revision_type": "shop",
                "revision_type_label": "Shop",
                "revision_notes": row.revision_note,
                "revision_feedback": row.revision_feedback,
                "requested_by": row.requested_by,
                "requested_by_name": user_name_map.get(row.requested_by, f"User {row.requested_by}" if row.requested_by else None),
                "assigned_to": row.assigned_to,
                "assigned_to_name": user_name_map.get(row.assigned_to, f"User {row.assigned_to}" if row.assigned_to else None),
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "revision_completed": bool(row.revision_completed),
                "completed_at": row.completed_at.isoformat() if row.completed_at else None,
                "job_id": row.job_id,
                "job_number": row.job_number,
                "job_name": row.job_name,
                "account_id": row.account_id,
                "account_name": row.account_name,
                "fab_type": row.fab_type,
                "input_area": row.input_area,
                "stone_type_name": row.stone_type_name,
                "stone_color_name": row.stone_color_name,
                "stone_thickness_value": row.stone_thickness_value,
                "edge_name": row.edge_name,
            }
        )

    client_type_filter = func.lower(func.coalesce(Revision.revision_type, "")) == "client"

    top_accounts_with_client_revisions = []
    if normalized_revision_type in {"", "client"}:
        top_accounts_query = (
            select(
                Account.id,
                Account.name,
                func.count(Revision.id).label("client_revision_count"),
            )
            .select_from(Revision)
            .join(Fab, Fab.id == Revision.fab_id)
            .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
            .join(Account, Account.id == BusinessJob.account_id, isouter=True)
            .where(active_fab_condition, client_type_filter)
            .group_by(Account.id, Account.name)
            .order_by(func.count(Revision.id).desc(), Account.name.asc())
            .limit(top_n)
        )
        if start_dt is not None:
            top_accounts_query = top_accounts_query.where(Revision.created_at >= start_dt)
        if end_dt is not None:
            top_accounts_query = top_accounts_query.where(Revision.created_at <= end_dt)
        if fab_type:
            top_accounts_query = top_accounts_query.where(func.lower(Fab.fab_type) == fab_type.strip().lower())
        if job_name:
            top_accounts_query = top_accounts_query.where(BusinessJob.name.ilike(f"%{job_name.strip()}%"))
        if account_name:
            top_accounts_query = top_accounts_query.where(Account.name.ilike(f"%{account_name.strip()}%"))
        top_accounts_with_client_revisions = [
            {
                "account_id": row[0],
                "account_name": row[1] or "Unassigned Account",
                "client_revision_count": int(row[2] or 0),
            }
            for row in (await db.execute(top_accounts_query)).all()
        ]

    top_jobs_with_client_revisions = []
    if normalized_revision_type in {"", "client"}:
        top_jobs_query = (
            select(
                BusinessJob.id,
                BusinessJob.job_number,
                BusinessJob.name,
                func.count(Revision.id).label("client_revision_count"),
            )
            .select_from(Revision)
            .join(Fab, Fab.id == Revision.fab_id)
            .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
            .where(active_fab_condition, client_type_filter)
            .group_by(BusinessJob.id, BusinessJob.job_number, BusinessJob.name)
            .order_by(func.count(Revision.id).desc(), BusinessJob.job_number.asc())
            .limit(top_n)
        )
        if start_dt is not None:
            top_jobs_query = top_jobs_query.where(Revision.created_at >= start_dt)
        if end_dt is not None:
            top_jobs_query = top_jobs_query.where(Revision.created_at <= end_dt)
        if fab_type:
            top_jobs_query = top_jobs_query.where(func.lower(Fab.fab_type) == fab_type.strip().lower())
        if job_name:
            top_jobs_query = top_jobs_query.where(BusinessJob.name.ilike(f"%{job_name.strip()}%"))
        if account_name:
            top_jobs_query = top_jobs_query.where(Account.name.ilike(f"%{account_name.strip()}%"))
        top_jobs_with_client_revisions = [
            {
                "job_id": row[0],
                "job_number": row[1],
                "job_name": row[2],
                "client_revision_count": int(row[3] or 0),
            }
            for row in (await db.execute(top_jobs_query)).all()
        ]

    return success_response(
        {
            "title": "Revision Report",
            "period": period_payload,
            "filters": {
                "revision_type": revision_type,
                "fab_type": fab_type,
                "job_name": job_name,
                "account_name": account_name,
            },
            "summary": {
                "sales_ct_revision_count": len(sales_ct_revisions),
                "shop_revision_count": len(shop_revisions),
                "active_fabs_with_sales_ct_revisions": len(active_sales_fab_ids),
                "active_fabs_with_shop_revisions": len(active_shop_fab_ids),
            },
            "sales_ct_revisions": sales_ct_revisions,
            "shop_revisions": shop_revisions,
            "top_accounts_with_client_revisions": top_accounts_with_client_revisions,
            "top_jobs_with_client_revisions": top_jobs_with_client_revisions,
            "employee_source": "users",
        },
        "Owner revision report generated",
    )


def _normalize_shop_stage_name(plan_name: Optional[str]) -> str:
    if not isinstance(plan_name, str):
        return "unplanned"
    normalized = plan_name.strip().lower()
    return normalized if normalized else "unplanned"


async def _get_shop_production_stage_counts(
    db: AsyncSession,
    start_date: Optional[date],
    end_date: Optional[date],
    status_id: int,
    include_non_shop_stages: bool,
) -> dict:
    filters = [Fab.status_id == status_id]
    edge_activity_names = {"edge", "edging"}

    if not include_non_shop_stages:
        filters.append(Fab.current_stage.in_(["shop", "cut_list"]))

    if start_date is not None:
        filters.append(Fab.shop_date_schedule >= start_date)
    if end_date is not None:
        filters.append(Fab.shop_date_schedule <= end_date)

    rows = (
        await db.execute(
            select(
                Fab.id,
                Fab.current_stage,
                Fab.fab_type,
                Fab.total_sqft,
                Fab.no_of_pieces,
                Fab.shop_date_schedule,
                Fab.updated_at,
                BusinessJob.job_number,
                BusinessJob.name,
                ShopCutPlan.sequence,
                ShopCutPlan.work_percentage,
                PlanningSection.plan_name,
            )
            .select_from(Fab)
            .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
            .join(ShopCutPlan, ShopCutPlan.fab_id == Fab.id, isouter=True)
            .join(PlanningSection, PlanningSection.id == ShopCutPlan.planning_section_id, isouter=True)
            .where(
                and_(*filters),
                or_(
                    PlanningSection.id.is_(None),
                    func.lower(func.trim(func.coalesce(PlanningSection.plan_name, ""))).notin_(edge_activity_names),
                ),
            )
            .order_by(Fab.id.asc(), ShopCutPlan.sequence.asc(), ShopCutPlan.id.asc())
        )
    ).all()

    fabs_map: dict[int, dict] = {}
    for (
        fab_id,
        current_stage,
        fab_type,
        total_sqft,
        no_of_pieces,
        shop_date_schedule,
        updated_at,
        job_number,
        job_name,
        sequence,
        work_percentage,
        plan_name,
    ) in rows:
        if fab_id not in fabs_map:
            fabs_map[fab_id] = {
                "fab_id": fab_id,
                "job_number": job_number,
                "job_name": job_name,
                "fab_type": fab_type,
                "current_stage": current_stage,
                "total_sqft": round(_to_float(total_sqft), 2),
                "no_of_pieces": int(_to_float(no_of_pieces)),
                "shop_date_schedule": shop_date_schedule.isoformat() if shop_date_schedule else None,
                "updated_at": updated_at.isoformat() if updated_at else None,
                "plans": [],
            }

        if plan_name is not None and _normalize_shop_stage_name(plan_name) not in edge_activity_names:
            fabs_map[fab_id]["plans"].append(
                {
                    "sequence": sequence,
                    "work_percentage": work_percentage,
                    "plan_name": plan_name,
                }
            )

    stage_buckets: dict[str, dict] = {}
    total_sqft = 0.0
    total_pieces = 0
    completed_fabs = 0
    in_progress_fabs = 0
    unplanned_fabs = 0

    fab_entries = []
    now_dt = datetime.now()
    shop_stage_order = {
        "unplanned": 0,
        "cut": 1,
        "wj": 2,
        "cnc": 3,
        "miter": 4,
        "hand work": 5,
        "touch up": 6,
        "resurfacing": 7,
    }

    def _shop_stage_sort_key(stage_name: Optional[str]) -> tuple[int, str]:
        normalized_stage = _normalize_shop_stage_name(stage_name)
        return shop_stage_order.get(normalized_stage, len(shop_stage_order)), normalized_stage

    for fab_data in fabs_map.values():
        plans = fab_data["plans"]
        resolved_stage = _normalize_shop_stage_name(_get_shop_current_stage(plans))

        if resolved_stage == "unplanned":
            unplanned_fabs += 1

        plan_percentages = [
            _to_float(plan.get("work_percentage"))
            for plan in plans
            if plan.get("work_percentage") is not None
        ]
        avg_work_percentage = round(
            (sum(plan_percentages) / len(plan_percentages)) if plan_percentages else 0.0,
            2,
        )
        is_completed = bool(plans) and all(p >= 100 for p in plan_percentages) and len(plan_percentages) == len(plans)

        if is_completed:
            completed_fabs += 1
        elif plans:
            in_progress_fabs += 1

        stage_bucket = stage_buckets.setdefault(
            resolved_stage,
            {
                "shop_current_stage": resolved_stage,
                "fab_count": 0,
                "total_sqft": 0.0,
                "total_pieces": 0,
                "avg_work_percentage_sum": 0.0,
                "completed_fab_count": 0,
                "in_progress_fab_count": 0,
            },
        )

        stage_bucket["fab_count"] += 1
        stage_bucket["total_sqft"] += fab_data["total_sqft"]
        stage_bucket["total_pieces"] += fab_data["no_of_pieces"]
        stage_bucket["avg_work_percentage_sum"] += avg_work_percentage
        if is_completed:
            stage_bucket["completed_fab_count"] += 1
        elif plans:
            stage_bucket["in_progress_fab_count"] += 1

        total_sqft += fab_data["total_sqft"]
        total_pieces += fab_data["no_of_pieces"]

        updated_at_raw = fab_data.get("updated_at")
        stale_days = None
        if updated_at_raw:
            try:
                updated_dt = datetime.fromisoformat(updated_at_raw)
                stale_days = max((now_dt.date() - updated_dt.date()).days, 0)
            except Exception:
                stale_days = None

        fab_entries.append(
            {
                **fab_data,
                "shop_current_stage": resolved_stage,
                "plan_count": len(plans),
                "avg_work_percentage": avg_work_percentage,
                "is_completed": is_completed,
                "stale_days": stale_days,
            }
        )

    stage_counts = []
    for stage_name, bucket in stage_buckets.items():
        fab_count = bucket["fab_count"]
        stage_counts.append(
            {
                "shop_current_stage": stage_name,
                "fab_count": fab_count,
                "total_sqft": round(bucket["total_sqft"], 2),
                "total_pieces": int(bucket["total_pieces"]),
                "avg_work_percentage": round(
                    (bucket["avg_work_percentage_sum"] / fab_count) if fab_count else 0.0,
                    2,
                ),
                "completed_fab_count": int(bucket["completed_fab_count"]),
                "in_progress_fab_count": int(bucket["in_progress_fab_count"]),
            }
        )

    stage_counts = sorted(stage_counts, key=lambda row: _shop_stage_sort_key(row["shop_current_stage"]))
    fab_entries = sorted(
        fab_entries,
        key=lambda row: (
            _shop_stage_sort_key(row["shop_current_stage"]),
            (row["stale_days"] if row["stale_days"] is not None else -1),
            row["fab_id"],
        ),
        reverse=False,
    )

    return {
        "summary": {
            "total_fabs": len(fab_entries),
            "total_sqft": round(total_sqft, 2),
            "total_pieces": int(total_pieces),
            "completed_fabs": int(completed_fabs),
            "in_progress_fabs": int(in_progress_fabs),
            "unplanned_fabs": int(unplanned_fabs),
        },
        "shop_current_stage_counts": stage_counts,
        "fabs": fab_entries,
    }


@router.get("/reports/shop-production-summary", response_model=SuccessResponse[dict])
@router.get("/reports/owner/shop-production-summary", response_model=SuccessResponse[dict])
async def get_shop_production_summary_report(
    start_date: Optional[date] = Query(None, description="Inclusive shop date start filter (shop_date_schedule)"),
    end_date: Optional[date] = Query(None, description="Inclusive shop date end filter (shop_date_schedule)"),
    status_id: int = Query(1, ge=0, description="FAB status to include (default: active=1)"),
    include_non_shop_stages: bool = Query(
        False,
        description="When true, include all current stages; when false include only shop/cut_list FABs",
    ),
    include_fab_details: bool = Query(True, description="Include per-FAB rows in report payload"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Shop Production Summary report grouped by derived shop_current_stage from plan progress."""
    _ = current_user

    report_data = await _get_shop_production_stage_counts(
        db=db,
        start_date=start_date,
        end_date=end_date,
        status_id=status_id,
        include_non_shop_stages=include_non_shop_stages,
    )

    response_payload = {
        "period": {
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
        },
        "filters": {
            "status_id": status_id,
            "include_non_shop_stages": include_non_shop_stages,
        },
        "summary": report_data["summary"],
        "shop_current_stage_counts": report_data["shop_current_stage_counts"],
    }

    if include_fab_details:
        response_payload["fabs"] = report_data["fabs"]

    return success_response(response_payload, "Shop Production Summary report generated")


@router.get("/reports/owner/stalled-install-jobs", response_model=SuccessResponse[dict])
async def get_owner_stalled_install_jobs_report(
    start_date: Optional[date] = Query(None, description="Inclusive start date filter"),
    end_date: Optional[date] = Query(None, description="Inclusive end date filter"),
    min_age_days: int = Query(0, ge=0, le=3650, description="Minimum age in days before including a job"),
    top_n: int = Query(50, ge=1, le=200, description="Maximum stalled jobs to return"),
    include_assigned: bool = Query(True, description="Include jobs that already have an installer assigned"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return stalled install jobs with job names, assignment details, and scheduling context."""
    start_dt, end_dt = _range_bounds(start_date, end_date)

    latest_schedule_query = (
        select(
            InstallScheduling.fab_id.label("fab_id"),
            func.max(InstallScheduling.id).label("install_scheduling_id"),
        )
        .group_by(InstallScheduling.fab_id)
        .subquery()
    )

    age_days_expr = func.date_part("day", func.now() - func.coalesce(Fab.updated_at, Fab.created_at))
    due_date_expr = func.coalesce(
        InstallScheduling.scheduled_install_date,
        Fab.installation_date,
        Fab.shop_est_completion_date,
        Fab.shop_date_schedule,
    )

    query = (
        select(
            Fab.id.label("fab_id"),
            Fab.job_id,
            BusinessJob.job_number,
            BusinessJob.name.label("job_name"),
            Account.name.label("account_name"),
            Fab.current_stage,
            Fab.status_id,
            Fab.created_at,
            Fab.updated_at,
            Fab.installation_date,
            Fab.shop_est_completion_date,
            Fab.shop_date_schedule,
            Fab.total_sqft,
            InstallScheduling.installer_id,
            InstallScheduling.scheduled_install_date,
            InstallScheduling.scheduled_end_date,
            InstallScheduling.actual_install_date,
            InstallScheduling.total_sqft.label("scheduled_total_sqft"),
            User.first_name,
            User.last_name,
        )
        .select_from(Fab)
        .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
        .join(Account, Account.id == BusinessJob.account_id, isouter=True)
        .join(latest_schedule_query, latest_schedule_query.c.fab_id == Fab.id, isouter=True)
        .join(InstallScheduling, InstallScheduling.id == latest_schedule_query.c.install_scheduling_id, isouter=True)
        .join(User, User.id == InstallScheduling.installer_id, isouter=True)
        .where(Fab.current_stage.in_(["install_scheduling", "resurface_scheduling"]))
        .where(InstallScheduling.actual_install_date.is_(None))
        .order_by(age_days_expr.desc(), due_date_expr.asc(), Fab.id.asc())
    )

    if start_dt is not None:
        query = query.where(Fab.created_at >= start_dt)
    if end_dt is not None:
        query = query.where(Fab.created_at <= end_dt)
    if min_age_days:
        query = query.where(age_days_expr >= min_age_days)
    if not include_assigned:
        query = query.where(InstallScheduling.installer_id.is_(None))

    rows = (await db.execute(query)).all()

    stalled_jobs = []
    unassigned_count = 0
    overdue_count = 0
    due_today_count = 0
    now_dt = datetime.now()

    for (
        fab_id,
        job_id,
        job_number,
        job_name,
        account_name,
        current_stage,
        status_id,
        created_at,
        updated_at,
        installation_date,
        shop_est_completion_date,
        shop_date_schedule,
        total_sqft,
        installer_id,
        scheduled_install_date,
        scheduled_end_date,
        actual_install_date,
        scheduled_total_sqft,
        installer_first_name,
        installer_last_name,
    ) in rows:
        due_date = scheduled_install_date or installation_date or shop_est_completion_date or shop_date_schedule
        age_days = max((now_dt.date() - (updated_at or created_at).date()).days, 0)
        days_overdue = None
        if due_date is not None:
            days_overdue = max((now_dt.date() - due_date.date()).days, 0)
            if due_date.date() < now_dt.date() and actual_install_date is None:
                overdue_count += 1
            if due_date.date() == now_dt.date():
                due_today_count += 1

        if installer_id is None:
            unassigned_count += 1

        installer_name = (
            f"{(installer_first_name or '').strip()} {(installer_last_name or '').strip()}".strip()
            or (f"User {installer_id}" if installer_id else "Unassigned")
        )

        stalled_jobs.append(
            {
                "fab_id": fab_id,
                "job_id": job_id,
                "job_number": job_number,
                "job_name": job_name,
                "account_name": account_name or "Unassigned Account",
                "current_stage": current_stage,
                "status_id": status_id,
                "age_days": age_days,
                "days_overdue": days_overdue,
                "due_date": due_date.isoformat() if due_date else None,
                "created_at": created_at.isoformat() if created_at else None,
                "updated_at": updated_at.isoformat() if updated_at else None,
                "installation_date": installation_date.isoformat() if installation_date else None,
                "shop_est_completion_date": shop_est_completion_date.isoformat() if shop_est_completion_date else None,
                "shop_date_schedule": shop_date_schedule.isoformat() if shop_date_schedule else None,
                "scheduled_install_date": scheduled_install_date.isoformat() if scheduled_install_date else None,
                "scheduled_end_date": scheduled_end_date.isoformat() if scheduled_end_date else None,
                "actual_install_date": actual_install_date.isoformat() if actual_install_date else None,
                "total_sqft": round(_to_float(total_sqft), 2),
                "scheduled_total_sqft": round(_to_float(scheduled_total_sqft), 2),
                "installer_id": installer_id,
                "installer_name": installer_name,
            }
        )

    total_stalled_jobs = len(stalled_jobs)
    stalled_jobs = stalled_jobs[:top_n]

    return success_response(
        {
            "period": {
                "start_date": start_date.isoformat() if start_date else None,
                "end_date": end_date.isoformat() if end_date else None,
            },
            "filters": {
                "min_age_days": min_age_days,
                "top_n": top_n,
                "include_assigned": include_assigned,
            },
            "summary": {
                "stalled_job_count": total_stalled_jobs,
                "unassigned_count": unassigned_count,
                "overdue_count": overdue_count,
                "due_today_count": due_today_count,
            },
            "stalled_install_jobs": stalled_jobs,
        },
        "Owner stalled install jobs report generated",
    )


@router.get("/reports/owner/largest-jobs", response_model=SuccessResponse[dict])
async def get_owner_largest_jobs_report(
    start_date: Optional[date] = Query(None, description="Inclusive start date filter against fab creation date"),
    end_date: Optional[date] = Query(None, description="Inclusive end date filter against fab creation date"),
    top_n: int = Query(20, ge=1, le=500, description="Maximum ranked jobs to return"),
    min_sqft: float = Query(0, ge=0, description="Minimum total square footage to include"),
    order_by: str = Query("sqft", description="Ranking basis: sqft or revenue"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return jobs ranked by square footage with optional revenue sorting and sqft threshold."""
    _ = current_user
    normalized_order_by = (order_by or "sqft").strip().lower()
    if normalized_order_by not in {"sqft", "revenue"}:
        normalized_order_by = "sqft"

    start_dt, end_dt = _range_bounds(start_date, end_date)
    ranked_sqft_expr = func.greatest(
        func.coalesce(func.max(BusinessJob.sq_ft), 0.0),
        func.coalesce(func.sum(Fab.total_sqft), 0.0),
    )
    fab_sqft_expr = func.coalesce(func.sum(Fab.total_sqft), 0.0)
    revenue_expr = func.coalesce(func.sum(Fab.revenue), 0.0)
    fab_count_expr = func.count(Fab.id)
    avg_fab_sqft_expr = func.coalesce(func.avg(Fab.total_sqft), 0.0)

    query = (
        select(
            BusinessJob.id.label("job_id"),
            BusinessJob.job_number,
            BusinessJob.name.label("job_name"),
            Account.name.label("account_name"),
            func.coalesce(func.max(BusinessJob.sq_ft), 0.0).label("job_declared_sqft"),
            fab_sqft_expr.label("fab_total_sqft"),
            ranked_sqft_expr.label("ranked_sqft"),
            revenue_expr.label("total_revenue"),
            fab_count_expr.label("fab_count"),
            avg_fab_sqft_expr.label("avg_fab_sqft"),
        )
        .select_from(BusinessJob)
        .join(Account, Account.id == BusinessJob.account_id, isouter=True)
        .join(Fab, Fab.job_id == BusinessJob.id)
        .group_by(BusinessJob.id, BusinessJob.job_number, BusinessJob.name, Account.name)
    )

    if start_dt is not None:
        query = query.where(Fab.created_at >= start_dt)
    if end_dt is not None:
        query = query.where(Fab.created_at <= end_dt)
    if min_sqft > 0:
        query = query.having(ranked_sqft_expr >= min_sqft)

    if normalized_order_by == "revenue":
        query = query.order_by(revenue_expr.desc(), ranked_sqft_expr.desc(), BusinessJob.job_number.asc())
    else:
        query = query.order_by(ranked_sqft_expr.desc(), revenue_expr.desc(), BusinessJob.job_number.asc())

    rows = (await db.execute(query.limit(top_n))).all()

    ranked_rows = []
    for index, row in enumerate(rows, start=1):
        ranked_rows.append(
            {
                "rank": index,
                "job_id": row.job_id,
                "job_number": row.job_number,
                "job_name": row.job_name,
                "account_name": row.account_name,
                "total_sqft": round(_to_float(row.ranked_sqft), 2),
                "fab_total_sqft": round(_to_float(row.fab_total_sqft), 2),
                "job_declared_sqft": round(_to_float(row.job_declared_sqft), 2),
                "fab_count": int(row.fab_count or 0),
                "avg_fab_sqft": round(_to_float(row.avg_fab_sqft), 2),
                "total_revenue": round(_to_float(row.total_revenue), 2),
            }
        )

    total_ranked_sqft = round(sum(_to_float(item.get("total_sqft")) for item in ranked_rows), 2)
    total_ranked_revenue = round(sum(_to_float(item.get("total_revenue")) for item in ranked_rows), 2)

    return success_response(
        {
            "period": {
                "start_date": start_date.isoformat() if start_date else None,
                "end_date": end_date.isoformat() if end_date else None,
            },
            "filters": {
                "top_n": top_n,
                "min_sqft": round(_to_float(min_sqft), 2),
                "order_by": normalized_order_by,
            },
            "summary": {
                "row_count": len(ranked_rows),
                "total_ranked_sqft": total_ranked_sqft,
                "total_ranked_revenue": total_ranked_revenue,
            },
            "rows": ranked_rows,
        },
        "Owner largest jobs report generated",
    )


@router.get("/reports/owner/install-performance", response_model=SuccessResponse[dict])
async def get_owner_install_performance_report(
    start_date: Optional[date] = Query(None, description="Inclusive start date filter"),
    end_date: Optional[date] = Query(None, description="Inclusive end date filter"),
    installer_id: Optional[int] = Query(None, gt=0, description="Optional installer user ID filter"),
    installer_name: Optional[str] = Query(None, description="Optional installer name filter"),
    top_n: int = Query(25, ge=1, le=100, description="Top installer rows to return"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Installer-focused output and labor efficiency based on completion and timer data."""
    effective_end_date = end_date or start_date or date.today()
    effective_start_date = start_date or end_date or (effective_end_date - timedelta(days=6))
    if effective_start_date > effective_end_date:
        effective_start_date, effective_end_date = effective_end_date, effective_start_date

    start_dt, end_dt = _range_bounds(effective_start_date, effective_end_date)

    matched_installer_ids: Optional[list[int]] = None
    if installer_name:
        installer_like = f"%{installer_name.strip().lower()}%"
        matched_user_rows = (
            await db.execute(
                select(User.id)
                .where(
                    or_(
                        func.lower(User.first_name).like(installer_like),
                        func.lower(User.last_name).like(installer_like),
                        func.lower(User.username).like(installer_like),
                        func.lower(User.email).like(installer_like),
                        func.lower(
                            func.concat(
                                func.coalesce(User.first_name, ""),
                                " ",
                                func.coalesce(User.last_name, ""),
                            )
                        ).like(installer_like),
                    )
                )
                .order_by(User.id.asc())
            )
        ).all()
        matched_installer_ids = [row[0] for row in matched_user_rows]
        if not matched_installer_ids:
            return success_response(
                {
                    "period": {
                        "start_date": effective_start_date.isoformat() if effective_start_date else None,
                        "end_date": effective_end_date.isoformat() if effective_end_date else None,
                    },
                    "filters": {
                        "installer_id": installer_id,
                        "installer_name": installer_name,
                    },
                    "summary": {
                        "installer_count": 0,
                        "total_sqft_installed": 0.0,
                        "total_work_hours": 0.0,
                        "total_labor_cost": 0.0,
                        "portfolio_labor_cost_per_sqft": 0.0,
                        "portfolio_sqft_per_hour": 0.0,
                    },
                    "installer_breakdown": [],
                },
                "Owner install performance report generated",
            )

    completion_filters = []
    _apply_datetime_filters(completion_filters, InstallCompletion.completion_date, start_dt, end_dt)
    if installer_id is not None:
        completion_filters.append(InstallCompletion.installer_id == installer_id)
    if matched_installer_ids is not None:
        completion_filters.append(InstallCompletion.installer_id.in_(matched_installer_ids))

    completion_query = select(
        InstallCompletion.installer_id,
        InstallCompletion.fab_id,
        InstallCompletion.total_sqft_installed,
        InstallCompletion.completion_date,
    )
    if completion_filters:
        completion_query = completion_query.where(and_(*completion_filters))

    completion_rows = (await db.execute(completion_query)).all()

    by_installer: dict[int, dict] = defaultdict(lambda: {
        "completed_installs": 0,
        "sqft_installed": 0.0,
        "first_completion_at": None,
        "last_completion_at": None,
    })

    for installer_id, _fab_id, sqft_str, completed_at in completion_rows:
        stats = by_installer[installer_id]
        stats["completed_installs"] += 1
        stats["sqft_installed"] += _to_float(sqft_str)

        if completed_at is not None:
            if stats["first_completion_at"] is None or completed_at < stats["first_completion_at"]:
                stats["first_completion_at"] = completed_at
            if stats["last_completion_at"] is None or completed_at > stats["last_completion_at"]:
                stats["last_completion_at"] = completed_at

    timer_filters = []
    _apply_datetime_filters(timer_filters, InstallerJobTimerSession.session_start_at, start_dt, end_dt)
    if installer_id is not None:
        timer_filters.append(InstallerJobTimerSession.installer_id == installer_id)
    if matched_installer_ids is not None:
        timer_filters.append(InstallerJobTimerSession.installer_id.in_(matched_installer_ids))
    timer_query = select(
        InstallerJobTimerSession.installer_id,
        func.sum(InstallerJobTimerSession.total_work_seconds),
        func.sum(InstallerJobTimerSession.total_pause_seconds),
    ).group_by(InstallerJobTimerSession.installer_id)
    if timer_filters:
        timer_query = timer_query.where(and_(*timer_filters))

    timer_rows = (await db.execute(timer_query)).all()

    for installer_id, work_seconds, pause_seconds in timer_rows:
        stats = by_installer[installer_id]
        stats["work_hours"] = round(_to_float(work_seconds) / 3600, 2)
        stats["pause_hours"] = round(_to_float(pause_seconds) / 3600, 2)

    installer_ids = [i for i in by_installer.keys() if i is not None]
    users_map: dict[int, str] = {}
    rates_map: dict[int, float] = {}
    if installer_ids:
        users = (
            await db.execute(
                select(User.id, User.first_name, User.last_name).where(User.id.in_(installer_ids))
            )
        ).all()
        users_map = {
            row[0]: (f"{(row[1] or '').strip()} {(row[2] or '').strip()}".strip() or f"User {row[0]}")
            for row in users
        }

        rate_rows = (
            await db.execute(
                select(
                    InstallerRateHistory.installer_id,
                    InstallerRateHistory.hourly_rate,
                    InstallerRateHistory.effective_from,
                )
                .where(
                    InstallerRateHistory.installer_id.in_(installer_ids),
                    InstallerRateHistory.is_active.is_(True),
                    or_(InstallerRateHistory.effective_to.is_(None), InstallerRateHistory.effective_to >= (start_dt or datetime.min)),
                    InstallerRateHistory.effective_from <= (end_dt or datetime.now()),
                )
                .order_by(InstallerRateHistory.installer_id, InstallerRateHistory.effective_from.desc())
            )
        ).all()

        for installer_id, hourly_rate, _effective_from in rate_rows:
            if installer_id not in rates_map:
                rates_map[installer_id] = _to_float(hourly_rate)

    installer_rows = []
    for installer_id, stats in by_installer.items():
        work_hours = _to_float(stats.get("work_hours", 0.0))
        sqft_installed = _to_float(stats.get("sqft_installed", 0.0))
        sqft_per_hour = round((sqft_installed / work_hours), 2) if work_hours else 0.0
        hourly_rate = _to_float(rates_map.get(installer_id, 0.0))
        labor_cost = round(work_hours * hourly_rate, 2)
        cost_per_sqft = round((labor_cost / sqft_installed), 2) if sqft_installed else 0.0

        installer_rows.append(
            {
                "installer_id": installer_id,
                "installer_name": users_map.get(installer_id, f"User {installer_id}" if installer_id else "Unknown"),
                "completed_installs": stats["completed_installs"],
                "sqft_installed": round(sqft_installed, 2),
                "work_hours": round(work_hours, 2),
                "pause_hours": round(_to_float(stats.get("pause_hours", 0.0)), 2),
                "sqft_per_hour": sqft_per_hour,
                "hourly_rate": round(hourly_rate, 2),
                "labor_cost": labor_cost,
                "labor_cost_per_sqft": cost_per_sqft,
                "first_completion_at": stats["first_completion_at"].isoformat() if stats.get("first_completion_at") else None,
                "last_completion_at": stats["last_completion_at"].isoformat() if stats.get("last_completion_at") else None,
            }
        )

    installer_rows.sort(key=lambda x: (x["sqft_installed"], x["completed_installs"]), reverse=True)
    installer_rows = installer_rows[:top_n]

    total_sqft_installed = round(sum(row["sqft_installed"] for row in installer_rows), 2)
    total_work_hours = round(sum(row["work_hours"] for row in installer_rows), 2)
    total_labor_cost = round(sum(row["labor_cost"] for row in installer_rows), 2)

    return success_response(
        {
            "period": {
                "start_date": effective_start_date.isoformat() if effective_start_date else None,
                "end_date": effective_end_date.isoformat() if effective_end_date else None,
            },
            "filters": {
                "installer_id": installer_id,
                "installer_name": installer_name,
            },
            "summary": {
                "installer_count": len(installer_rows),
                "total_sqft_installed": total_sqft_installed,
                "total_work_hours": total_work_hours,
                "total_labor_cost": total_labor_cost,
                "portfolio_labor_cost_per_sqft": round((total_labor_cost / total_sqft_installed), 2) if total_sqft_installed else 0.0,
                "portfolio_sqft_per_hour": round((total_sqft_installed / total_work_hours), 2) if total_work_hours else 0.0,
            },
            "installer_breakdown": installer_rows,
        },
        "Owner install performance report generated",
    )


@router.get("/reports/owner/weekly-trends", response_model=SuccessResponse[dict])
async def get_owner_weekly_trends_report(
    weeks: int = Query(12, ge=4, le=52, description="How many trailing weeks to include"),
    from_date: Optional[date] = Query(None, description="Inclusive from date filter"),
    to_date: Optional[date] = Query(None, description="Inclusive to date filter"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Weekly trendline for owner review with fab, install, and template metrics."""
    _ = current_user
    if from_date is not None or to_date is not None:
        effective_from = from_date or to_date
        effective_to = to_date or from_date
        start_dt, end_dt = _range_bounds(effective_from, effective_to)
    else:
        start_dt = datetime.now() - timedelta(days=weeks * 7)
        end_dt = None

    week_bucket = func.date_trunc(literal_column("'week'"), Fab.created_at)

    fab_filters = []
    _apply_datetime_filters(fab_filters, Fab.created_at, start_dt, end_dt)

    fab_query = (
        select(
            week_bucket.label("week_start"),
            func.count(Fab.id).label("fabs_created"),
            func.sum(Fab.revenue).label("revenue"),
            func.sum(Fab.gp).label("gp"),
        )
        .group_by(week_bucket)
        .order_by(week_bucket)
    )
    if fab_filters:
        fab_query = fab_query.where(and_(*fab_filters))

    fab_week_rows = (
        await db.execute(fab_query)
    ).all()

    install_activity_date = func.coalesce(InstallCompletion.completion_date, InstallCompletion.install_date)
    install_filters = []
    _apply_datetime_filters(install_filters, install_activity_date, start_dt, end_dt)

    install_query = select(
        install_activity_date.label("activity_date"),
        InstallCompletion.is_completed,
        InstallCompletion.total_sqft_installed,
    )
    if install_filters:
        install_query = install_query.where(and_(*install_filters))

    install_rows = (
        await db.execute(install_query)
    ).all()

    template_activity_date = func.coalesce(Templating.actual_end_date, Templating.actual_start_date, Templating.created_at)
    template_filters = []
    _apply_datetime_filters(template_filters, template_activity_date, start_dt, end_dt)

    template_query = select(
        template_activity_date.label("activity_date"),
        Templating.total_sqft,
        Templating.is_completed,
    )
    if template_filters:
        template_query = template_query.where(and_(*template_filters))

    template_rows = (
        await db.execute(template_query)
    ).all()

    by_week: dict[str, dict] = {}

    def _blank_week_row(week_key: str) -> dict:
        return {
            "week_start": week_key,
            "fabs_created": 0,
            "installs_completed": 0,
            "installs_not_completed": 0,
            "revenue": 0.0,
            "gross_profit": 0.0,
            "sqft_installed": 0.0,
            "sqft_templated": 0.0,
            "sqft_not_templated": 0.0,
        }

    for week_start, fabs_created, revenue, gp in fab_week_rows:
        key = week_start.date().isoformat()
        by_week[key] = {
            "week_start": key,
            "fabs_created": int(fabs_created or 0),
            "installs_completed": 0,
            "installs_not_completed": 0,
            "revenue": round(_to_float(revenue), 2),
            "gross_profit": round(_to_float(gp), 2),
            "sqft_installed": 0.0,
            "sqft_templated": 0.0,
            "sqft_not_templated": 0.0,
        }

    installs_by_week: dict[str, dict] = defaultdict(
        lambda: {
            "installs_completed": 0,
            "installs_not_completed": 0,
            "sqft_installed": 0.0,
        }
    )
    for activity_at, is_completed, sqft_installed in install_rows:
        if activity_at is None:
            continue
        activity_day = activity_at.date() if hasattr(activity_at, "date") else activity_at
        week_start = activity_day - timedelta(days=activity_day.weekday())
        key = week_start.isoformat()
        if bool(is_completed):
            installs_by_week[key]["installs_completed"] += 1
            installs_by_week[key]["sqft_installed"] += _to_float(sqft_installed)
        else:
            installs_by_week[key]["installs_not_completed"] += 1

    templates_by_week: dict[str, dict] = defaultdict(
        lambda: {
            "sqft_templated": 0.0,
            "sqft_not_templated": 0.0,
        }
    )
    for activity_at, total_sqft, is_completed in template_rows:
        if activity_at is None:
            continue
        activity_day = activity_at.date() if hasattr(activity_at, "date") else activity_at
        week_start = activity_day - timedelta(days=activity_day.weekday())
        key = week_start.isoformat()
        if bool(is_completed):
            templates_by_week[key]["sqft_templated"] += _to_float(total_sqft)
        else:
            templates_by_week[key]["sqft_not_templated"] += _to_float(total_sqft)

    for key, values in installs_by_week.items():
        if key not in by_week:
            by_week[key] = _blank_week_row(key)
        by_week[key]["installs_completed"] = values["installs_completed"]
        by_week[key]["installs_not_completed"] = values["installs_not_completed"]
        by_week[key]["sqft_installed"] = round(values["sqft_installed"], 2)

    for key, values in templates_by_week.items():
        if key not in by_week:
            by_week[key] = _blank_week_row(key)
        by_week[key]["sqft_templated"] = round(values["sqft_templated"], 2)
        by_week[key]["sqft_not_templated"] = round(values["sqft_not_templated"], 2)

    weekly_rows = [by_week[key] for key in sorted(by_week.keys())]

    return success_response(
        {
            "weeks": weeks,
            "period": {
                "from_date": start_dt.date().isoformat() if start_dt else None,
                "to_date": end_dt.date().isoformat() if end_dt else None,
            },
            "weekly_trends": weekly_rows,
        },
        "Owner weekly trends report generated",
    )


@router.get("/reports/owner/installation-template", response_model=SuccessResponse[dict])
async def get_owner_installation_template_report(
    start_date: Optional[date] = Query(None, description="Inclusive start date filter"),
    end_date: Optional[date] = Query(None, description="Inclusive end date filter"),
    activity: str = Query("both", pattern="^(both|installation|template)$"),
    limit: int = Query(250, ge=1, le=2000),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Combined Installation and Template report for table-based owner view."""
    start_dt, end_dt = _range_bounds(start_date, end_date)
    rows: list[dict] = []
    installer_ids: set[int] = set()

    if activity in ("both", "installation"):
        install_filters = []
        _apply_datetime_filters(install_filters, InstallCompletion.completion_date, start_dt, end_dt)

        install_query = (
            select(
                InstallCompletion.installer_id,
                InstallCompletion.fab_id,
                InstallCompletion.total_sqft_installed,
                InstallCompletion.is_completed,
                InstallCompletion.completion_notes,
                InstallCompletion.completion_date,
                BusinessJob.name,
                BusinessJob.sq_ft,
                func.coalesce(func.sum(InstallerJobTimerSession.total_work_seconds), 0).label("work_seconds"),
            )
            .join(Fab, Fab.id == InstallCompletion.fab_id, isouter=True)
            .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
            .join(
                InstallerJobTimerSession,
                and_(
                    InstallerJobTimerSession.fab_id == InstallCompletion.fab_id,
                    InstallerJobTimerSession.installer_id == InstallCompletion.installer_id,
                ),
                isouter=True,
            )
            .group_by(
                InstallCompletion.installer_id,
                InstallCompletion.fab_id,
                InstallCompletion.total_sqft_installed,
                InstallCompletion.is_completed,
                InstallCompletion.completion_notes,
                InstallCompletion.completion_date,
                BusinessJob.name,
                BusinessJob.sq_ft,
            )
            .order_by(InstallCompletion.completion_date.desc())
        )
        if install_filters:
            install_query = install_query.where(and_(*install_filters))

        install_rows = (await db.execute(install_query)).all()

        for installer_id, _fab_id, sqft_installed_raw, is_completed, completion_notes, completed_at, job_name, job_sqft, work_seconds in install_rows:
            installed_sqft = round(_to_float(sqft_installed_raw), 2)
            target_sqft = _to_float(job_sqft)
            incomplete_sqft = round(max(target_sqft - installed_sqft, 0.0), 2) if target_sqft > 0 else 0.0
            work_seconds_int = int(_to_float(work_seconds))

            if installer_id is not None:
                installer_ids.add(installer_id)

            rows.append(
                {
                    "activity_type": "Installation",
                    "activity_date": completed_at.isoformat() if completed_at else None,
                    "installer_id": installer_id,
                    "installer": None,
                    "installer_hours": round(work_seconds_int / 3600, 2),
                    "job_name": job_name or "Unknown Job",
                    "activity_complete": bool(is_completed),
                    "time_duration": _format_duration_hhmm(work_seconds_int),
                    "sq_ft_installed": installed_sqft,
                    "sq_ft_incomplete": incomplete_sqft,
                    "reason_if_not_complete": None if bool(is_completed) else (_notes_to_text(completion_notes) or "Not marked complete"),
                }
            )

    if activity in ("both", "template"):
        template_activity_date = func.coalesce(Templating.actual_end_date, Templating.actual_start_date, Templating.created_at)
        template_filters = []
        _apply_datetime_filters(template_filters, template_activity_date, start_dt, end_dt)

        template_query = (
            select(
                Templating.technician_id,
                Templating.total_sqft,
                Templating.is_completed,
                Templating.duration,
                Templating.notes,
                template_activity_date.label("activity_date"),
                BusinessJob.name,
            )
            .join(Fab, Fab.id == Templating.fab_id, isouter=True)
            .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
            .order_by(template_activity_date.desc())
        )
        if template_filters:
            template_query = template_query.where(and_(*template_filters))

        template_rows = (await db.execute(template_query)).all()

        for technician_id, total_sqft_raw, is_completed, duration_minutes, notes, activity_date_value, job_name in template_rows:
            total_sqft = round(_to_float(total_sqft_raw), 2)
            is_done = bool(is_completed)
            duration_seconds = int(_to_float(duration_minutes) * 60)

            if technician_id is not None:
                installer_ids.add(technician_id)

            rows.append(
                {
                    "activity_type": "Template",
                    "activity_date": activity_date_value.isoformat() if activity_date_value else None,
                    "installer_id": technician_id,
                    "installer": None,
                    "installer_hours": round(duration_seconds / 3600, 2),
                    "job_name": job_name or "Unknown Job",
                    "activity_complete": is_done,
                    "time_duration": _format_duration_hhmm(duration_seconds),
                    "sq_ft_installed": total_sqft if is_done else 0.0,
                    "sq_ft_incomplete": 0.0 if is_done else total_sqft,
                    "reason_if_not_complete": None if is_done else (_notes_to_text(notes) or "Not marked complete"),
                }
            )

    name_map: dict[int, str] = {}
    if installer_ids:
        user_rows = (
            await db.execute(
                select(User.id, User.first_name, User.last_name).where(User.id.in_(list(installer_ids)))
            )
        ).all()
        name_map = {
            row[0]: (f"{(row[1] or '').strip()} {(row[2] or '').strip()}".strip() or f"User {row[0]}")
            for row in user_rows
        }

    for row in rows:
        installer_id = row.get("installer_id")
        row["installer"] = name_map.get(installer_id, f"User {installer_id}" if installer_id else "Unknown")

    rows.sort(key=lambda item: item.get("activity_date") or "", reverse=True)
    rows = rows[:limit]

    return success_response(
        {
            "title": "Installation and Template Report",
            "period": {
                "start_date": start_date.isoformat() if start_date else None,
                "end_date": end_date.isoformat() if end_date else None,
            },
            "columns": [
                "installer",
                "installer_hours",
                "job_name",
                "activity_complete",
                "time_duration",
                "sq_ft_installed",
                "sq_ft_incomplete",
                "reason_if_not_complete",
            ],
            "summary": {
                "row_count": len(rows),
                "total_installer_hours": round(sum(_to_float(r.get("installer_hours", 0.0)) for r in rows), 2),
                "total_sq_ft_installed": round(sum(_to_float(r.get("sq_ft_installed", 0.0)) for r in rows), 2),
                "total_sq_ft_incomplete": round(sum(_to_float(r.get("sq_ft_incomplete", 0.0)) for r in rows), 2),
            },
            "rows": rows,
        },
        "Owner installation and template report generated",
    )


@router.patch("/reports/owner/installation-template-dashboard", response_model=SuccessResponse[dict])
async def update_owner_installation_template_dashboard(
    request: InstallationTemplateDashboardPatchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update fields for a specific templater or installer row in the dashboard."""
    _ = current_user

    if request.type.lower() == "templater":
        # Update Templating
        templating_record = (await db.execute(select(Templating).where(Templating.fab_id == request.fab_id))).scalars().first()
        if templating_record:
            if request.activity_complete is not None:
                templating_record.is_completed = request.activity_complete
            if request.reason is not None:
                templating_record.notes = [request.reason]
            if request.duration is not None:
                templating_record.duration = request.duration
            templating_record.updated_at = datetime.now()
            templating_record.updated_by = current_user.id
            db.add(templating_record)

        # Update TemplaterJobTimerSession
        if request.sqft_templated is not None or request.sqft_not_templated is not None:
            if request.job_id is not None:
                query = select(TemplaterJobTimerSession).where(
                    TemplaterJobTimerSession.job_id == request.job_id
                )
                if request.installer_id:
                    query = query.where(TemplaterJobTimerSession.templater_id == request.installer_id)
                query = query.order_by(TemplaterJobTimerSession.id.desc())
                timer_session = (await db.execute(query)).scalars().first()
                if timer_session:
                    if request.sqft_templated is not None:
                        timer_session.sqft_templated = request.sqft_templated
                    if request.sqft_not_templated is not None:
                        timer_session.sqft_not_templated = request.sqft_not_templated
                    timer_session.updated_at = datetime.now()
                    timer_session.updated_by = current_user.id
                    db.add(timer_session)

    elif request.type.lower() == "installer":
        # Update InstallCompletion
        install_record = (await db.execute(select(InstallCompletion).where(InstallCompletion.fab_id == request.fab_id))).scalars().first()
        if install_record:
            if request.activity_complete is not None:
                install_record.is_completed = request.activity_complete
            if request.reason is not None:
                install_record.completion_notes = request.reason
            install_record.updated_at = datetime.now()
            install_record.updated_by = current_user.id
            db.add(install_record)
    else:
        return error_response(status.HTTP_400_BAD_REQUEST, "Invalid type. Must be 'templater' or 'installer'")

    await db.commit()
    return success_response({}, "Dashboard record updated successfully")


@router.get("/reports/owner/installation-template-dashboard", response_model=SuccessResponse[dict])
async def get_owner_installation_template_dashboard_report(
    from_date: Optional[date] = Query(None, description="Inclusive from date filter"),
    to_date: Optional[date] = Query(None, description="Inclusive to date filter"),
    search: Optional[str] = Query(None, description="Search by job name, job number, or FAB ID"),
    fab_type: Optional[str] = Query(None, description="Optional FAB type filter"),
    sales_person_id: Optional[int] = Query(None, gt=0, description="Optional sales person filter"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Grouped installation and template dashboard report for the screenshot-based UI."""
    _ = current_user
    start_dt, end_dt = _range_bounds(from_date, to_date)
    from sqlalchemy.orm import aliased

    InstallerUser = aliased(User)
    SalesPersonUser = aliased(User)

    def _matches_common_filters(base_query, activity_date_field):
        if start_dt is not None:
            base_query = base_query.where(activity_date_field >= start_dt)
        if end_dt is not None:
            base_query = base_query.where(activity_date_field <= end_dt)
        if fab_type:
            base_query = base_query.where(func.lower(Fab.fab_type) == fab_type.strip().lower())
        if sales_person_id is not None:
            base_query = base_query.where(Fab.sales_person_id == sales_person_id)
        if search:
            search_term = f"%{search.strip()}%"
            base_query = base_query.where(
                or_(
                    BusinessJob.job_number.ilike(search_term),
                    BusinessJob.name.ilike(search_term),
                    cast(Fab.id, String).ilike(search_term),
                )
            )
        return base_query

    grouped_rows_map: dict[tuple[str, str], dict] = {}
    flat_rows: list[dict] = []

    def _department_rank(department: str) -> int:
        if department == "Templater":
            return 0
        if department == "Installer":
            return 1
        return 2

    installer_timer_totals = (
        select(
            InstallerJobTimerSession.job_id.label("job_id"),
            func.coalesce(func.sum(InstallerJobTimerSession.total_work_seconds), 0).label("work_seconds"),
        )
        .group_by(InstallerJobTimerSession.job_id)
        .subquery("installer_timer_totals")
    )

    templater_timer_totals = (
        select(
            TemplaterJobTimerSession.job_id.label("job_id"),
            func.coalesce(func.sum(TemplaterJobTimerSession.total_work_seconds), 0).label("work_seconds"),
            func.coalesce(func.sum(TemplaterJobTimerSession.sqft_templated), 0).label("sqft_templated"),
            func.coalesce(func.sum(TemplaterJobTimerSession.sqft_not_templated), 0).label("sqft_not_templated"),
        )
        .group_by(TemplaterJobTimerSession.job_id)
        .subquery("templater_timer_totals")
    )

    install_query = _matches_common_filters(
        select(
            InstallCompletion.installer_id,
            InstallerUser.first_name.label("installer_first_name"),
            InstallerUser.last_name.label("installer_last_name"),
            BusinessJob.id.label("job_id"),
            BusinessJob.job_number,
            BusinessJob.name.label("job_name"),
            Account.name.label("account_name"),
            Fab.id.label("fab_id"),
            Fab.fab_type,
            Fab.sales_person_id,
            SalesPersonUser.first_name.label("sales_person_first_name"),
            SalesPersonUser.last_name.label("sales_person_last_name"),
            InstallCompletion.is_completed,
            InstallCompletion.completion_notes,
            InstallCompletion.completion_date,
            InstallCompletion.total_sqft_installed,
            BusinessJob.sq_ft,
            func.coalesce(installer_timer_totals.c.work_seconds, 0).label("work_seconds"),
        )
        .select_from(InstallCompletion)
        .join(Fab, Fab.id == InstallCompletion.fab_id)
        .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
        .join(Account, Account.id == BusinessJob.account_id, isouter=True)
        .join(InstallerUser, InstallerUser.id == InstallCompletion.installer_id, isouter=True)
        .join(SalesPersonUser, SalesPersonUser.id == Fab.sales_person_id, isouter=True)
        .join(
            installer_timer_totals,
            installer_timer_totals.c.job_id == BusinessJob.id,
            isouter=True,
        )
        .order_by(InstallCompletion.completion_date.desc(), Fab.id.desc())
    , InstallCompletion.completion_date)

    install_rows_db = (await db.execute(install_query)).all()

    template_activity_date = func.coalesce(Templating.actual_end_date, Templating.actual_start_date, Templating.created_at)
    template_query = _matches_common_filters(
        select(
            Templating.technician_id,
            InstallerUser.first_name.label("installer_first_name"),
            InstallerUser.last_name.label("installer_last_name"),
            BusinessJob.id.label("job_id"),
            BusinessJob.job_number,
            BusinessJob.name.label("job_name"),
            Account.name.label("account_name"),
            Fab.id.label("fab_id"),
            Fab.fab_type,
            Fab.sales_person_id,
            SalesPersonUser.first_name.label("sales_person_first_name"),
            SalesPersonUser.last_name.label("sales_person_last_name"),
            Templating.is_completed,
            Templating.notes,
            template_activity_date.label("activity_date"),
            Templating.total_sqft,
            BusinessJob.sq_ft,
            Templating.duration,
            func.coalesce(templater_timer_totals.c.work_seconds, 0).label("work_seconds"),
            func.coalesce(templater_timer_totals.c.sqft_templated, 0).label("timer_sqft_templated"),
            func.coalesce(templater_timer_totals.c.sqft_not_templated, 0).label("timer_sqft_not_templated"),
        )
        .select_from(Templating)
        .join(Fab, Fab.id == Templating.fab_id, isouter=True)
        .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
        .join(Account, Account.id == BusinessJob.account_id, isouter=True)
        .join(InstallerUser, InstallerUser.id == Templating.technician_id, isouter=True)
        .join(SalesPersonUser, SalesPersonUser.id == Fab.sales_person_id, isouter=True)
        .join(
            templater_timer_totals,
            templater_timer_totals.c.job_id == BusinessJob.id,
            isouter=True,
        )
        .order_by(template_activity_date.desc())
    , template_activity_date)

    template_rows_db = (await db.execute(template_query)).all()

    def _get_group(department: str, installer_name: str, installer_id: Optional[int]) -> dict:
        group_key = (department, installer_name)
        group = grouped_rows_map.get(group_key)
        if group is None:
            group = {
                "department": department,
                "installer": installer_name,
                "installer_id": installer_id,
                "activity_label": department,
                "total_seconds": 0,
                "job_count": 0,
                "rows": [],
            }
            grouped_rows_map[group_key] = group
        return group

    for row in install_rows_db:
        (
            installer_id,
            installer_first_name,
            installer_last_name,
            job_id,
            job_number,
            job_name,
            account_name,
            fab_row_id,
            row_fab_type,
            row_sales_person_id,
            sales_person_first_name,
            sales_person_last_name,
            is_completed,
            completion_notes,
            completed_at,
            sqft_installed_raw,
            job_sqft,
            work_seconds,
        ) = row

        installer_name = f"{(installer_first_name or '').strip()} {(installer_last_name or '').strip()}".strip() or (f"User {installer_id}" if installer_id else "Unknown")
        sales_person_name = f"{(sales_person_first_name or '').strip()} {(sales_person_last_name or '').strip()}".strip() or (f"User {row_sales_person_id}" if row_sales_person_id else None)
        installed_sqft = round(_to_float(sqft_installed_raw), 2)
        incomplete_sqft = round(max(_to_float(job_sqft) - installed_sqft, 0.0), 2) if _to_float(job_sqft) > 0 else 0.0
        total_seconds = int(_to_float(work_seconds))
        group = _get_group("Installer", installer_name, installer_id)
        group["total_seconds"] += total_seconds
        group["job_count"] += 1

        flat_rows.append(
            {
            "department": "Installer",
                "installer": installer_name,
                "installer_id": installer_id,
                "installer_hours": round(total_seconds / 3600, 2),
                "activity_type": "Installation",
                "activity_date": completed_at.isoformat() if completed_at else None,
                "job_id": job_id,
                "fab_id": fab_row_id,
                "fab_type": row_fab_type,
                "job_number": job_number,
                "job_name": job_name or "Unknown Job",
                "account_name": account_name,
                "activity_complete": bool(is_completed),
                "duration": _format_duration_hhmm(total_seconds),
                "sq_ft_installed": installed_sqft,
                "sq_ft_incomplete": incomplete_sqft,
                "sqft_templated": installed_sqft,
                "sqft_not_templated": incomplete_sqft,
                "reason_if_not_complete": None if bool(is_completed) else (_notes_to_text(completion_notes) or "Not marked complete"),
                "sales_person_id": row_sales_person_id,
                "sales_person_name": sales_person_name,
            }
        )
        group["rows"].append(flat_rows[-1])

    for row in template_rows_db:
        (
            technician_id,
            installer_first_name,
            installer_last_name,
            job_id,
            job_number,
            job_name,
            account_name,
            fab_row_id,
            row_fab_type,
            row_sales_person_id,
            sales_person_first_name,
            sales_person_last_name,
            is_completed,
            notes,
            activity_date_value,
            total_sqft_raw,
            job_sqft,
            duration_minutes,
            work_seconds,
            timer_sqft_templated_raw,
            timer_sqft_not_templated_raw,
        ) = row

        installer_name = f"{(installer_first_name or '').strip()} {(installer_last_name or '').strip()}".strip() or (f"User {technician_id}" if technician_id else "Unknown")
        sales_person_name = f"{(sales_person_first_name or '').strip()} {(sales_person_last_name or '').strip()}".strip() or (f"User {row_sales_person_id}" if row_sales_person_id else None)
        total_sqft = round(_to_float(total_sqft_raw), 2)
        templated_sqft = round(_to_float(timer_sqft_templated_raw), 2)
        not_templated_sqft = round(_to_float(timer_sqft_not_templated_raw), 2)
        if templated_sqft <= 0 and not_templated_sqft <= 0 and total_sqft > 0:
            if bool(is_completed):
                templated_sqft = total_sqft
                not_templated_sqft = 0.0
            else:
                templated_sqft = 0.0
                not_templated_sqft = total_sqft
        total_seconds = int(_to_float(work_seconds))
        if total_seconds <= 0 and duration_minutes is not None:
            total_seconds = int(_to_float(duration_minutes) * 60)
        group = _get_group("Templater", installer_name, technician_id)
        group["total_seconds"] += total_seconds
        group["job_count"] += 1

        flat_rows.append(
            {
                "department": "Templater",
                "installer": installer_name,
                "installer_id": technician_id,
                "installer_hours": round(total_seconds / 3600, 2),
                "activity_type": "Template",
                "activity_date": activity_date_value.isoformat() if activity_date_value else None,
                "job_id": job_id,
                "fab_id": fab_row_id,
                "fab_type": row_fab_type,
                "job_number": job_number,
                "job_name": job_name or "Unknown Job",
                "account_name": account_name,
                "activity_complete": bool(is_completed),
                "duration": _format_duration_hhmm(total_seconds),
                "sq_ft_installed": templated_sqft,
                "sq_ft_incomplete": not_templated_sqft,
                "sqft_templated": templated_sqft,
                "sqft_not_templated": not_templated_sqft,
                "reason_if_not_complete": None if bool(is_completed) else (_notes_to_text(notes) or "Not marked complete"),
                "sales_person_id": row_sales_person_id,
                "sales_person_name": sales_person_name,
            }
        )
        group["rows"].append(flat_rows[-1])

    grouped_rows = []
    for group_data in sorted(
        grouped_rows_map.values(),
        key=lambda item: (_department_rank(item["department"]), item["installer"].lower()),
    ):
        total_seconds = int(group_data.get("total_seconds") or 0)
        group = {k: v for k, v in group_data.items() if k != "total_seconds"}
        group["installer_hours"] = round(total_seconds / 3600, 2)
        group["installer_hours_display"] = _format_duration_hhmm(total_seconds)
        group["rows"] = sorted(group["rows"], key=lambda item: item.get("activity_date") or "", reverse=True)
        grouped_rows.append(group)

    flat_rows.sort(key=lambda item: item.get("activity_date") or "", reverse=True)

    total_seconds_templated = sum(
        int(group["total_seconds"])
        for group in grouped_rows_map.values()
        if group["department"] == "Templater"
    )
    total_seconds_installed = sum(
        int(group["total_seconds"])
        for group in grouped_rows_map.values()
        if group["department"] == "Installer"
    )
    sqft_templated = round(
        sum(_to_float(row["sq_ft_installed"]) for row in flat_rows if row["activity_type"] == "Template" and row["activity_complete"]),
        2,
    )
    sqft_not_templated = round(
        sum(_to_float(row["sq_ft_incomplete"]) for row in flat_rows if row["department"] == "Templater"),
        2,
    )
    installs_sq_ft = round(
        sum(_to_float(row["sq_ft_installed"]) for row in flat_rows if row["activity_type"] == "Installation"),
        2,
    )
    incomplete_sq_ft = round(sum(_to_float(row["sq_ft_incomplete"]) for row in flat_rows), 2)

    sales_person_rows = (
        await db.execute(
            select(User.id, User.first_name, User.last_name)
            .where(User.status == 1)
            .order_by(User.first_name.asc(), User.last_name.asc())
        )
    ).all()
    sales_person_options = [
        {
            "id": row[0],
            "name": f"{(row[1] or '').strip()} {(row[2] or '').strip()}".strip() or f"User {row[0]}",
        }
        for row in sales_person_rows
    ]

    fab_type_options = sorted({row.get("fab_type") for row in flat_rows if row.get("fab_type")})

    return success_response(
        {
            "title": "Installation and Template Report",
            "period": {
                "from_date": from_date.isoformat() if from_date else None,
                "to_date": to_date.isoformat() if to_date else None,
            },
            "columns": [
                "installer",
                "installer_hours",
                "job_id",
                "job_name",
                "account_name",
                "activity_complete",
                "duration",
                "sq_ft_installed",
                "sq_ft_incomplete",
                "sqft_templated",
                "sqft_not_templated",
                "reason_if_not_complete",
            ],
            "filters": {
                "search": search,
                "fab_type": fab_type,
                "sales_person_id": sales_person_id,
            },
            "filter_options": {
                "sales_person_options": sales_person_options,
                "fab_types": fab_type_options,
            },
            "summary": {
                "total_hours_templated": _format_duration_hhmm(total_seconds_templated),
                "total_hours_installed": _format_duration_hhmm(total_seconds_installed),
                "sqft_installed": installs_sq_ft,
                "sqft_not_installed": incomplete_sq_ft,
                "sqft_templated": sqft_templated,
                "sqft_not_templated": sqft_not_templated,
                "row_count": len(flat_rows),
                "group_count": len(grouped_rows),
            },
            "groups": grouped_rows,
            "rows": flat_rows,
        },
        "Owner installation and template dashboard report generated",
    )


@router.get("/reports/owner/installation-template-dashboard/pdf")
async def get_owner_installation_template_dashboard_pdf(
    from_date: Optional[date] = Query(None, description="Inclusive from date filter"),
    to_date: Optional[date] = Query(None, description="Inclusive to date filter"),
    search: Optional[str] = Query(None, description="Search by job name, job number, or FAB ID"),
    fab_type: Optional[str] = Query(None, description="Optional FAB type filter"),
    sales_person_id: Optional[int] = Query(None, gt=0, description="Optional sales person filter"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download a PDF version of the installation and template dashboard report."""
    import json as _json
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    # Reuse existing data-fetching logic
    json_response = await get_owner_installation_template_dashboard_report(
        from_date=from_date,
        to_date=to_date,
        search=search,
        fab_type=fab_type,
        sales_person_id=sales_person_id,
        db=db,
        current_user=current_user,
    )
    payload = _json.loads(json_response.body)
    data = payload.get("data", {})

    summary = data.get("summary", {})
    groups = data.get("groups", [])
    period = data.get("period", {})

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(letter),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    brand_blue = colors.HexColor("#1E3A5F")
    brand_light = colors.HexColor("#E8EDF3")
    green_ok = colors.HexColor("#2E7D32")
    red_no = colors.HexColor("#C62828")
    grey_mid = colors.HexColor("#6B7280")

    title_style = ParagraphStyle("ReportTitle", parent=styles["Heading1"], fontSize=16, textColor=brand_blue, spaceAfter=4)
    subtitle_style = ParagraphStyle("ReportSubtitle", parent=styles["Normal"], fontSize=9, textColor=grey_mid, spaceAfter=2)
    group_header_style = ParagraphStyle("GroupHeader", parent=styles["Normal"], fontSize=10, textColor=brand_blue, fontName="Helvetica-Bold", spaceBefore=8, spaceAfter=3)
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7, leading=9)
    small_style = ParagraphStyle("Small", parent=styles["Normal"], fontSize=7, textColor=grey_mid)

    story = []

    # ── Title & period ──
    period_from = period.get("from_date") or "All time"
    period_to = period.get("to_date") or "All time"
    period_label = f"{period_from} – {period_to}" if period.get("from_date") or period.get("to_date") else "All time"

    story.append(Paragraph("Installation & Template Dashboard Report", title_style))
    story.append(Paragraph(f"Period: {period_label}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))

    active_filters = []
    if search:
        active_filters.append(f"Search: {search}")
    if fab_type:
        active_filters.append(f"FAB Type: {fab_type}")
    if sales_person_id:
        active_filters.append(f"Sales Person ID: {sales_person_id}")
    if active_filters:
        story.append(Paragraph("Filters: " + "  |  ".join(active_filters), subtitle_style))

    story.append(Spacer(1, 6))

    # ── Summary bar ──
    summary_data = [
        ["Total Hours Templated", "Total Hours Installed", "Sqft Templated", "Sqft Not Templated", "Sqft Installed", "Sqft Not Installed", "Jobs"],
        [
            summary.get("total_hours_templated", "–"),
            summary.get("total_hours_installed", "–"),
            str(summary.get("sqft_templated", 0)),
            str(summary.get("sqft_not_templated", 0)),
            str(summary.get("sqft_installed", 0)),
            str(summary.get("sqft_not_installed", 0)),
            str(summary.get("row_count", 0)),
        ],
    ]
    col_w = doc.width / len(summary_data[0])
    summary_table = Table(summary_data, colWidths=[col_w] * len(summary_data[0]))
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand_blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, 1), [brand_light]),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.white),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 10))

    # ── Per-group tables ──
    row_headers = ["#", "Date", "Job #", "Job Name", "Account", "FAB Type", "Sqft ✓", "Sqft ✗", "Duration", "Complete", "Notes"]
    col_widths = [
        0.25 * inch,  # #
        0.65 * inch,  # Date
        0.65 * inch,  # Job #
        1.4 * inch,   # Job Name
        1.0 * inch,   # Account
        0.7 * inch,   # FAB Type
        0.55 * inch,  # Sqft ✓
        0.55 * inch,  # Sqft ✗
        0.55 * inch,  # Duration
        0.55 * inch,  # Complete
        1.7 * inch,   # Notes
    ]

    for group in groups:
        department = group.get("department", "")
        installer = group.get("installer", "Unknown")
        hours_display = group.get("installer_hours_display", "0:00")
        job_count = group.get("job_count", 0)
        rows = group.get("rows", [])

        dept_color = brand_blue if department == "Installer" else colors.HexColor("#4A6741")
        story.append(Paragraph(
            f"{department} — {installer}   |   Total Hours: {hours_display}   |   Jobs: {job_count}",
            ParagraphStyle("GH", parent=group_header_style, textColor=dept_color),
        ))

        table_data = [row_headers]
        for idx, r in enumerate(rows, start=1):
            activity_date = (r.get("activity_date") or "")[:10]
            complete = "Yes" if r.get("activity_complete") else "No"
            notes = r.get("reason_if_not_complete") or ""
            sqft_ok = r.get("sqft_templated" if department == "Templater" else "sq_ft_installed", 0)
            sqft_no = r.get("sqft_not_templated" if department == "Templater" else "sq_ft_incomplete", 0)
            table_data.append([
                str(idx),
                activity_date,
                r.get("job_number") or "–",
                Paragraph(r.get("job_name") or "–", cell_style),
                Paragraph(r.get("account_name") or "–", cell_style),
                r.get("fab_type") or "–",
                str(sqft_ok),
                str(sqft_no),
                r.get("duration") or "–",
                complete,
                Paragraph(notes, cell_style),
            ])

        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        row_fills = []
        for i in range(1, len(table_data)):
            bg = colors.white if i % 2 == 0 else brand_light
            row_fills.append(("BACKGROUND", (0, i), (-1, i), bg))

        complete_col_idx = row_headers.index("Complete")
        complete_style_cmds = []
        for i, r in enumerate(rows, start=1):
            txt_color = green_ok if r.get("activity_complete") else red_no
            complete_style_cmds.append(("TEXTCOLOR", (complete_col_idx, i), (complete_col_idx, i), txt_color))

        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), dept_color),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (3, 1), (4, -1), "LEFT"),
            ("ALIGN", (complete_col_idx, 1), (complete_col_idx, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            *row_fills,
            *complete_style_cmds,
        ]))
        story.append(tbl)
        story.append(Spacer(1, 6))

    doc.build(story)
    buf.seek(0)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"installation_template_dashboard_{stamp}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/reports/owner/daily-completion", response_model=SuccessResponse[dict])
@router.get("/reports/owner/daily-shop-completion", response_model=SuccessResponse[dict])
async def get_owner_daily_completion_report(
    from_date: Optional[date] = Query(None, description="Inclusive from date filter"),
    to_date: Optional[date] = Query(None, description="Inclusive to date filter"),
    weekdays: Optional[int] = Query(None, ge=0, description="Optional weekday override for AVG PER WEEKDAY"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Daily completion rollup by stage (Template, Draft, SCT, Final, Resurface, Cut, Shop Fab)."""
    _ = current_user

    today = date.today()
    effective_to = to_date or from_date or today
    effective_from = from_date or to_date or (effective_to - timedelta(days=6))
    if effective_from > effective_to:
        effective_from, effective_to = effective_to, effective_from

    def _date_key(value) -> Optional[str]:
        if value is None:
            return None
        if hasattr(value, "date"):
            return value.date().isoformat()
        return str(value)

    day_rows_map: dict[str, dict] = {}
    current_day = effective_from
    while current_day <= effective_to:
        key = current_day.isoformat()
        day_rows_map[key] = {
            "date": key,
            "template_sqft": 0.0,
            "draft_sqft": 0.0,
            "sct_sqft": 0.0,
            "final_programming_sqft": 0.0,
            "resurface_sqft": 0.0,
            "cut_sqft": 0.0,
            "fab_sqft": 0.0,
            "revenue": 0.0,
            "gp": 0.0,
        }
        current_day += timedelta(days=1)

    template_day_expr = func.date(func.coalesce(Templating.actual_end_date, Templating.actual_start_date, Templating.updated_at, Templating.created_at))
    template_rows = (
        await db.execute(
            select(
                template_day_expr.label("report_day"),
                func.coalesce(func.sum(_safe_numeric_col(Templating.total_sqft)), 0).label("sqft"),
            )
            .where(
                Templating.is_completed.is_(True),
                template_day_expr >= effective_from,
                template_day_expr <= effective_to,
            )
            .group_by(template_day_expr)
        )
    ).all()
    for report_day, sqft in template_rows:
        key = _date_key(report_day)
        if key in day_rows_map:
            day_rows_map[key]["template_sqft"] = round(_to_float(sqft), 2)

    draft_day_expr = func.date(func.coalesce(DraftingSession.session_end_time, DraftingSession.session_start_time))
    draft_rows = (
        await db.execute(
            select(
                draft_day_expr.label("report_day"),
                func.coalesce(func.sum(_safe_numeric_col(DraftingSession.cumulative_sqft_drafted)), 0).label("sqft"),
            )
            .where(
                draft_day_expr >= effective_from,
                draft_day_expr <= effective_to,
            )
            .group_by(draft_day_expr)
        )
    ).all()
    for report_day, sqft in draft_rows:
        key = _date_key(report_day)
        if key in day_rows_map:
            day_rows_map[key]["draft_sqft"] = round(_to_float(sqft), 2)

    sct_day_expr = func.date(Fab.sct_completed_date)
    sct_rows = (
        await db.execute(
            select(
                sct_day_expr.label("report_day"),
                func.coalesce(func.sum(Fab.total_sqft), 0).label("sqft"),
            )
            .where(
                Fab.sct_completed_date.isnot(None),
                sct_day_expr >= effective_from,
                sct_day_expr <= effective_to,
            )
            .group_by(sct_day_expr)
        )
    ).all()
    for report_day, sqft in sct_rows:
        key = _date_key(report_day)
        if key in day_rows_map:
            day_rows_map[key]["sct_sqft"] = round(_to_float(sqft), 2)

    final_day_expr = func.date(Fab.final_programming_completed_date)
    final_rows = (
        await db.execute(
            select(
                final_day_expr.label("report_day"),
                func.coalesce(func.sum(Fab.total_sqft), 0).label("sqft"),
            )
            .where(
                Fab.final_programming_completed_date.isnot(None),
                final_day_expr >= effective_from,
                final_day_expr <= effective_to,
            )
            .group_by(final_day_expr)
        )
    ).all()
    for report_day, sqft in final_rows:
        key = _date_key(report_day)
        if key in day_rows_map:
            day_rows_map[key]["final_programming_sqft"] = round(_to_float(sqft), 2)

    resurface_day_expr = func.date(func.coalesce(ResurfaceScheduling.actual_end_date, ResurfaceScheduling.actual_start_date))
    resurface_rows = (
        await db.execute(
            select(
                resurface_day_expr.label("report_day"),
                func.coalesce(func.sum(_safe_numeric_col(ResurfaceScheduling.completed_sqft)), 0).label("sqft"),
            )
            .where(
                resurface_day_expr >= effective_from,
                resurface_day_expr <= effective_to,
            )
            .group_by(resurface_day_expr)
        )
    ).all()
    for report_day, sqft in resurface_rows:
        key = _date_key(report_day)
        if key in day_rows_map:
            day_rows_map[key]["resurface_sqft"] = round(_to_float(sqft), 2)

    cut_day_expr = func.date(ShopCutPlan.actual_end_date)
    cut_rows = (
        await db.execute(
            select(
                cut_day_expr.label("report_day"),
                func.coalesce(func.sum(Fab.total_sqft), 0).label("sqft"),
            )
            .select_from(ShopCutPlan)
            .join(Fab, Fab.id == ShopCutPlan.fab_id)
            .where(
                ShopCutPlan.work_percentage == 100,
                ShopCutPlan.actual_end_date.isnot(None),
                cut_day_expr >= effective_from,
                cut_day_expr <= effective_to,
            )
            .group_by(cut_day_expr)
        )
    ).all()
    for report_day, sqft in cut_rows:
        key = _date_key(report_day)
        if key in day_rows_map:
            day_rows_map[key]["cut_sqft"] = round(_to_float(sqft), 2)

    fab_cut_day_subquery = (
        select(
            cut_day_expr.label("report_day"),
            ShopCutPlan.fab_id.label("fab_id"),
            func.max(Fab.total_sqft).label("fab_sqft"),
            func.max(_safe_numeric_col(Fab.revenue)).label("fab_revenue"),
            func.max(_safe_numeric_col(Fab.gp)).label("fab_gp"),
        )
        .select_from(ShopCutPlan)
        .join(Fab, Fab.id == ShopCutPlan.fab_id)
        .where(
            ShopCutPlan.work_percentage == 100,
            ShopCutPlan.actual_end_date.isnot(None),
            cut_day_expr >= effective_from,
            cut_day_expr <= effective_to,
        )
        .group_by(cut_day_expr, ShopCutPlan.fab_id)
        .subquery("fab_cut_day_subquery")
    )

    shop_fab_rows = (
        await db.execute(
            select(
                fab_cut_day_subquery.c.report_day,
                func.coalesce(func.sum(fab_cut_day_subquery.c.fab_sqft), 0).label("sqft"),
                func.coalesce(func.sum(fab_cut_day_subquery.c.fab_revenue), 0).label("revenue"),
                func.coalesce(func.sum(fab_cut_day_subquery.c.fab_gp), 0).label("gp"),
            )
            .group_by(fab_cut_day_subquery.c.report_day)
        )
    ).all()
    for report_day, sqft, revenue, gp in shop_fab_rows:
        key = _date_key(report_day)
        if key in day_rows_map:
            day_rows_map[key]["fab_sqft"] = round(_to_float(sqft), 2)
            day_rows_map[key]["revenue"] = round(_to_float(revenue), 2)
            day_rows_map[key]["gp"] = round(_to_float(gp), 2)

    daily_rows = [day_rows_map[key] for key in sorted(day_rows_map.keys())]

    totals = {
        "template_sqft": round(sum(_to_float(row["template_sqft"]) for row in daily_rows), 2),
        "draft_sqft": round(sum(_to_float(row["draft_sqft"]) for row in daily_rows), 2),
        "sct_sqft": round(sum(_to_float(row["sct_sqft"]) for row in daily_rows), 2),
        "final_programming_sqft": round(sum(_to_float(row["final_programming_sqft"]) for row in daily_rows), 2),
        "resurface_sqft": round(sum(_to_float(row["resurface_sqft"]) for row in daily_rows), 2),
        "cut_sqft": round(sum(_to_float(row["cut_sqft"]) for row in daily_rows), 2),
        "fab_sqft": round(sum(_to_float(row["fab_sqft"]) for row in daily_rows), 2),
        "revenue": round(sum(_to_float(row["revenue"]) for row in daily_rows), 2),
        "gp": round(sum(_to_float(row["gp"]) for row in daily_rows), 2),
    }

    calculated_weekdays = sum(1 for offset in range((effective_to - effective_from).days + 1) if (effective_from + timedelta(days=offset)).weekday() < 5)
    weekday_count = weekdays if weekdays is not None else calculated_weekdays

    avg_per_weekday = {
        "template_sqft": round(_safe_div(totals["template_sqft"], weekday_count), 2) if weekday_count else 0.0,
        "draft_sqft": round(_safe_div(totals["draft_sqft"], weekday_count), 2) if weekday_count else 0.0,
        "sct_sqft": round(_safe_div(totals["sct_sqft"], weekday_count), 2) if weekday_count else 0.0,
        "final_programming_sqft": round(_safe_div(totals["final_programming_sqft"], weekday_count), 2) if weekday_count else 0.0,
        "resurface_sqft": round(_safe_div(totals["resurface_sqft"], weekday_count), 2) if weekday_count else 0.0,
        "cut_sqft": round(_safe_div(totals["cut_sqft"], weekday_count), 2) if weekday_count else 0.0,
        "fab_sqft": round(_safe_div(totals["fab_sqft"], weekday_count), 2) if weekday_count else 0.0,
        "revenue": round(_safe_div(totals["revenue"], weekday_count), 2) if weekday_count else 0.0,
        "gp": round(_safe_div(totals["gp"], weekday_count), 2) if weekday_count else 0.0,
    }

    return success_response(
        {
            "title": "Daily Completion Report",
            "period": {
                "from_date": effective_from.isoformat(),
                "to_date": effective_to.isoformat(),
            },
            "summary": {
                "weekdays": weekday_count,
                "totals": totals,
                "avg_per_weekday": avg_per_weekday,
                "row_count": len(daily_rows),
            },
            "columns": [
                "date",
                "template_sqft",
                "draft_sqft",
                "sct_sqft",
                "final_programming_sqft",
                "resurface_sqft",
                "cut_sqft",
                "fab_sqft",
                "revenue",
                "gp",
            ],
            "rows": daily_rows,
        },
        "Owner daily completion report generated",
    )


@router.get("/reports/owner/monthly-install-completion", response_model=SuccessResponse[dict])
async def get_owner_monthly_install_completion_report(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    from_date: Optional[date] = Query(None, description="Optional from date filter (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="Optional to date filter (YYYY-MM-DD)"),
    fab_id: Optional[int] = Query(None, gt=0, description="Optional FAB ID filter"),
    job_number: Optional[str] = Query(None, description="Optional job number filter"),
    installer_name: Optional[str] = Query(None, description="Optional installer name filter"),
    fab_type: Optional[str] = Query(None, description="Optional FAB type filter"),
    fab_type_sort: str = Query("asc", pattern="^(asc|desc)$", description="Sort order for FAB type"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Monthly install completion report modeled after legacy spreadsheet format."""
    month_start_dt, month_end_dt = _month_bounds(year, month)
    effective_start_date = from_date if from_date is not None else month_start_dt.date()
    effective_end_date = to_date if to_date is not None else month_end_dt.date()
    if effective_start_date > effective_end_date:
        effective_start_date, effective_end_date = effective_end_date, effective_start_date
    start_dt, end_dt = _range_bounds(effective_start_date, effective_end_date)
    fab_type_sort_normalized = (fab_type_sort or "asc").strip().lower()
    fab_type_order = (
        func.lower(Fab.fab_type).desc()
        if fab_type_sort_normalized == "desc"
        else func.lower(Fab.fab_type).asc()
    )

    query = (
        select(
            InstallCompletion.completion_date,
            Fab.fab_type,
            Fab.id,
            BusinessJob.job_number,
            InstallCompletion.installer_id,
            User.first_name.label("installer_first_name"),
            User.last_name.label("installer_last_name"),
            BusinessJob.name.label("job_name"),
            Account.name.label("account_name"),
            StoneType.name.label("stone_type_name"),
            StoneColor.name.label("stone_color_name"),
            Edge.name.label("edge_name"),
            StoneThickness.thickness.label("stone_thickness_value"),
            Fab.input_area,
            Fab.no_of_pieces,
            Fab.total_sqft,
            Fab.revenue,
            Fab.cost_of_stone,
            CostOfStone.total_cost,
        )
        .join(Fab, Fab.id == InstallCompletion.fab_id)
        .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
        .join(User, User.id == InstallCompletion.installer_id, isouter=True)
        .join(Account, Account.id == BusinessJob.account_id, isouter=True)
        .join(StoneType, StoneType.id == Fab.stone_type_id, isouter=True)
        .join(StoneColor, StoneColor.id == Fab.stone_color_id, isouter=True)
        .join(Edge, Edge.id == Fab.edge_id, isouter=True)
        .join(StoneThickness, StoneThickness.id == Fab.stone_thickness_id, isouter=True)
        .join(CostOfStone, CostOfStone.id == Fab.cost_of_stone_id, isouter=True)
        .where(
            InstallCompletion.is_completed.is_(True),
            InstallCompletion.completion_date >= start_dt,
            InstallCompletion.completion_date <= end_dt,
        )
        .order_by(InstallCompletion.completion_date.asc(), fab_type_order, Fab.id.asc())
    )

    if fab_type:
        query = query.where(func.lower(Fab.fab_type) == fab_type.strip().lower())
    if fab_id is not None:
        query = query.where(Fab.id == fab_id)
    if job_number:
        query = query.where(BusinessJob.job_number.ilike(f"%{job_number.strip()}%"))
    if installer_name:
        installer_like = f"%{installer_name.strip().lower()}%"
        query = query.where(
            or_(
                func.lower(User.first_name).like(installer_like),
                func.lower(User.last_name).like(installer_like),
                func.lower(User.username).like(installer_like),
                func.lower(User.email).like(installer_like),
                func.lower(
                    func.concat(
                        func.coalesce(User.first_name, ""),
                        " ",
                        func.coalesce(User.last_name, ""),
                    )
                ).like(installer_like),
            )
        )

    records = (await db.execute(query)).all()

    rows = []
    daily_rollup: dict[str, dict] = defaultdict(lambda: {
        "pieces": 0,
        "sq_ft": 0.0,
        "revenue": 0.0,
        "cost_of_stone": 0.0,
        "gp": 0.0,
        "row_count": 0,
    })

    for completion_date, row_fab_type, fab_id, job_number, installer_id, installer_first_name, installer_last_name, job_name, account_name, stone_type_name, stone_color_name, edge_name, stone_thickness_value, input_area, pieces, sq_ft, revenue, fab_cost_of_stone, cos_total_cost in records:
        sq_ft_value = round(_to_float(sq_ft), 2)
        revenue_value = round(_to_float(revenue), 2)
        cost_value = round(_to_float(fab_cost_of_stone if fab_cost_of_stone is not None else cos_total_cost), 2)
        gp_value = round(revenue_value - cost_value, 2)
        pieces_value = int(_to_float(pieces))
        revenue_per_sqft = round((revenue_value / sq_ft_value), 2) if sq_ft_value else 0.0
        day_key = completion_date.date().isoformat()
        installer_name = f"{(installer_first_name or '').strip()} {(installer_last_name or '').strip()}".strip() or (
            f"Installer {installer_id}" if installer_id else None
        )

        daily_rollup[day_key]["pieces"] += pieces_value
        daily_rollup[day_key]["sq_ft"] += sq_ft_value
        daily_rollup[day_key]["revenue"] += revenue_value
        daily_rollup[day_key]["cost_of_stone"] += cost_value
        daily_rollup[day_key]["gp"] += gp_value
        daily_rollup[day_key]["row_count"] += 1

        rows.append(
            {
                "install_date": day_key,
                "fab_type": row_fab_type,
                "fab_id": fab_id,
                "job_number": job_number,
                "installer_id": installer_id,
                "installer_name": installer_name,
                "job_name": job_name,
                "account_name": account_name,
                "stone_type_name": stone_type_name,
                "stone_color_name": stone_color_name,
                "edge_name": edge_name,
                "stone_thickness_value": stone_thickness_value,
                "input_area": input_area,
                "pieces": pieces_value,
                "sq_ft": sq_ft_value,
                "revenue": revenue_value,
                "revenue_per_sq_ft": revenue_per_sqft,
                "cost_of_stone": cost_value,
                "gp": gp_value,
            }
        )

    daily_totals = []
    for day_key in sorted(daily_rollup.keys()):
        item = daily_rollup[day_key]
        day_sqft = round(item["sq_ft"], 2)
        day_revenue = round(item["revenue"], 2)
        daily_totals.append(
            {
                "install_date": day_key,
                "pieces": int(item["pieces"]),
                "sq_ft": day_sqft,
                "revenue": day_revenue,
                "revenue_per_sq_ft": round((day_revenue / day_sqft), 2) if day_sqft else 0.0,
                "cost_of_stone": round(item["cost_of_stone"], 2),
                "gp": round(item["gp"], 2),
                "row_count": int(item["row_count"]),
            }
        )

    total_sqft = round(sum(item["sq_ft"] for item in daily_rollup.values()), 2)
    total_revenue = round(sum(item["revenue"] for item in daily_rollup.values()), 2)
    total_cost = round(sum(item["cost_of_stone"] for item in daily_rollup.values()), 2)
    total_gp = round(sum(item["gp"] for item in daily_rollup.values()), 2)

    return success_response(
        {
            "title": "Monthly Install Completion",
            "year": year,
            "month": month,
            "period": {
                "start_date": effective_start_date.isoformat() if effective_start_date else None,
                "end_date": effective_end_date.isoformat() if effective_end_date else None,
            },
            "filters": {
                "from_date": effective_start_date.isoformat() if effective_start_date else None,
                "to_date": effective_end_date.isoformat() if effective_end_date else None,
                "fab_id": fab_id,
                "job_number": job_number,
                "installer_name": installer_name,
                "fab_type": fab_type,
            },
            "columns": [
                "install_date",
                "fab_type",
                "fab_id",
                "job_number",
                "installer_id",
                "installer_name",
                "job_name",
                "account_name",
                "stone_type_name",
                "stone_color_name",
                "edge_name",
                "stone_thickness_value",
                "input_area",
                "pieces",
                "sq_ft",
                "revenue",
                "revenue_per_sq_ft",
                "cost_of_stone",
                "gp",
            ],
            "summary": {
                "pieces": int(sum(item["pieces"] for item in daily_rollup.values())),
                "sq_ft": total_sqft,
                "revenue": total_revenue,
                "revenue_per_sq_ft": round((total_revenue / total_sqft), 2) if total_sqft else 0.0,
                "cost_of_stone": total_cost,
                "gp": total_gp,
                "total_cost_of_stone": total_cost,
                "total_gp": total_gp,
                "row_count": len(rows),
            },
            "daily_totals": daily_totals,
            "rows": rows,
        },
        "Owner monthly install completion report generated",
    )



@router.get("/reports/owner/monthly-cut-completion", response_model=SuccessResponse[dict])
async def get_owner_monthly_cut_completion_report(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    from_date: Optional[date] = Query(None, description="Optional from date filter (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="Optional to date filter (YYYY-MM-DD)"),
    fab_id: Optional[int] = Query(None, gt=0, description="Optional FAB ID filter"),
    job_number: Optional[str] = Query(None, description="Optional job number filter"),
    installer_name: Optional[str] = Query(None, description="Optional installer name filter"),
    fab_type: Optional[str] = Query(None, description="Optional FAB type filter"),
    fab_type_sort: str = Query("asc", pattern="^(asc|desc)$", description="Sort order for FAB type"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Monthly cut completion report modeled after legacy spreadsheet format."""
    month_start_dt, month_end_dt = _month_bounds(year, month)
    effective_start_date = from_date if from_date is not None else month_start_dt.date()
    effective_end_date = to_date if to_date is not None else month_end_dt.date()
    if effective_start_date > effective_end_date:
        effective_start_date, effective_end_date = effective_end_date, effective_start_date
    start_dt, end_dt = _range_bounds(effective_start_date, effective_end_date)
    cut_date_expr = func.coalesce(Fab.shop_date_schedule, Fab.final_programming_completed_date)
    fab_type_sort_normalized = (fab_type_sort or "asc").strip().lower()
    fab_type_order = (
        func.lower(Fab.fab_type).desc()
        if fab_type_sort_normalized == "desc"
        else func.lower(Fab.fab_type).asc()
    )

    query = (
        select(
            cut_date_expr.label("cut_date"),
            Fab.fab_type,
            Fab.id,
            BusinessJob.job_number,
            BusinessJob.name.label("job_name"),
            Account.name.label("account_name"),
            StoneType.name.label("stone_type_name"),
            StoneColor.name.label("stone_color_name"),
            Edge.name.label("edge_name"),
            StoneThickness.thickness.label("stone_thickness_value"),
            Fab.input_area,
            Fab.no_of_pieces,
            Fab.total_sqft,
            Fab.revenue,
            Fab.cost_of_stone,
            CostOfStone.total_cost,
        )
        .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
        .join(Account, Account.id == BusinessJob.account_id, isouter=True)
        .join(StoneType, StoneType.id == Fab.stone_type_id, isouter=True)
        .join(StoneColor, StoneColor.id == Fab.stone_color_id, isouter=True)
        .join(Edge, Edge.id == Fab.edge_id, isouter=True)
        .join(StoneThickness, StoneThickness.id == Fab.stone_thickness_id, isouter=True)
        .join(CostOfStone, CostOfStone.id == Fab.cost_of_stone_id, isouter=True)
        .join(CutList, CutList.fab_id == Fab.id, isouter=True)
        .where(cut_date_expr >= start_dt, cut_date_expr <= end_dt)
        .order_by(cut_date_expr.asc(), fab_type_order, Fab.id.asc())
    )

    if fab_type:
        query = query.where(func.lower(Fab.fab_type) == fab_type.strip().lower())
    if fab_id is not None:
        query = query.where(Fab.id == fab_id)
    if job_number:
        query = query.where(BusinessJob.job_number.ilike(f"%{job_number.strip()}%"))
    if installer_name:
        installer_like = f"%{installer_name.strip().lower()}%"
        installer_exists = (
            select(ShopCutPlan.id)
            .join(User, User.id == ShopCutPlan.user_id)
            .where(
                ShopCutPlan.fab_id == Fab.id,
                or_(
                    func.lower(User.first_name).like(installer_like),
                    func.lower(User.last_name).like(installer_like),
                    func.lower(User.username).like(installer_like),
                    func.lower(User.email).like(installer_like),
                    func.lower(
                        func.concat(
                            func.coalesce(User.first_name, ""),
                            " ",
                            func.coalesce(User.last_name, ""),
                        )
                    ).like(installer_like),
                ),
            )
            .limit(1)
            .exists()
        )
        query = query.where(installer_exists)

    records = (await db.execute(query)).all()

    rows = []
    daily_rollup: dict[str, dict] = defaultdict(lambda: {
        "pieces": 0,
        "sq_ft": 0.0,
        "revenue": 0.0,
        "cost_of_stone": 0.0,
        "gp": 0.0,
        "row_count": 0,
    })

    for cut_date, row_fab_type, fab_id, job_number, job_name, account_name, stone_type_name, stone_color_name, edge_name, stone_thickness_value, input_area, pieces, sq_ft, revenue, fab_cost_of_stone, cos_total_cost in records:
        if cut_date is None:
            continue

        sq_ft_value = round(_to_float(sq_ft), 2)
        revenue_value = round(_to_float(revenue), 2)
        cost_value = round(_to_float(fab_cost_of_stone if fab_cost_of_stone is not None else cos_total_cost), 2)
        gp_value = round(revenue_value - cost_value, 2)
        pieces_value = int(_to_float(pieces))
        revenue_per_sqft = round((revenue_value / sq_ft_value), 2) if sq_ft_value else 0.0
        day_key = cut_date.date().isoformat()

        daily_rollup[day_key]["pieces"] += pieces_value
        daily_rollup[day_key]["sq_ft"] += sq_ft_value
        daily_rollup[day_key]["revenue"] += revenue_value
        daily_rollup[day_key]["cost_of_stone"] += cost_value
        daily_rollup[day_key]["gp"] += gp_value
        daily_rollup[day_key]["row_count"] += 1

        rows.append(
            {
                "cut_date": day_key,
                "fab_type": row_fab_type,
                "fab_id": fab_id,
                "job_number": job_number,
                "job_name": job_name,
                "account_name": account_name,
                "stone_type_name": stone_type_name,
                "stone_color_name": stone_color_name,
                "edge_name": edge_name,
                "stone_thickness_value": stone_thickness_value,
                "input_area": input_area,
                "pieces": pieces_value,
                "sq_ft": sq_ft_value,
                "revenue": revenue_value,
                "revenue_per_sq_ft": revenue_per_sqft,
                "cost_of_stone": cost_value,
                "gp": gp_value,
            }
        )

    daily_totals = []
    for day_key in sorted(daily_rollup.keys()):
        item = daily_rollup[day_key]
        day_sqft = round(item["sq_ft"], 2)
        day_revenue = round(item["revenue"], 2)
        daily_totals.append(
            {
                "cut_date": day_key,
                "pieces": int(item["pieces"]),
                "sq_ft": day_sqft,
                "revenue": day_revenue,
                "revenue_per_sq_ft": round((day_revenue / day_sqft), 2) if day_sqft else 0.0,
                "cost_of_stone": round(item["cost_of_stone"], 2),
                "gp": round(item["gp"], 2),
                "row_count": int(item["row_count"]),
            }
        )

    total_sqft = round(sum(item["sq_ft"] for item in daily_rollup.values()), 2)
    total_revenue = round(sum(item["revenue"] for item in daily_rollup.values()), 2)
    total_cost = round(sum(item["cost_of_stone"] for item in daily_rollup.values()), 2)
    total_gp = round(sum(item["gp"] for item in daily_rollup.values()), 2)

    return success_response(
        {
            "title": "Monthly Cut Completion",
            "year": year,
            "month": month,
            "period": {
                "start_date": effective_start_date.isoformat() if effective_start_date else None,
                "end_date": effective_end_date.isoformat() if effective_end_date else None,
            },
            "filters": {
                "from_date": effective_start_date.isoformat() if effective_start_date else None,
                "to_date": effective_end_date.isoformat() if effective_end_date else None,
                "fab_id": fab_id,
                "job_number": job_number,
                "installer_name": installer_name,
                "fab_type": fab_type,
            },
            "columns": [
                "cut_date",
                "fab_type",
                "fab_id",
                "job_number",
                "job_name",
                "account_name",
                "stone_type_name",
                "stone_color_name",
                "edge_name",
                "stone_thickness_value",
                "input_area",
                "pieces",
                "sq_ft",
                "revenue",
                "revenue_per_sq_ft",
                "cost_of_stone",
                "gp",
            ],
            "summary": {
                "pieces": int(sum(item["pieces"] for item in daily_rollup.values())),
                "sq_ft": total_sqft,
                "revenue": total_revenue,
                "revenue_per_sq_ft": round((total_revenue / total_sqft), 2) if total_sqft else 0.0,
                "cost_of_stone": total_cost,
                "gp": total_gp,
                "total_cost_of_stone": total_cost,
                "total_gp": total_gp,
                "row_count": len(rows),
            },
            "daily_totals": daily_totals,
            "rows": rows,
        },
        "Owner monthly cut completion report generated",
    )


@router.patch("/reports/owner/monthly-cut-completion/{monthly_cut_completion_id}", response_model=SuccessResponse[dict])
async def patch_owner_monthly_cut_completion(
    monthly_cut_completion_id: int,
    patch: MonthlyCutCompletionPatchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update monthly cut completion row fields by FAB id."""
    if patch.revenue is None and patch.cost_of_stone is None and patch.revenue_per_sq_ft is None:
        raise error_response("At least one field is required", 400)

    fab = (await db.execute(select(Fab).where(Fab.id == monthly_cut_completion_id))).scalar_one_or_none()
    if not fab:
        raise error_response("Monthly cut completion record not found", 404)

    now = datetime.now()

    if patch.revenue is not None:
        fab.revenue = round(patch.revenue, 2)
    elif patch.revenue_per_sq_ft is not None:
        sqft_value = _to_float(fab.total_sqft)
        if sqft_value <= 0:
            raise error_response("Cannot apply revenue_per_sq_ft when sq_ft is missing or zero", 400)
        fab.revenue = round(patch.revenue_per_sq_ft * sqft_value, 2)

    if patch.cost_of_stone is not None:
        fab.cost_of_stone = round(patch.cost_of_stone, 2)
        cost_record = None
        if fab.cost_of_stone_id:
            cost_record = await db.get(CostOfStone, fab.cost_of_stone_id)
        if cost_record is None:
            cost_record = (
                await db.execute(select(CostOfStone).where(CostOfStone.fab_id == fab.id))
            ).scalar_one_or_none()
        if cost_record is None:
            cost_record = CostOfStone(
                fab_id=fab.id,
                stone_color_id=fab.stone_color_id,
                stone_type_id=fab.stone_type_id,
                total_sqft=str(fab.total_sqft) if fab.total_sqft is not None else None,
                total_cost=f"{patch.cost_of_stone:.2f}",
                status_id=1,
                created_at=now,
                updated_at=now,
                updated_by=current_user.id,
            )
            db.add(cost_record)
            await db.flush()
            fab.cost_of_stone_id = cost_record.id
        else:
            cost_record.total_cost = f"{patch.cost_of_stone:.2f}"
            cost_record.updated_at = now
            cost_record.updated_by = current_user.id

    fab.updated_at = now
    fab.updated_by = current_user.id
    await db.commit()

    current_sqft = round(_to_float(fab.total_sqft), 2)
    current_revenue = round(_to_float(fab.revenue), 2)
    current_cost = round(_to_float(fab.cost_of_stone), 2)

    return success_response(
        {
            "monthly_cut_completion_id": monthly_cut_completion_id,
            "fab_id": fab.id,
            "revenue": current_revenue,
            "cost_of_stone": current_cost,
            "revenue_per_sq_ft": round((current_revenue / current_sqft), 2) if current_sqft else 0.0,
        },
        "Monthly cut completion record updated successfully",
    )


@router.patch("/reports/owner/monthly-install-completion/{monthly_install_completion_id}", response_model=SuccessResponse[dict])
async def patch_owner_monthly_install_completion(
    monthly_install_completion_id: int,
    patch: MonthlyInstallCompletionPatchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update monthly install completion row fields by FAB id."""
    if (
        patch.revenue is None
        and patch.sq_ft is None
        and patch.revenue_per_sq_ft is None
        and patch.installer_id is None
        and patch.installer_name is None
    ):
        raise error_response("At least one field is required", 400)

    fab = (await db.execute(select(Fab).where(Fab.id == monthly_install_completion_id))).scalar_one_or_none()
    if not fab:
        raise error_response("Monthly install completion record not found", 404)

    completion = await _get_latest_install_completion_for_fab(db, fab.id)
    if completion is None:
        raise error_response("Install completion record not found for this FAB", 404)

    now = datetime.now()

    if patch.sq_ft is not None:
        completion.total_sqft_installed = f"{patch.sq_ft:.2f}"

    if patch.revenue is not None:
        fab.revenue = round(patch.revenue, 2)
    elif patch.revenue_per_sq_ft is not None:
        sqft_value = patch.sq_ft if patch.sq_ft is not None else _to_float(completion.total_sqft_installed)
        if sqft_value <= 0:
            raise error_response("Cannot apply revenue_per_sq_ft when sq_ft is missing or zero", 400)
        fab.revenue = round(patch.revenue_per_sq_ft * sqft_value, 2)

    installer = None
    if patch.installer_id is not None:
        installer = await db.get(User, patch.installer_id)
        if installer is None:
            raise error_response("Installer not found", 404)
        completion.installer_id = installer.id
    elif patch.installer_name is not None:
        installer = await _resolve_user_by_name(db, patch.installer_name)
        if installer is None:
            raise error_response("Installer not found", 404)
        completion.installer_id = installer.id

    completion.updated_at = now
    completion.updated_by = current_user.id
    fab.updated_at = now
    fab.updated_by = current_user.id

    await db.commit()

    if installer is None and completion.installer_id:
        installer = await db.get(User, completion.installer_id)

    installer_name = None
    if installer is not None:
        installer_name = f"{(installer.first_name or '').strip()} {(installer.last_name or '').strip()}".strip() or f"Installer {installer.id}"

    current_sqft = round(_to_float(completion.total_sqft_installed), 2)
    current_revenue = round(_to_float(fab.revenue), 2)

    return success_response(
        {
            "monthly_install_completion_id": monthly_install_completion_id,
            "fab_id": fab.id,
            "revenue": current_revenue,
            "sq_ft": current_sqft,
            "revenue_per_sq_ft": round((current_revenue / current_sqft), 2) if current_sqft else 0.0,
            "installer_id": completion.installer_id,
            "installer_name": installer_name,
        },
        "Monthly install completion record updated successfully",
    )


@router.patch("/reports/owner/daily-install-completion/{daily_install_completion_id}", response_model=SuccessResponse[dict])
async def patch_owner_daily_install_completion(
    daily_install_completion_id: int,
    patch: DailyInstallCompletionPatchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update daily install completion row fields by FAB id."""
    if patch.revenue is None and patch.sq_ft is None and patch.installer_id is None and patch.installer_name is None:
        raise error_response("At least one field is required", 400)

    fab = (await db.execute(select(Fab).where(Fab.id == daily_install_completion_id))).scalar_one_or_none()
    if not fab:
        raise error_response("Daily install completion record not found", 404)

    completion = await _get_latest_install_completion_for_fab(db, fab.id)
    if completion is None:
        raise error_response("Install completion record not found for this FAB", 404)

    now = datetime.now()

    if patch.revenue is not None:
        fab.revenue = round(patch.revenue, 2)
    if patch.sq_ft is not None:
        completion.total_sqft_installed = f"{patch.sq_ft:.2f}"

    installer = None
    if patch.installer_id is not None:
        installer = await db.get(User, patch.installer_id)
        if installer is None:
            raise error_response("Installer not found", 404)
        completion.installer_id = installer.id
    elif patch.installer_name is not None:
        installer = await _resolve_user_by_name(db, patch.installer_name)
        if installer is None:
            raise error_response("Installer not found", 404)
        completion.installer_id = installer.id

    completion.updated_at = now
    completion.updated_by = current_user.id
    fab.updated_at = now
    fab.updated_by = current_user.id

    await db.commit()

    if installer is None and completion.installer_id:
        installer = await db.get(User, completion.installer_id)

    installer_name = None
    if installer is not None:
        installer_name = f"{(installer.first_name or '').strip()} {(installer.last_name or '').strip()}".strip() or f"Installer {installer.id}"

    return success_response(
        {
            "daily_install_completion_id": daily_install_completion_id,
            "fab_id": fab.id,
            "revenue": round(_to_float(fab.revenue), 2),
            "sq_ft": round(_to_float(completion.total_sqft_installed), 2),
            "installer_id": completion.installer_id,
            "installer_name": installer_name,
        },
        "Daily install completion record updated successfully",
    )

@router.get("/reports/owner/daily-install-completion", response_model=SuccessResponse[dict])
@router.get("/reports/daily-install-completion", response_model=SuccessResponse[dict])
async def get_daily_install_completion_report(
    start_date: Optional[date] = Query(None, description="Inclusive start date filter"),
    end_date: Optional[date] = Query(None, description="Inclusive end date filter"),
    job_number: Optional[str] = Query(None, description="Optional job number filter"),
    fab_type: Optional[str] = Query(None, description="Optional FAB type filter"),
    fab_id: Optional[int] = Query(None, gt=0, description="Optional FAB ID filter"),
    installer_id: Optional[int] = Query(None, gt=0, description="Optional installer user ID filter"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Daily install completion report based on completed installs with optional filters and daily/grand totals."""
    effective_end_date = end_date or start_date or date.today()
    effective_start_date = start_date or end_date or (effective_end_date - timedelta(days=6))
    if effective_start_date > effective_end_date:
        effective_start_date, effective_end_date = effective_end_date, effective_start_date

    start_dt, end_dt = _range_bounds(effective_start_date, effective_end_date)

    query = (
        select(
            InstallCompletion.completion_date,
            Fab.id,
            Fab.fab_type,
            BusinessJob.job_number,
            BusinessJob.name,
            Account.name,
            StoneType.name,
            StoneColor.name,
            StoneThickness.thickness,
            Edge.name,
            Fab.input_area,
            InstallCompletion.installer_id,
            User.first_name,
            User.last_name,
            Fab.total_sqft,
            Fab.revenue,
            Fab.cost_of_stone,
            CostOfStone.total_cost,
        )
        .select_from(InstallCompletion)
        .join(Fab, Fab.id == InstallCompletion.fab_id)
        .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
        .join(Account, Account.id == BusinessJob.account_id, isouter=True)
        .join(StoneType, StoneType.id == Fab.stone_type_id, isouter=True)
        .join(StoneColor, StoneColor.id == Fab.stone_color_id, isouter=True)
        .join(StoneThickness, StoneThickness.id == Fab.stone_thickness_id, isouter=True)
        .join(Edge, Edge.id == Fab.edge_id, isouter=True)
        .join(CostOfStone, CostOfStone.id == Fab.cost_of_stone_id, isouter=True)
        .join(User, User.id == InstallCompletion.installer_id, isouter=True)
        .where(
            InstallCompletion.completion_date.is_not(None),
            InstallCompletion.is_completed.is_(True),
        )
        .order_by(InstallCompletion.completion_date.asc(), Fab.id.asc())
    )

    if start_dt is not None:
        query = query.where(InstallCompletion.completion_date >= start_dt)
    if end_dt is not None:
        query = query.where(InstallCompletion.completion_date <= end_dt)
    if job_number:
        query = query.where(BusinessJob.job_number.ilike(f"%{job_number.strip()}%"))
    if fab_type:
        query = query.where(func.lower(Fab.fab_type) == fab_type.strip().lower())
    if fab_id is not None:
        query = query.where(Fab.id == fab_id)
    if installer_id is not None:
        query = query.where(InstallCompletion.installer_id == installer_id)

    records = (await db.execute(query)).all()

    entries = []
    daily_totals_map: dict[str, dict] = defaultdict(lambda: {
        "total_sqft": 0.0,
        "total_revenue": 0.0,
        "total_cost_of_stone": 0.0,
        "total_gp": 0.0,
        "count": 0,
    })

    grand_total_sqft = 0.0
    grand_total_revenue = 0.0
    grand_total_cost_of_stone = 0.0
    grand_total_gp = 0.0

    for (
        completion_date,
        row_fab_id,
        row_fab_type,
        row_job_number,
        row_job_name,
        row_account_name,
        row_stone_type_name,
        row_stone_color_name,
        row_stone_thickness_value,
        row_edge_name,
        row_input_area,
        row_installer_id,
        installer_first_name,
        installer_last_name,
        sqft_installed_raw,
        revenue_raw,
        fab_cost_of_stone,
        cos_total_cost,
    ) in records:
        day_key = completion_date.date().isoformat()
        sqft_value = round(_to_float(sqft_installed_raw), 2)
        revenue_value = round(_to_float(revenue_raw), 2)
        cost_value = round(_to_float(fab_cost_of_stone if fab_cost_of_stone is not None else cos_total_cost), 2)
        gp_value = round(revenue_value - cost_value, 2)

        daily_totals_map[day_key]["total_sqft"] += sqft_value
        daily_totals_map[day_key]["total_revenue"] += revenue_value
        daily_totals_map[day_key]["total_cost_of_stone"] += cost_value
        daily_totals_map[day_key]["total_gp"] += gp_value
        daily_totals_map[day_key]["count"] += 1

        grand_total_sqft += sqft_value
        grand_total_revenue += revenue_value
        grand_total_cost_of_stone += cost_value
        grand_total_gp += gp_value

        installer_name = (
            f"{(installer_first_name or '').strip()} {(installer_last_name or '').strip()}".strip()
            or (f"User {row_installer_id}" if row_installer_id else "Unknown")
        )
        entries.append(
            {
                "install_date": day_key,
                "fab_id": row_fab_id,
                "fab_type": row_fab_type,
                "job_number": row_job_number,
                "job_name": row_job_name,
                "account_name": row_account_name,
                "stone_type_name": row_stone_type_name,
                "stone_color_name": row_stone_color_name,
                "stone_thickness_value": row_stone_thickness_value,
                "edge_name": row_edge_name,
                "input_area": row_input_area,
                "installer_id": row_installer_id,
                "installer_name": installer_name,
                "sqft": sqft_value,
                "revenue": revenue_value,
                "cost_of_stone": cost_value,
                "gp": gp_value,
            }
        )

    daily_totals = []
    for day_key in sorted(daily_totals_map.keys()):
        item = daily_totals_map[day_key]
        daily_totals.append(
            {
                "install_date": day_key,
                "total_sqft": round(item["total_sqft"], 2),
                "total_revenue": round(item["total_revenue"], 2),
                "total_cost_of_stone": round(item["total_cost_of_stone"], 2),
                "total_gp": round(item["total_gp"], 2),
                "entry_count": int(item["count"]),
            }
        )

    return success_response(
        {
            "period": {
                "start_date": effective_start_date.isoformat() if effective_start_date else None,
                "end_date": effective_end_date.isoformat() if effective_end_date else None,
            },
            "filters": {
                "job_number": job_number,
                "fab_type": fab_type,
                "fab_id": fab_id,
                "installer_id": installer_id,
            },
            "grand_totals": {
                "total_sqft": round(grand_total_sqft, 2),
                "total_revenue": round(grand_total_revenue, 2),
                "total_cost_of_stone": round(grand_total_cost_of_stone, 2),
                "total_gp": round(grand_total_gp, 2),
                "entry_count": len(entries),
            },
            "daily_totals": daily_totals,
            "entries": entries,
        },
        "Daily Install Completion data retrieved successfully",
    )


@router.get("/reports/monthly-cut-completion", response_model=SuccessResponse[dict])
async def get_monthly_cut_completion_report(
    month: str = Query(..., description="Month number (1-12) or month name (e.g. April)"),
    year: int = Query(..., ge=2000, le=2100),
    fab_type: Optional[str] = Query(None, description="Optional FAB type filter"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Monthly cut completion report with summary and detailed entries."""
    month_num = _parse_month_input(month)
    if month_num is None:
        return success_response(None, "Invalid month. Use month number (1-12) or month name", status_code=400)

    start_dt, end_dt = _month_bounds(year, month_num)
    cut_date_expr = func.coalesce(Fab.shop_date_schedule, Fab.final_programming_completed_date)
    cut_or_wj_plan_exists = (
        select(ShopCutPlan.id)
        .join(PlanningSection, PlanningSection.id == ShopCutPlan.planning_section_id)
        .where(
            ShopCutPlan.fab_id == Fab.id,
            func.lower(func.trim(PlanningSection.plan_name)).in_(["cut", "wj"]),
        )
        .exists()
    )

    query = (
        select(
            cut_date_expr.label("cut_date"),
            Fab.fab_type,
            Fab.id,
            BusinessJob.job_number,
            Account.name.label("account_name"),
            StoneType.name.label("stone_type_name"),
            StoneColor.name.label("stone_color_name"),
            StoneThickness.thickness.label("stone_thickness"),
            Fab.no_of_pieces,
            Fab.total_sqft,
            CostOfStone.cost_per_sqft,
            Fab.revenue,
            Fab.cost_of_stone,
            CostOfStone.total_cost,
            Fab.gp,
        )
        .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
        .join(Account, Account.id == BusinessJob.account_id, isouter=True)
        .join(StoneType, StoneType.id == Fab.stone_type_id, isouter=True)
        .join(StoneColor, StoneColor.id == Fab.stone_color_id, isouter=True)
        .join(StoneThickness, StoneThickness.id == Fab.stone_thickness_id, isouter=True)
        .join(CostOfStone, CostOfStone.id == Fab.cost_of_stone_id, isouter=True)
        .where(cut_date_expr >= start_dt, cut_date_expr <= end_dt, cut_or_wj_plan_exists)
        .order_by(cut_date_expr.asc(), Fab.id.asc())
    )

    if fab_type:
        query = query.where(func.lower(Fab.fab_type) == fab_type.strip().lower())

    records = (await db.execute(query)).all()

    entries = []
    total_pieces = 0
    total_sqft = 0.0
    total_revenue = 0.0
    total_cost_of_stone = 0.0
    total_gp = 0.0

    for (
        cut_date,
        row_fab_type,
        fab_id,
        job_number,
        account_name,
        stone_type_name,
        stone_color_name,
        stone_thickness,
        no_of_pieces,
        sqft,
        cost_per_sqft_raw,
        revenue_raw,
        fab_cost_of_stone,
        cos_total_cost,
        gp_raw,
    ) in records:
        if cut_date is None:
            continue

        pieces_value = int(_to_float(no_of_pieces))
        sqft_value = round(_to_float(sqft), 2)
        cost_per_sf = round(_to_float(cost_per_sqft_raw), 2)
        revenue_value = round((-cost_per_sf * sqft_value), 2)
        cost_of_stone_value = round(_to_float(fab_cost_of_stone if fab_cost_of_stone is not None else cos_total_cost), 2)
        gp_value = round((-revenue_value - cost_of_stone_value), 2)

        total_pieces += pieces_value
        total_sqft += sqft_value
        total_revenue += revenue_value
        total_cost_of_stone += cost_of_stone_value
        total_gp += gp_value

        info_parts = [account_name, stone_type_name, stone_color_name, stone_thickness]
        fab_info = " - ".join(str(part).strip() for part in info_parts if part and str(part).strip())

        entries.append(
            {
                "cut_date": cut_date.date().isoformat(),
                "fab_type": row_fab_type,
                "fab_id": fab_id,
                "job_number": job_number,
                "fab_info": fab_info or None,
                "no_of_pieces": pieces_value,
                "sqft": sqft_value,
                "cost/sf": cost_per_sf,
                "revenue": revenue_value,
                "revenue_per_sqft": round((revenue_value / sqft_value), 2) if sqft_value else 0.0,
                "cost_of_stone": cost_of_stone_value,
                "gp": gp_value,
            }
        )

    total_sqft = round(total_sqft, 2)
    total_revenue = round(total_revenue, 2)
    total_cost_of_stone = round(total_cost_of_stone, 2)
    total_gp = round(total_gp, 2)

    return success_response(
        {
            "month": calendar.month_name[month_num],
            "year": year,
            "summary": {
                "total_no_of_pieces": int(total_pieces),
                "total_sqft": total_sqft,
                "total_revenue": total_revenue,
                "total_revenue/sqft": round((total_revenue / total_sqft), 2) if total_sqft else 0.0,
                "total_cost_of_stone": total_cost_of_stone,
                "total_gp": total_gp,
            },
            "entries": entries,
        },
        "Monthly Cut Completion data retrieved successfully",
    )


@router.get("/reports/owner/turnaround-times", response_model=SuccessResponse[dict])
async def get_owner_turnaround_times_report(
    year: Optional[int] = Query(None, ge=2000, le=2100, description="Required with month when from_date/to_date is not provided"),
    month: Optional[int] = Query(None, ge=1, le=12, description="Required with year when from_date/to_date is not provided"),
    from_date: Optional[date] = Query(None, description="Inclusive from date filter (YYYY-MM-DD)"),
    to_date: Optional[date] = Query(None, description="Inclusive to date filter (YYYY-MM-DD)"),
    fab_id: Optional[int] = Query(None, gt=0, description="Optional FAB ID filter"),
    job_number: Optional[str] = Query(None, description="Optional job number filter"),
    fab_type: Optional[str] = Query(None, description="Optional FAB type filter"),
    limit: int = Query(2000, ge=1, le=10000),
    threshold_days: Optional[int] = Query(
        None,
        ge=0,
        description="Optional threshold used to list jobs that exceeded the specified days in each stage and total days",
    ),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Turnaround-times report with full stage dates/days and stage statistics."""
    if from_date is not None or to_date is not None:
        effective_from_date = from_date or to_date
        effective_to_date = to_date or from_date
        start_dt, end_dt = _range_bounds(effective_from_date, effective_to_date)
        report_year = effective_from_date.year if effective_from_date is not None else None
        report_month = effective_from_date.month if effective_from_date is not None else None
    else:
        if (year is None) != (month is None):
            return success_response(
                None,
                "Provide both year and month, or use from_date/to_date",
                status_code=400,
            )
        if year is None or month is None:
            return success_response(
                None,
                "Either provide year and month, or provide from_date/to_date",
                status_code=400,
            )
        start_dt, end_dt = _month_bounds(year, month)
        report_year = year
        report_month = month

    cut_end_subquery = (
        select(
            ShopCutPlan.fab_id.label("fab_id"),
            func.max(ShopCutPlan.actual_end_date).label("cut_end_date"),
        )
        .join(PlanningSection, PlanningSection.id == ShopCutPlan.planning_section_id)
        .where(func.lower(func.trim(PlanningSection.plan_name)).in_(["cut", "wj"]))
        .group_by(ShopCutPlan.fab_id)
        .subquery()
    )

    cnc_end_subquery = (
        select(
            CNCDrafting.fab_id.label("fab_id"),
            func.max(CNCDrafting.drafter_end_date).label("cnc_date"),
        )
        .group_by(CNCDrafting.fab_id)
        .subquery()
    )

    revision_start_subquery = (
        select(
            Revision.fab_id.label("fab_id"),
            func.max(func.coalesce(Revision.actual_start_date, Revision.scheduled_start_date, Revision.created_at)).label("revision_start_date"),
        )
        .group_by(Revision.fab_id)
        .subquery()
    )

    query = (
        select(
            Fab.id,
            BusinessJob.job_number,
            BusinessJob.name.label("job_name"),
            Account.name.label("account_name"),
            StoneType.name.label("stone_type_name"),
            StoneColor.name.label("stone_color_name"),
            StoneThickness.thickness.label("stone_thickness"),
            Fab.no_of_pieces,
            Fab.total_sqft,
            Fab.template_completed_date,
            Fab.predraft_completed_date,
            Fab.draft_completed_date,
            Fab.slabsmith_completed_date,
            Fab.revision_completed_date,
            revision_start_subquery.c.revision_start_date,
            func.coalesce(Fab.sales_ct_completed_date, Fab.sct_completed_date).label("sct_date"),
            Fab.final_programming_completed_date,
            cnc_end_subquery.c.cnc_date,
            Fab.shop_date_schedule,
            cut_end_subquery.c.cut_end_date,
            Fab.shop_est_completion_date,
        )
        .select_from(Fab)
        .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
        .join(Account, Account.id == BusinessJob.account_id, isouter=True)
        .join(StoneType, StoneType.id == Fab.stone_type_id, isouter=True)
        .join(StoneColor, StoneColor.id == Fab.stone_color_id, isouter=True)
        .join(StoneThickness, StoneThickness.id == Fab.stone_thickness_id, isouter=True)
        .join(cut_end_subquery, cut_end_subquery.c.fab_id == Fab.id, isouter=True)
        .join(cnc_end_subquery, cnc_end_subquery.c.fab_id == Fab.id, isouter=True)
        .join(revision_start_subquery, revision_start_subquery.c.fab_id == Fab.id, isouter=True)
        .where(Fab.shop_est_completion_date.is_not(None))
        .order_by(Fab.shop_est_completion_date.desc(), Fab.id.asc())
        .limit(limit)
    )

    if start_dt is not None:
        query = query.where(Fab.shop_est_completion_date >= start_dt)
    if end_dt is not None:
        query = query.where(Fab.shop_est_completion_date <= end_dt)
    if fab_id is not None:
        query = query.where(Fab.id == fab_id)
    if job_number:
        query = query.where(BusinessJob.job_number.ilike(f"%{job_number.strip()}%"))
    if fab_type:
        query = query.where(func.lower(Fab.fab_type) == fab_type.strip().lower())

    records = (await db.execute(query)).all()
    rows: list[dict] = []

    day_columns = {
        "predraft_days": [],
        "draft_days": [],
        "slabsmith_days": [],
        "revision_days": [],
        "sct_days": [],
        "final_prog_days": [],
        "cnc_days": [],
        "cut_days": [],
        "fab_days": [],
        "total_days": [],
    }
    stage_labels = {
        "predraft_days": "predraft",
        "draft_days": "draft",
        "slabsmith_days": "slabsmith",
        "revision_days": "revision",
        "sct_days": "sct",
        "final_prog_days": "final_programming",
        "cnc_days": "cnc",
        "cut_days": "cut",
        "fab_days": "fab",
        "total_days": "total",
    }

    for (
        fab_id,
        job_number,
        job_name,
        account_name,
        stone_type_name,
        stone_color_name,
        stone_thickness,
        pieces,
        total_sqft,
        template_date,
        predraft_date,
        draft_date,
        slabsmith_date,
        revision_date,
        revision_start_date,
        sct_date,
        final_prog_date,
        cnc_date,
        cut_date,
        cut_end_date,
        fab_complete_date,
    ) in records:
        predraft_days = _days_between(template_date, predraft_date)
        # Draft stage should count any positive partial day (e.g. minutes) as 1 day.
        draft_days = _days_between_count_partial_day_as_full(predraft_date, draft_date)
        slabsmith_days = _days_between(draft_date, slabsmith_date)
        revision_days = _days_between(revision_start_date, revision_date)
        sct_days = _days_between(draft_date, sct_date)
        final_prog_days = _days_between(sct_date, final_prog_date)
        cnc_days = _days_between(final_prog_date, cnc_date)
        cut_days = _days_between(cut_date, cut_end_date)
        fab_days = _days_between(cut_end_date, fab_complete_date)

        total_days = sum(
            value or 0
            for value in [
                predraft_days,
                draft_days,
                slabsmith_days,
                revision_days,
                sct_days,
                final_prog_days,
                cnc_days,
                cut_days,
            ]
        )

        info_parts = [account_name, job_name, stone_type_name, stone_color_name, stone_thickness]
        fab_info = " - ".join(str(part).strip() for part in info_parts if part and str(part).strip()) or None

        row = {
            "fab_id": fab_id,
            "job_number": job_number,
            "fab_info": fab_info,
            "no_of_pieces": int(_to_float(pieces)),
            "total_sqft": round(_to_float(total_sqft), 2),
            "template_date": template_date.isoformat() if template_date else None,
            "predraft_date": predraft_date.isoformat() if predraft_date else None,
            "predraft_days": predraft_days,
            "draft_date": draft_date.isoformat() if draft_date else None,
            "draft_days": draft_days,
            "slabsmith_date": slabsmith_date.isoformat() if slabsmith_date else None,
            "slabsmith_days": slabsmith_days,
            "revision_date": revision_date.isoformat() if revision_date else None,
            "revision_start_date": revision_start_date.isoformat() if revision_start_date else None,
            "revision_days": revision_days,
            "sct_date": sct_date.isoformat() if sct_date else None,
            "sct_days": sct_days,
            "final_prog_date": final_prog_date.isoformat() if final_prog_date else None,
            "final_prog_days": final_prog_days,
            "cnc_date": cnc_date.isoformat() if cnc_date else None,
            "cnc_days": cnc_days,
            "cut_date": cut_date.isoformat() if cut_date else None,
            "cut_end_date": cut_end_date.isoformat() if cut_end_date else None,
            "cut_days": cut_days,
            "fab_complete_date": fab_complete_date.isoformat() if fab_complete_date else None,
            "fab_days": fab_days,
            "total_days": total_days,
        }
        rows.append(row)

        for key in day_columns.keys():
            value = row.get(key)
            if value is not None:
                day_columns[key].append(
                    {
                        "fab_id": fab_id,
                        "job_number": job_number,
                        "fab_info": fab_info,
                        "value": int(value),
                    }
                )

    def _stage_stats(values: list[dict]) -> dict:
        if not values:
            return {
                "average_days": None,
                "highest": None,
                "lowest": None,
            }

        sorted_values = sorted(values, key=lambda item: item["value"])
        average_days = round(sum(item["value"] for item in values) / len(values), 2)

        return {
            "average_days": average_days,
            "highest": sorted_values[-1],
            "lowest": sorted_values[0],
        }

    summary = {
        key: _stage_stats(values)
        for key, values in day_columns.items()
    }
    summary["row_count"] = len(rows)

    stage_averages = {
        stage_labels[key]: stats["average_days"]
        for key, stats in summary.items()
        if key in stage_labels
    }

    threshold_analysis = None
    if threshold_days is not None:
        jobs_by_stage = {}
        counts_by_stage = {}

        for key, values in day_columns.items():
            stage_name = stage_labels[key]
            exceeded = [
                {
                    "fab_id": item["fab_id"],
                    "job_number": item["job_number"],
                    "fab_info": item.get("fab_info"),
                    "days": item["value"],
                    "days_over_threshold": item["value"] - threshold_days,
                }
                for item in values
                if item["value"] > threshold_days
            ]
            exceeded.sort(key=lambda item: (item["days"], item["job_number"] or "", item["fab_id"]), reverse=True)
            jobs_by_stage[stage_name] = exceeded
            counts_by_stage[stage_name] = len(exceeded)

        threshold_analysis = {
            "threshold_days": threshold_days,
            "counts_by_stage": counts_by_stage,
            "jobs_by_stage": jobs_by_stage,
        }

    return success_response(
        {
            "title": "Turnaround Times Report",
            "year": report_year,
            "month": report_month,
            "period": {
                "from_date": (from_date or (start_dt.date() if start_dt else None)).isoformat() if (from_date or start_dt) else None,
                "to_date": (to_date or (end_dt.date() if end_dt else None)).isoformat() if (to_date or end_dt) else None,
            },
            "filters": {
                "fab_id": fab_id,
                "job_number": job_number,
                "fab_type": fab_type,
            },
            "summary": summary,
            "stage_averages": stage_averages,
            "average_draft_days": stage_averages.get("draft"),
            "threshold_analysis": threshold_analysis,
            "rows": rows,
        },
        "Owner turnaround times report generated",
    )


@router.get("/reports/owner/service-level", response_model=SuccessResponse[dict])
async def get_owner_service_level_report(
    start_date: Optional[date] = Query(None, description="Inclusive start date filter"),
    end_date: Optional[date] = Query(None, description="Inclusive end date filter"),
    date_basis: str = Query("completed", pattern="^(created|scheduled|completed)$"),
    sla_days: int = Query(14, ge=1, le=365, description="SLA threshold in days"),
    stage: Optional[str] = Query(None, description="Optional stage filter (e.g. Drafting, SCT, CNC)"),
    status: Optional[str] = Query(None, description="Optional status filter (on_track/at_risk/over_sla or green/yellow/red)"),
    sort_by: str = Query("stage", pattern="^(fab_id|job_number|stage)$", description="Sort fab_status_rows by fab_id, job_number, or stage"),
    sort_order: str = Query("asc", pattern="^(asc|desc)$", description="Sort direction for fab_status_rows"),
    breach_limit: int = Query(500, ge=1, le=5000),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Service-level report with KPI widgets, stage heat map, and Fab-level bottleneck rows."""
    start_dt, end_dt = _range_bounds(start_date, end_date)

    templating_exists = (
        select(Templating.id)
        .where(Templating.fab_id == Fab.id)
        .limit(1)
        .exists()
    )
    active_fab_filter = or_(
        Fab.template_needed.is_(False),
        and_(Fab.template_needed.is_(True), templating_exists),
    )
    non_resurface_filter = func.upper(func.trim(func.coalesce(Fab.fab_type, ""))) != "RESURFACE"
    normalized_stage_value = func.replace(
        func.replace(
            func.replace(func.lower(func.coalesce(Fab.current_stage, "")), " ", ""),
            "_",
            "",
        ),
        "-",
        "",
    )
    not_cutlist_stage_filter = ~normalized_stage_value.like("%cutlist%")
    eligible_fab_filter = and_(
        active_fab_filter,
        non_resurface_filter,
        not_cutlist_stage_filter,
    )

    # ── Load SLA settings from DB ───────────────────────────────────────────
    # Build a two-level lookup:  sla_map[normalized_fab_type][stage_name] = {target, at_risk}
    sla_rows = (
        await db.execute(select(ServiceLevelSetting).where(ServiceLevelSetting.is_applicable.is_(True)))
    ).scalars().all()
    sla_map: dict[str, dict[str, dict]] = {}
    for sla_row in sla_rows:
        ft = sla_row.fab_type.strip().upper()
        if ft not in sla_map:
            sla_map[ft] = {}
        sla_map[ft][sla_row.stage_name] = {
            "target_days": sla_row.target_days,
            "at_risk_days": sla_row.at_risk_days,
        }

    def _sla_for(fab_type_raw: Optional[str], stage_name: str) -> dict:
        """Return {target_days, at_risk_days} from DB, falling back to DEFAULT then hardcoded."""
        _hardcoded = {
            "Pre-Draft Review": {"target_days": 2.0, "at_risk_days": 1.0},
            "Drafting": {"target_days": 3.0, "at_risk_days": 1.0},
            "SCT": {"target_days": 3.0, "at_risk_days": 1.0},
            "SlabSmith": {"target_days": 2.0, "at_risk_days": 1.0},
            "Final Programming": {"target_days": 2.0, "at_risk_days": 0.0},
            "CNC": {"target_days": 1.0, "at_risk_days": 0.0},
            "Revisions": {"target_days": 2.0, "at_risk_days": 1.0},
        }
        normalized_ft = (fab_type_raw or "").strip().upper()
        # 1. Exact fab-type match
        if normalized_ft in sla_map and stage_name in sla_map[normalized_ft]:
            return sla_map[normalized_ft][stage_name]
        # 2. DEFAULT row
        if "DEFAULT" in sla_map and stage_name in sla_map["DEFAULT"]:
            return sla_map["DEFAULT"][stage_name]
        # 3. Hardcoded fallback
        return _hardcoded.get(stage_name, {"target_days": float(sla_days), "at_risk_days": 0.0})
    # ────────────────────────────────────────────────────────────────────────

    def _revision_subtype(value: Optional[str]) -> Optional[str]:
        text = (value or "").strip().lower()
        if not text:
            return None
        if "cad" in text:
            return "cad"
        if "template" in text:
            return "template"
        if "sales" in text:
            return "sales"
        if "client" in text:
            return "client"
        return None

    def _target_days_for(stage_name: str, revision_type: Optional[str], fab_type_raw: Optional[str] = None) -> float:
        sla = _sla_for(fab_type_raw, stage_name)
        return float(sla["target_days"])

    def _risk_color(age_days: float, target_days: float, at_risk_days: float = 0.0) -> str:
        red_threshold = target_days + at_risk_days
        if age_days > red_threshold:
            return "red"
        if age_days > target_days:
            return "yellow"
        return "green"

    def _normalize_stage(stage_value: Optional[str]) -> str:
        stage = (stage_value or "").strip().lower()
        if "pre_draft" in stage or "predraft" in stage:
            return "Pre-Draft Review"
        if "revision" in stage:
            return "Revisions"
        if stage in {"sct", "sales_ct"} or "sales_ct" in stage or "salesct" in stage:
            return "SCT"
        if "slabsmith" in stage or "slab_smith" in stage:
            return "SlabSmith"
        if "final_program" in stage:
            return "Final Programming"
        if "cnc" in stage:
            return "CNC"
        if "draft" in stage:
            return "Drafting"
        return "Other"

    def _normalize_filter_text(value: Optional[str]) -> str:
        return " ".join((value or "").replace("_", " ").replace("-", " ").strip().lower().split())

    stage_filter_value = _normalize_filter_text(stage) if stage else None
    status_filter_value = _normalize_filter_text(status) if status else None

    completion_query = (
        select(
            Fab.id.label("fab_id"),
            BusinessJob.job_number,
            Fab.created_at,
            Fab.template_completed_date,
            InstallCompletion.install_date,
            InstallCompletion.completion_date,
            Fab.shop_date_schedule,
            Fab.shop_est_completion_date,
            Fab.installation_date,
            InstallCompletion.installer_id,
            InstallCompletion.total_sqft_installed,
        )
        .select_from(InstallCompletion)
        .join(Fab, Fab.id == InstallCompletion.fab_id)
        .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
        .where(eligible_fab_filter)
    )

    if date_basis == "created":
        if start_dt is not None:
            completion_query = completion_query.where(Fab.created_at >= start_dt)
        if end_dt is not None:
            completion_query = completion_query.where(Fab.created_at <= end_dt)
    elif date_basis == "scheduled":
        scheduled_basis = func.coalesce(
            InstallCompletion.install_date,
            Fab.installation_date,
            Fab.shop_est_completion_date,
            Fab.shop_date_schedule,
        )
        if start_dt is not None:
            completion_query = completion_query.where(scheduled_basis >= start_dt)
        if end_dt is not None:
            completion_query = completion_query.where(scheduled_basis <= end_dt)
    else:
        if start_dt is not None:
            completion_query = completion_query.where(InstallCompletion.completion_date >= start_dt)
        if end_dt is not None:
            completion_query = completion_query.where(InstallCompletion.completion_date <= end_dt)

    completion_rows = (await db.execute(completion_query)).all()

    total_completed = len(completion_rows)
    on_time_eligible = 0
    on_time_count = 0
    cycle_days: list[int] = []
    cycle_rows = []

    for (
        fab_id,
        job_number,
        created_at,
        template_completed_at,
        install_date,
        completion_date,
        shop_date_schedule,
        shop_est_completion_date,
        installation_date,
        installer_id,
        total_sqft_installed,
    ) in completion_rows:
        due_date = install_date or installation_date or shop_est_completion_date or shop_date_schedule
        if due_date is not None:
            on_time_eligible += 1
            if completion_date is not None and completion_date <= due_date:
                on_time_count += 1

        cycle_day = _days_between(template_completed_at, completion_date)
        if cycle_day is not None:
            cycle_days.append(cycle_day)

        cycle_rows.append(
            {
                "fab_id": fab_id,
                "job_number": job_number,
                "installer_id": installer_id,
                "created_at": created_at.isoformat() if created_at else None,
                "template_completed_at": template_completed_at.isoformat() if template_completed_at else None,
                "due_date": due_date.isoformat() if due_date else None,
                "completion_date": completion_date.isoformat() if completion_date else None,
                "cycle_days": cycle_day,
                "is_on_time": bool(due_date is not None and completion_date is not None and completion_date <= due_date),
                "total_sqft_installed": round(_to_float(total_sqft_installed), 2),
            }
        )

    cycle_stats = {
        "count": len(cycle_days),
        "min_days": min(cycle_days) if cycle_days else None,
        "max_days": max(cycle_days) if cycle_days else None,
        "avg_days": round(sum(cycle_days) / len(cycle_days), 2) if cycle_days else None,
    }

    completed_subquery = (
        select(
            InstallCompletion.fab_id.label("fab_id"),
            func.max(InstallCompletion.completion_date).label("completion_date"),
        )
        .group_by(InstallCompletion.fab_id)
        .subquery()
    )

    schedule_subquery = (
        select(
            InstallScheduling.fab_id.label("fab_id"),
            func.max(InstallScheduling.scheduled_install_date).label("scheduled_install_date"),
        )
        .group_by(InstallScheduling.fab_id)
        .subquery()
    )

    revision_subquery = (
        select(
            Revision.fab_id.label("fab_id"),
            Revision.revision_type,
            Revision.assigned_to,
            Revision.actual_start_date,
            Revision.scheduled_start_date,
            Revision.created_at,
        )
        .subquery()
    )

    backlog_query = (
        select(
            Fab.id,
            BusinessJob.job_number,
            BusinessJob.name,
            Account.name.label("account_name"),
            BusinessJob.priority,
            Fab.current_stage,
            Fab.fab_type,
            Fab.input_area,
            Fab.total_sqft,
            Fab.no_of_pieces,
            StoneType.name.label("stone_type_name"),
            StoneColor.name.label("stone_color_name"),
            StoneThickness.thickness.label("stone_thickness_value"),
            Edge.name.label("edge_name"),
            Fab.created_at,
            Fab.updated_at,
            Fab.template_completed_date,
            Fab.predraft_completed_date,
            Fab.draft_completed_date,
            Fab.slabsmith_completed_date,
            Fab.sales_ct_completed_date,
            Fab.sct_completed_date,
            Fab.final_programming_completed_date,
            Fab.installation_date,
            Fab.shop_est_completion_date,
            Fab.shop_date_schedule,
            Fab.status_id,
            Fab.drafter_id,
            Fab.sales_person_id,
            schedule_subquery.c.scheduled_install_date,
            completed_subquery.c.completion_date,
            revision_subquery.c.revision_type,
            revision_subquery.c.assigned_to,
            revision_subquery.c.actual_start_date,
            revision_subquery.c.scheduled_start_date,
            revision_subquery.c.created_at,
        )
        .select_from(Fab)
        .join(BusinessJob, BusinessJob.id == Fab.job_id, isouter=True)
        .join(Account, Account.id == BusinessJob.account_id, isouter=True)
        .join(StoneType, StoneType.id == Fab.stone_type_id, isouter=True)
        .join(StoneColor, StoneColor.id == Fab.stone_color_id, isouter=True)
        .join(StoneThickness, StoneThickness.id == Fab.stone_thickness_id, isouter=True)
        .join(Edge, Edge.id == Fab.edge_id, isouter=True)
        .join(schedule_subquery, schedule_subquery.c.fab_id == Fab.id, isouter=True)
        .join(completed_subquery, completed_subquery.c.fab_id == Fab.id, isouter=True)
        .join(revision_subquery, revision_subquery.c.fab_id == Fab.id, isouter=True)
        .where(completed_subquery.c.completion_date.is_(None), Fab.status_id == 1, eligible_fab_filter)
    )

    if date_basis == "created":
        if start_dt is not None:
            backlog_query = backlog_query.where(Fab.created_at >= start_dt)
        if end_dt is not None:
            backlog_query = backlog_query.where(Fab.created_at <= end_dt)
    elif date_basis == "scheduled":
        scheduled_backlog_basis = func.coalesce(
            schedule_subquery.c.scheduled_install_date,
            Fab.installation_date,
            Fab.shop_est_completion_date,
            Fab.shop_date_schedule,
        )
        if start_dt is not None:
            backlog_query = backlog_query.where(scheduled_backlog_basis >= start_dt)
        if end_dt is not None:
            backlog_query = backlog_query.where(scheduled_backlog_basis <= end_dt)

    backlog_rows = (await db.execute(backlog_query)).all()

    bucket_totals = {
        "0_7_days": 0,
        "8_14_days": 0,
        "15_30_days": 0,
        "31_plus_days": 0,
    }

    stage_display_names = [
        "Pre-Draft Review",
        "Drafting",
        "Revisions",
        "SCT",
        "SlabSmith",
        "Final Programming",
        "CNC",
    ]
    stage_rollup: dict[str, dict] = {
        stage_name: {
            "green": 0,
            "yellow": 0,
            "red": 0,
            "total_wip": 0,
            "total_age_days": 0,
        }
        for stage_name in stage_display_names
    }

    total_green = 0
    total_yellow = 0
    total_red = 0
    oldest_open_job = None
    breach_details = []
    fab_status_rows = []

    def _first_name(names: dict[int, str], user_id: Optional[int]) -> Optional[str]:
        if user_id is None:
            return None
        return names.get(user_id)

    backlog_records = backlog_rows
    user_ids: set[int] = set()
    for row in backlog_records:
        # Keep tuple-index mapping aligned to backlog_query select order.
        drafter_id = row[27]
        sales_person_id = row[28]
        revision_assigned_to = row[32]
        if isinstance(drafter_id, int):
            user_ids.add(drafter_id)
        if isinstance(sales_person_id, int):
            user_ids.add(sales_person_id)
        if isinstance(revision_assigned_to, int):
            user_ids.add(revision_assigned_to)

    user_names: dict[int, str] = {}
    if user_ids:
        user_rows = (
            await db.execute(
                select(User.id, User.first_name, User.last_name).where(User.id.in_(list(user_ids)))
            )
        ).all()
        user_names = {
            user_id: (f"{(first_name or '').strip()} {(last_name or '').strip()}".strip() or f"User {user_id}")
            for user_id, first_name, last_name in user_rows
        }

    now_dt = datetime.now()
    for (
        fab_id,
        job_number,
        job_name,
        account_name,
        job_priority,
        current_stage,
        fab_type,
        input_area,
        total_sqft,
        no_of_pieces,
        stone_type_name,
        stone_color_name,
        stone_thickness_value,
        edge_name,
        created_at,
        updated_at,
        template_completed_at,
        predraft_completed_at,
        draft_completed_at,
        slabsmith_completed_at,
        sales_ct_completed_at,
        sct_completed_at,
        final_programming_completed_at,
        installation_date,
        shop_est_completion_date,
        shop_date_schedule,
        _status_id,
        drafter_id,
        sales_person_id,
        scheduled_install_date,
        _completion_date,
        revision_type,
        revision_assigned_to,
        revision_actual_start,
        revision_scheduled_start,
        revision_created_at,
    ) in backlog_rows:
        due_date = scheduled_install_date or installation_date or shop_est_completion_date or shop_date_schedule
        cycle_anchor = template_completed_at or created_at
        stage_name = _normalize_stage(current_stage)

        if stage_name == "Pre-Draft Review":
            stage_anchor = template_completed_at or created_at
        elif stage_name == "Drafting":
            stage_anchor = predraft_completed_at or template_completed_at or created_at
        elif stage_name == "Revisions":
            stage_anchor = revision_actual_start or revision_scheduled_start or revision_created_at or updated_at or created_at
        elif stage_name == "SCT":
            stage_anchor = draft_completed_at or updated_at or created_at
        elif stage_name == "SlabSmith":
            stage_anchor = sales_ct_completed_at or sct_completed_at or updated_at or created_at
        elif stage_name == "Final Programming":
            stage_anchor = slabsmith_completed_at or sales_ct_completed_at or sct_completed_at or updated_at or created_at
        elif stage_name == "CNC":
            stage_anchor = final_programming_completed_at or updated_at or created_at
        else:
            stage_anchor = updated_at or created_at

        if stage_anchor is None:
            continue

        days_since_template = _days_between(template_completed_at, now_dt)
        days_in_stage = max((now_dt.date() - stage_anchor.date()).days, 0)
        sla = _sla_for(fab_type, stage_name)
        target_days = float(sla["target_days"])
        at_risk_days = float(sla["at_risk_days"])
        risk_color = _risk_color(float(days_in_stage), target_days, at_risk_days)
        status_label = "On Track" if risk_color == "green" else "At Risk" if risk_color == "yellow" else "Over SLA"
        effective_stage_name = stage_name if stage_name != "Other" else (current_stage or "Other")

        if stage_filter_value:
            if _normalize_filter_text(effective_stage_name) != stage_filter_value:
                continue

        if status_filter_value:
            status_aliases = {
                "on track": "on track",
                "ontrack": "on track",
                "at risk": "at risk",
                "atrisk": "at risk",
                "over sla": "over sla",
                "oversla": "over sla",
                "green": "green",
                "yellow": "yellow",
                "red": "red",
            }
            row_status_key = status_aliases.get(_normalize_filter_text(status_label), _normalize_filter_text(status_label))
            row_risk_key = status_aliases.get(_normalize_filter_text(risk_color), _normalize_filter_text(risk_color))
            filter_key = status_aliases.get(status_filter_value, status_filter_value)
            if filter_key not in {row_status_key, row_risk_key}:
                continue

        age_days = days_in_stage
        if age_days <= 7:
            bucket_totals["0_7_days"] += 1
        elif age_days <= 14:
            bucket_totals["8_14_days"] += 1
        elif age_days <= 30:
            bucket_totals["15_30_days"] += 1
        else:
            bucket_totals["31_plus_days"] += 1

        if risk_color == "green":
            total_green += 1
        elif risk_color == "yellow":
            total_yellow += 1
        else:
            total_red += 1

        if stage_name in stage_rollup:
            stage_rollup[stage_name][risk_color] += 1
            stage_rollup[stage_name]["total_wip"] += 1
            stage_rollup[stage_name]["total_age_days"] += age_days

        if oldest_open_job is None or age_days > oldest_open_job["age_days"]:
            oldest_open_job = {
                "fab_id": fab_id,
                "job_number": job_number,
                "current_stage": stage_name if stage_name != "Other" else current_stage,
                "age_days": age_days,
                "due_date": due_date.isoformat() if due_date else None,
                "stage_anchor": stage_anchor.isoformat() if stage_anchor else None,
            }

        if stage_name == "Revisions":
            assigned_user_id = revision_assigned_to or drafter_id
        elif stage_name in {"Drafting", "Pre-Draft Review", "Final Programming", "CNC", "SlabSmith"}:
            assigned_user_id = drafter_id
        elif stage_name == "SCT":
            assigned_user_id = sales_person_id
        else:
            assigned_user_id = drafter_id or sales_person_id

        priority_flag = str(job_priority or "").strip() or ("High" if risk_color == "red" else "Medium")

        fab_status_rows.append(
            {
                "fab_type": fab_type,
                "fab_id": fab_id,
                "job_number": job_number,
                "job_name": job_name,
                "account_name": account_name,
                "input_area": input_area,
                "stone_type_name": stone_type_name,
                "stone_color_name": stone_color_name,
                "stone_thickness_value": stone_thickness_value,
                "edge_name": edge_name,
                "fab_info": (f"{job_name or ''} | {round(_to_float(total_sqft), 2)} sqft | {int(_to_float(no_of_pieces))} pcs").strip(" |"),
                "current_stage": effective_stage_name,
                "days_since_template": days_since_template,
                "days_in_stage": days_in_stage,
                "revision_type": revision_type,
                "assigned_user": _first_name(user_names, assigned_user_id),
                "status": status_label,
                "priority_flag": priority_flag,
                "risk_color": risk_color,
                "stage_target_days": target_days,
            }
        )

        if risk_color == "red":
            breach_details.append(
                {
                    "fab_id": fab_id,
                    "job_number": job_number,
                    "job_name": job_name,
                    "account_name": account_name,
                    "input_area": input_area,
                    "stone_type_name": stone_type_name,
                    "stone_color_name": stone_color_name,
                    "stone_thickness_value": stone_thickness_value,
                    "edge_name": edge_name,
                    "current_stage": effective_stage_name,
                    "age_days": days_in_stage,
                    "sla_days": target_days,
                    "days_over_sla": round(days_in_stage - target_days, 2),
                    "stage_anchor": stage_anchor.isoformat() if stage_anchor else None,
                    "due_date": due_date.isoformat() if due_date else None,
                }
            )

    breach_details.sort(key=lambda row: row["days_over_sla"], reverse=True)
    total_breach_count = len(breach_details)
    breach_details = breach_details[:breach_limit]

    aging_backlog = [
        {"bucket": "0-7 days", "count": bucket_totals["0_7_days"]},
        {"bucket": "8-14 days", "count": bucket_totals["8_14_days"]},
        {"bucket": "15-30 days", "count": bucket_totals["15_30_days"]},
        {"bucket": "31+ days", "count": bucket_totals["31_plus_days"]},
    ]

    stage_bottleneck_heat_map = []
    for stage_name in stage_display_names:
        row = stage_rollup[stage_name]
        total_wip = int(row["total_wip"])
        avg_days = round((row["total_age_days"] / total_wip), 2) if total_wip else 0.0
        default_sla = _sla_for(None, stage_name)
        target_days = float(default_sla["target_days"])
        stage_bottleneck_heat_map.append(
            {
                "stage": stage_name,
                "target_days": target_days,
                "green": int(row["green"]),
                "yellow": int(row["yellow"]),
                "red": int(row["red"]),
                "total_wip": total_wip,
                "avg_days": avg_days,
                "sla_breach_percent": round((row["red"] / total_wip) * 100, 2) if total_wip else 0.0,
            }
        )

    on_time_percent = round((on_time_count / on_time_eligible) * 100, 2) if on_time_eligible else 0.0
    total_active_fab_ids = len(backlog_rows)

    widgets = {
        "total_active_fab_ids": total_active_fab_ids,
        "on_track_green": total_green,
        "at_risk_yellow": total_yellow,
        "overdue_red": total_red,
        "avg_cycle_time_days": cycle_stats["avg_days"],
        "oldest_open_job": oldest_open_job,
    }

    sort_by_normalized = (sort_by or "stage").strip().lower()
    sort_order_normalized = (sort_order or "asc").strip().lower()
    reverse_sort = sort_order_normalized == "desc"

    if sort_by_normalized == "fab_id":
        fab_status_rows.sort(key=lambda row: int(_to_float(row.get("fab_id", 0))), reverse=reverse_sort)
    elif sort_by_normalized == "job_number":
        fab_status_rows.sort(key=lambda row: (row.get("job_number") or "").lower(), reverse=reverse_sort)
    else:
        fab_status_rows.sort(key=lambda row: (row.get("current_stage") or "").lower(), reverse=reverse_sort)

    return success_response(
        {
            "title": "Service Level Report",
            "period": {
                "start_date": start_date.isoformat() if start_date else None,
                "end_date": end_date.isoformat() if end_date else None,
                "date_basis": date_basis,
                "sla_days": sla_days,
            },
            "sort": {
                "sort_by": sort_by_normalized,
                "sort_order": sort_order_normalized,
            },
            "filters": {
                "stage": stage,
                "status": status,
            },
            "widgets": widgets,
            "summary": {
                "total_completed": total_completed,
                "on_time_eligible": on_time_eligible,
                "on_time_count": on_time_count,
                "on_time_percent": on_time_percent,
                "open_backlog_count": sum(bucket_totals.values()),
                "sla_breach_count": total_breach_count,
                "green_count": total_green,
                "yellow_count": total_yellow,
                "red_count": total_red,
                "cycle_time_avg_days": cycle_stats["avg_days"],
                "cycle_time_min_days": cycle_stats["min_days"],
                "cycle_time_max_days": cycle_stats["max_days"],
            },
            "stage_bottleneck_heat_map": stage_bottleneck_heat_map,
            "sla_rules": {
                "green": "At or below target days for the FAB type and stage",
                "yellow": "Beyond target days but within at-risk window",
                "red": "Exceeds target + at-risk days threshold",
                "settings": [
                    {
                        "id": r.id,
                        "fab_type": r.fab_type,
                        "stage_name": r.stage_name,
                        "target_days": r.target_days,
                        "at_risk_days": r.at_risk_days,
                        "is_applicable": r.is_applicable,
                    }
                    for r in sla_rows
                ],
            },
            "fab_status_rows": fab_status_rows,
            "cycle_time_stats": [cycle_stats],
            "aging_backlog": aging_backlog,
            "breach_details": breach_details,
            "completed_rows": cycle_rows,
        },
        "Owner service level report generated",
    )


@router.get("/reports/owner/installer-rates", response_model=SuccessResponse[list[dict]])
async def get_installer_rates(
    installer_id: Optional[int] = Query(None, gt=0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(
        InstallerRateHistory.id,
        InstallerRateHistory.installer_id,
        InstallerRateHistory.hourly_rate,
        InstallerRateHistory.effective_from,
        InstallerRateHistory.effective_to,
        InstallerRateHistory.is_active,
        User.first_name,
        User.last_name,
    ).join(User, User.id == InstallerRateHistory.installer_id, isouter=True)

    if installer_id is not None:
        query = query.where(InstallerRateHistory.installer_id == installer_id)

    query = query.order_by(InstallerRateHistory.installer_id, InstallerRateHistory.effective_from.desc())
    rows = (await db.execute(query)).all()

    data = [
        {
            "id": row[0],
            "installer_id": row[1],
            "installer_name": (f"{(row[6] or '').strip()} {(row[7] or '').strip()}".strip() or f"User {row[1]}"),
            "hourly_rate": round(_to_float(row[2]), 2),
            "effective_from": row[3].isoformat() if row[3] else None,
            "effective_to": row[4].isoformat() if row[4] else None,
            "is_active": bool(row[5]),
        }
        for row in rows
    ]

    return success_response(data, "Installer rates retrieved")


@router.post("/reports/owner/installer-rates", response_model=SuccessResponse[dict])
async def create_installer_rate(
    payload: InstallerRateUpsert,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    installer_exists = (
        await db.execute(select(User.id).where(User.id == payload.installer_id))
    ).scalar_one_or_none()
    if not installer_exists:
        return success_response(None, "Installer not found", status_code=404)

    if payload.effective_to and payload.effective_to < payload.effective_from:
        return success_response(None, "effective_to cannot be before effective_from", status_code=400)

    # Close previously active open-ended ranges to avoid overlapping active rates.
    open_rates = (
        await db.execute(
            select(InstallerRateHistory)
            .where(
                InstallerRateHistory.installer_id == payload.installer_id,
                InstallerRateHistory.is_active.is_(True),
                InstallerRateHistory.effective_to.is_(None),
                InstallerRateHistory.effective_from <= payload.effective_from,
            )
            .order_by(InstallerRateHistory.effective_from.desc())
        )
    ).scalars().all()

    for rate in open_rates:
        rate.effective_to = payload.effective_from
        rate.updated_at = datetime.now()
        rate.updated_by = current_user.id

    new_rate = InstallerRateHistory(
        installer_id=payload.installer_id,
        hourly_rate=payload.hourly_rate,
        effective_from=payload.effective_from,
        effective_to=payload.effective_to,
        is_active=payload.is_active,
        created_by=current_user.id,
    )
    db.add(new_rate)
    await db.commit()
    await db.refresh(new_rate)

    return success_response(
        {
            "id": new_rate.id,
            "installer_id": new_rate.installer_id,
            "hourly_rate": new_rate.hourly_rate,
            "effective_from": new_rate.effective_from.isoformat(),
            "effective_to": new_rate.effective_to.isoformat() if new_rate.effective_to else None,
            "is_active": new_rate.is_active,
        },
        "Installer rate created",
    )


@router.get("/reports/owner/service-level-settings", response_model=SuccessResponse[list[dict]])
async def get_service_level_settings(
    fab_type: Optional[str] = Query(None, description="Filter by fab type"),
    stage_name: Optional[str] = Query(None, description="Filter by stage name"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return all SLA target-day settings, optionally filtered by fab_type or stage_name."""
    _ = current_user
    query = select(ServiceLevelSetting).order_by(
        ServiceLevelSetting.fab_type.asc(), ServiceLevelSetting.stage_name.asc()
    )
    if fab_type is not None:
        query = query.where(ServiceLevelSetting.fab_type.ilike(fab_type.strip()))
    if stage_name is not None:
        query = query.where(ServiceLevelSetting.stage_name.ilike(stage_name.strip()))
    rows = (await db.execute(query)).scalars().all()
    return success_response(
        [
            {
                "id": row.id,
                "fab_type": row.fab_type,
                "stage_name": row.stage_name,
                "target_days": row.target_days,
                "at_risk_days": row.at_risk_days,
                "is_applicable": row.is_applicable,
                "updated_at": row.updated_at.isoformat() if row.updated_at else None,
                "updated_by": row.updated_by,
            }
            for row in rows
        ],
        "Service level settings retrieved",
    )


@router.post("/reports/owner/service-level-settings", response_model=SuccessResponse[dict])
async def create_service_level_setting(
    payload: ServiceLevelSettingCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create a new SLA target-day setting row."""
    # Prevent duplicates on (fab_type, stage_name)
    existing = (
        await db.execute(
            select(ServiceLevelSetting).where(
                ServiceLevelSetting.fab_type == payload.fab_type,
                ServiceLevelSetting.stage_name == payload.stage_name,
            )
        )
    ).scalar_one_or_none()
    if existing:
        return error_response(
            f"Setting for fab_type='{payload.fab_type}' stage='{payload.stage_name}' already exists (id={existing.id}). Use PATCH to update.",
            400,
        )

    row = ServiceLevelSetting(
        fab_type=payload.fab_type,
        stage_name=payload.stage_name,
        target_days=payload.target_days,
        at_risk_days=payload.at_risk_days,
        is_applicable=payload.is_applicable,
        updated_by=current_user.id,
        updated_at=datetime.now(),
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return success_response(
        {
            "id": row.id,
            "fab_type": row.fab_type,
            "stage_name": row.stage_name,
            "target_days": row.target_days,
            "at_risk_days": row.at_risk_days,
            "is_applicable": row.is_applicable,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
            "updated_by": row.updated_by,
        },
        "Service level setting created",
    )


@router.patch("/reports/owner/service-level-settings/{setting_id}", response_model=SuccessResponse[dict])
async def update_service_level_setting(
    setting_id: int,
    payload: ServiceLevelSettingUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update target_days, at_risk_days, or is_applicable for a single SLA setting row."""
    if payload.target_days is None and payload.at_risk_days is None and payload.is_applicable is None:
        return error_response("At least one field is required", 400)

    row = (await db.execute(select(ServiceLevelSetting).where(ServiceLevelSetting.id == setting_id))).scalar_one_or_none()
    if not row:
        return error_response("Service level setting not found", 404)

    if payload.target_days is not None:
        row.target_days = payload.target_days
    if payload.at_risk_days is not None:
        row.at_risk_days = payload.at_risk_days
    if payload.is_applicable is not None:
        row.is_applicable = payload.is_applicable

    row.updated_at = datetime.now()
    row.updated_by = current_user.id
    await db.commit()
    await db.refresh(row)

    return success_response(
        {
            "id": row.id,
            "fab_type": row.fab_type,
            "stage_name": row.stage_name,
            "target_days": row.target_days,
            "at_risk_days": row.at_risk_days,
            "is_applicable": row.is_applicable,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
            "updated_by": row.updated_by,
        },
        "Service level setting updated",
    )


@router.delete("/reports/owner/service-level-settings/{setting_id}", response_model=SuccessResponse[dict])
async def delete_service_level_setting(
    setting_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a single SLA setting row."""
    row = (await db.execute(select(ServiceLevelSetting).where(ServiceLevelSetting.id == setting_id))).scalar_one_or_none()
    if not row:
        return error_response("Service level setting not found", 404)
    await db.delete(row)
    await db.commit()
    return success_response({"id": setting_id}, "Service level setting deleted")


@router.get("/reports/owner/management-packet", response_model=SuccessResponse[dict])
async def get_owner_management_packet(
    start_date: Optional[date] = Query(None, description="Inclusive start date filter"),
    end_date: Optional[date] = Query(None, description="Inclusive end date filter"),
    weeks: int = Query(12, ge=4, le=52, description="How many trailing weeks to include"),
    top_n: int = Query(10, ge=1, le=50, description="Top N for hotspot/installer blocks"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    overview = _unwrap_success_data(
        await get_owner_overview_report(start_date=start_date, end_date=end_date, db=db, current_user=current_user)
    )
    redo = _unwrap_success_data(
        await get_owner_redo_analysis_report(
            start_date=start_date,
            end_date=end_date,
            top_n=top_n,
            db=db,
            current_user=current_user,
        )
    )
    shop_status = _unwrap_success_data(
        await get_owner_shop_status_report(start_date=start_date, end_date=end_date, db=db, current_user=current_user)
    )
    install_perf = _unwrap_success_data(
        await get_owner_install_performance_report(
            start_date=start_date,
            end_date=end_date,
            top_n=top_n,
            db=db,
            current_user=current_user,
        )
    )
    weekly = _unwrap_success_data(
        await get_owner_weekly_trends_report(weeks=weeks, db=db, current_user=current_user)
    )

    return success_response(
        {
            "generated_at": datetime.now().isoformat(),
            "period": {
                "start_date": start_date.isoformat() if start_date else None,
                "end_date": end_date.isoformat() if end_date else None,
                "weeks": weeks,
            },
            "overview": overview,
            "redo_analysis": redo,
            "shop_status": shop_status,
            "install_performance": install_perf,
            "weekly_trends": weekly,
        },
        "Owner management packet generated",
    )


@router.post("/reports/owner/end-of-month-status/send-test", response_model=SuccessResponse[dict])
async def send_end_of_month_status_report_test(
    year: Optional[int] = Query(None, ge=2000, le=2100, description="Target year; defaults to previous month year"),
    month: Optional[int] = Query(None, ge=1, le=12, description="Target month; defaults to previous month"),
    email: Optional[str] = Query(None, description="Optional recipient email(s), comma-separated"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    del db  # Route keeps consistent dependency pattern; service manages its own DB session.

    if (year is None) != (month is None):
        return success_response(None, "Provide both year and month, or neither", status_code=400)

    if year is None and month is None:
        previous_month_date = date.today().replace(day=1) - timedelta(days=1)
        year = previous_month_date.year
        month = previous_month_date.month

    send_result = await send_monthly_end_of_month_status_report(
        year=year,
        month=month,
        recipients_override=email,
    )

    if not send_result.get("sent"):
        return success_response(send_result, "Unable to send end-of-month status report", status_code=400)

    return success_response(send_result, "End-of-month status report sent successfully")


@router.get("/reports/owner/export/{report_key}")
async def export_owner_report(
    report_key: str,
    export_format: str = Query("csv", pattern="^(csv|xlsx|json)$"),
    layout: str = Query("default", pattern="^(default|client)$"),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    search: Optional[str] = Query(None),
    fab_type: Optional[str] = Query(None),
    sales_person_id: Optional[int] = Query(None, gt=0),
    date_basis: str = Query("completed", pattern="^(created|scheduled|completed)$"),
    sla_days: int = Query(14, ge=1, le=365),
    weeks: int = Query(12, ge=4, le=52),
    top_n: int = Query(10, ge=1, le=50),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    month: Optional[int] = Query(None, ge=1, le=12),
    total_employees: int = Query(40, ge=0),
    overhead_per_week: float = Query(38512.69, ge=0),
    week_ending_weekday: int = Query(4, ge=0, le=6),
    payroll_overrides_json: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    key = report_key.strip().lower()

    if key == "overview":
        data = _unwrap_success_data(
            await get_owner_overview_report(start_date=start_date, end_date=end_date, db=db, current_user=current_user)
        )
    elif key == "redo-analysis":
        data = _unwrap_success_data(
            await get_owner_redo_analysis_report(start_date=start_date, end_date=end_date, top_n=top_n, db=db, current_user=current_user)
        )
    elif key == "shop-status":
        data = _unwrap_success_data(
            await get_owner_shop_status_report(start_date=start_date, end_date=end_date, db=db, current_user=current_user)
        )
    elif key == "install-performance":
        data = _unwrap_success_data(
            await get_owner_install_performance_report(start_date=start_date, end_date=end_date, top_n=top_n, db=db, current_user=current_user)
        )
    elif key == "installation-template":
        data = _unwrap_success_data(
            await get_owner_installation_template_report(start_date=start_date, end_date=end_date, db=db, current_user=current_user)
        )
    elif key == "installation-template-dashboard":
        data = _unwrap_success_data(
            await get_owner_installation_template_dashboard_report(
                from_date=from_date or start_date,
                to_date=to_date or end_date,
                search=search,
                fab_type=fab_type,
                sales_person_id=sales_person_id,
                db=db,
                current_user=current_user,
            )
        )
    elif key == "monthly-install-completion":
        if year is None or month is None:
            return success_response(None, "year and month are required for monthly-install-completion", status_code=400)
        data = _unwrap_success_data(
            await get_owner_monthly_install_completion_report(year=year, month=month, db=db, current_user=current_user)
        )
    elif key == "daily-install-completion":
        data = _unwrap_success_data(
            await get_daily_install_completion_report(start_date=start_date, end_date=end_date, db=db, current_user=current_user)
        )
    elif key == "monthly-cut-completion":
        if year is None or month is None:
            return success_response(None, "year and month are required for monthly-cut-completion", status_code=400)
        data = _unwrap_success_data(
            await get_owner_monthly_cut_completion_report(year=year, month=month, db=db, current_user=current_user)
        )
    elif key == "turnaround-times":
        if year is None or month is None:
            return success_response(None, "year and month are required for turnaround-times", status_code=400)
        data = _unwrap_success_data(
            await get_owner_turnaround_times_report(year=year, month=month, db=db, current_user=current_user)
        )
    elif key == "service-level":
        data = _unwrap_success_data(
            await get_owner_service_level_report(
                start_date=start_date,
                end_date=end_date,
                date_basis=date_basis,
                sla_days=sla_days,
                db=db,
                current_user=current_user,
            )
        )
    elif key == "weekly-trends":
        data = _unwrap_success_data(
            await get_owner_weekly_trends_report(weeks=weeks, db=db, current_user=current_user)
        )
    elif key == "management-packet":
        data = _unwrap_success_data(
            await get_owner_management_packet(
                start_date=start_date,
                end_date=end_date,
                weeks=weeks,
                top_n=top_n,
                db=db,
                current_user=current_user,
            )
        )
    elif key == "weekly-fabrication-labor-cost":
        if year is None or month is None:
            return success_response(None, "year and month are required for weekly-fabrication-labor-cost", status_code=400)
        data = _unwrap_success_data(
            await get_owner_weekly_fabrication_labor_cost_report(
                year=year,
                month=month,
                total_employees=total_employees,
                overhead_per_week=overhead_per_week,
                week_ending_weekday=week_ending_weekday,
                payroll_overrides_json=payroll_overrides_json,
                db=db,
                current_user=current_user,
            )
        )
    elif key == "weekly-installer-labor-cost":
        if year is None or month is None:
            return success_response(None, "year and month are required for weekly-installer-labor-cost", status_code=400)
        data = _unwrap_success_data(
            await get_owner_weekly_installer_labor_cost_report(
                year=year,
                month=month,
                total_employees=total_employees,
                overhead_per_week=overhead_per_week,
                week_ending_weekday=week_ending_weekday,
                payroll_overrides_json=payroll_overrides_json,
                db=db,
                current_user=current_user,
            )
        )
    else:
        return success_response(None, f"Unsupported report_key '{report_key}'", status_code=400)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{key}_{stamp}.{export_format}"

    if export_format == "json":
        content = json.dumps(data, indent=2).encode("utf-8")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    if export_format == "csv":
        content = _csv_bytes(key, data, layout=layout)
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    content = _xlsx_bytes(key, data, layout=layout)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
